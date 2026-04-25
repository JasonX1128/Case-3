from __future__ import annotations

import argparse
import csv
import ctypes
import heapq
import json
import math
import multiprocessing as mp
import os
import re
import shutil
import signal
import subprocess
import sys
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor
from dataclasses import dataclass
from datetime import datetime, time as dt_time
from pathlib import Path
from time import perf_counter
from typing import Any, Callable

import gurobipy as gp
from gurobipy import GRB
from openpyxl import load_workbook

try:
    import numpy as np
except ImportError:  # pragma: no cover - optional acceleration
    np = None


DEPOT_NODE = 0
REDUCED_COST_TOL = 1e-6
INTEGRALITY_TOL = 1e-6
HEURISTIC_MASTER_TIME_LIMIT = 5.0
ARTIFICIAL_COVER_PENALTY = 1_000_000.0
INITIAL_ACTIVE_COVER_CANDIDATES_PER_STOP = 3
INITIAL_ACTIVE_CHEAPEST_ROUTES_PER_TYPE = 12
INITIAL_ACTIVE_RECENT_ROUTES_PER_TYPE = 12
ROOT_PRICING_COLUMNS_MIN = 30
ROOT_PRICING_COLUMNS_MULTIPLIER = 10
EXACT_PRICING_FALLBACK_MAX_ROUTES = 1
ROOT_EXACT_PRICING_FALLBACK_MAX_ROUTES = 16
HEURISTIC_PRICING_SINGLETON_SEED_LIMIT = 40
HEURISTIC_PRICING_BEAM_WIDTH = 8
ROUTE_BASED_PARTIAL_PRICING = True
ROOT_BOOTSTRAP_MAX_ITERATIONS = 200
ROOT_BOOTSTRAP_MAX_SECONDS = 8.0
ROOT_BOOTSTRAP_TIME_FRACTION = 0.20
ROOT_BOOTSTRAP_ROUTES_PER_TYPE = 18
ROOT_PRICING_EXTRA_SWEEPS = 8
ROOT_PRICING_TARGET_NEW_ROUTES_PER_LP = 48
EXACT_PRICING_NATIVE_MIN_BUCKET_SIZE = 2
EXACT_PRICING_NUMPY_MIN_BUCKET_SIZE = 16
# Bump whenever the default arc-flow MIP variable set or feasibility logic changes
# in a way that can invalidate persisted .mst/.sol starts.
DEFAULT_MIP_FORMULATION_VERSION = 12
DEFAULT_MIP_PROFILE_PATH = Path(__file__).with_name("gurobi_mip_profiles.json")
DEFAULT_INCUMBENT_REPORT_MIN_IMPROVEMENT = 1.0
_NATIVE_PRICING_LIBRARY: Any | None | bool = None
_NATIVE_PRICING_HELPER: Any | None | bool = None
_NATIVE_PRICING_REWARD_HELPER: Any | None | bool = None


def native_pricing_library() -> Any | None:
    global _NATIVE_PRICING_LIBRARY
    if _NATIVE_PRICING_LIBRARY is False:
        return None
    if _NATIVE_PRICING_LIBRARY is not None:
        return _NATIVE_PRICING_LIBRARY
    if np is None:
        _NATIVE_PRICING_LIBRARY = False
        return None

    source_path = Path(__file__).with_name("native_pricing_helper.c")
    if not source_path.is_file():
        _NATIVE_PRICING_LIBRARY = False
        return None

    cache_dir = Path(__file__).resolve().parent / ".native_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    suffix = ".dylib" if sys.platform == "darwin" else ".so"
    binary_path = cache_dir / f"{source_path.stem}{suffix}"

    try:
        rebuild_required = (
            not binary_path.is_file()
            or binary_path.stat().st_mtime < source_path.stat().st_mtime
        )
    except OSError:
        rebuild_required = True

    if rebuild_required:
        temp_binary_path = binary_path.with_name(
            f"{binary_path.stem}.{os.getpid()}{binary_path.suffix}"
        )
        compile_command = [os.environ.get("CC", "cc"), "-O3", "-std=c99"]
        if sys.platform == "darwin":
            compile_command.extend(["-dynamiclib", "-o", str(temp_binary_path)])
        else:
            compile_command.extend(["-shared", "-fPIC", "-o", str(temp_binary_path)])
        compile_command.append(str(source_path))
        try:
            subprocess.run(
                compile_command,
                check=True,
                capture_output=True,
                text=True,
            )
            os.replace(temp_binary_path, binary_path)
        except Exception:
            temp_binary_path.unlink(missing_ok=True)
            if not binary_path.is_file():
                _NATIVE_PRICING_LIBRARY = False
                return None

    try:
        library = ctypes.CDLL(str(binary_path))
    except Exception:
        _NATIVE_PRICING_LIBRARY = False
        return None

    _NATIVE_PRICING_LIBRARY = library
    return library


def native_pricing_helper() -> Any | None:
    global _NATIVE_PRICING_HELPER
    if _NATIVE_PRICING_HELPER is False:
        return None
    if _NATIVE_PRICING_HELPER is not None:
        return _NATIVE_PRICING_HELPER
    library = native_pricing_library()
    if library is None:
        _NATIVE_PRICING_HELPER = False
        return None
    try:
        function = library.pricing_dominance_scan
        function.argtypes = [
            ctypes.POINTER(ctypes.c_int32),
            ctypes.c_size_t,
            ctypes.POINTER(ctypes.c_double),
            ctypes.POINTER(ctypes.c_double),
            ctypes.POINTER(ctypes.c_double),
            ctypes.POINTER(ctypes.c_uint64),
            ctypes.POINTER(ctypes.c_uint8),
            ctypes.c_double,
            ctypes.c_double,
            ctypes.c_double,
            ctypes.c_uint64,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.POINTER(ctypes.c_int32),
            ctypes.POINTER(ctypes.c_size_t),
        ]
        function.restype = ctypes.c_int
    except Exception:
        _NATIVE_PRICING_HELPER = False
        return None
    _NATIVE_PRICING_HELPER = function
    return function


def native_pricing_reward_helper() -> Any | None:
    global _NATIVE_PRICING_REWARD_HELPER
    if _NATIVE_PRICING_REWARD_HELPER is False:
        return None
    if _NATIVE_PRICING_REWARD_HELPER is not None:
        return _NATIVE_PRICING_REWARD_HELPER
    library = native_pricing_library()
    if library is None:
        _NATIVE_PRICING_REWARD_HELPER = False
        return None
    try:
        function = library.pricing_optimistic_reward
        function.argtypes = [
            ctypes.POINTER(ctypes.c_uint64),
            ctypes.POINTER(ctypes.c_double),
            ctypes.POINTER(ctypes.c_double),
            ctypes.POINTER(ctypes.c_double),
            ctypes.POINTER(ctypes.c_uint64),
            ctypes.POINTER(ctypes.c_int32),
            ctypes.POINTER(ctypes.c_uint64),
            ctypes.c_size_t,
            ctypes.POINTER(ctypes.c_int32),
            ctypes.c_size_t,
            ctypes.POINTER(ctypes.c_int32),
            ctypes.c_size_t,
            ctypes.c_uint64,
            ctypes.c_uint64,
            ctypes.c_int32,
            ctypes.c_double,
            ctypes.c_double,
        ]
        function.restype = ctypes.c_double
    except Exception:
        _NATIVE_PRICING_REWARD_HELPER = False
        return None
    _NATIVE_PRICING_REWARD_HELPER = function
    return function

# ============================================================================
# Hardcoded business rules
# Keep operational exceptions centralized here so they are easy to spot when
# scrolling through the file. The workbook remains the source of the baseline
# data; this section owns the extra "real world" routing rules.
# ============================================================================


@dataclass(frozen=True)
class RequiredAdjacentStopPairRule:
    stop_a: int
    stop_b: int
    combined_service_min: float
    description: str

    @property
    def ordered_pair(self) -> tuple[int, int]:
        return tuple(sorted((self.stop_a, self.stop_b)))

    @property
    def directed_arcs(self) -> tuple[tuple[int, int], tuple[int, int]]:
        return ((self.stop_a, self.stop_b), (self.stop_b, self.stop_a))


@dataclass(frozen=True)
class MandatoryLunchBreakRule:
    window_start_min: float
    window_end_min: float
    duration_min: float
    description: str

    @property
    def latest_start_min(self) -> float:
        return self.window_end_min - self.duration_min


@dataclass(frozen=True)
class StopServiceWindowRule:
    stop_id: int
    earliest_min: float | None
    latest_min: float | None
    description: str


@dataclass(frozen=True)
class StopAllowedVehicleTypesRule:
    stop_id: int
    allowed_vehicle_types: frozenset[str]
    description: str


@dataclass(frozen=True)
class VehicleStopExclusionRule:
    vehicle_id: int
    forbidden_stops: frozenset[int]
    description: str


@dataclass(frozen=True)
class VehicleDemandServiceTimeAdjustmentRule:
    vehicle_id: int
    demand_threshold_kg: float
    extra_service_min: float
    description: str


@dataclass(frozen=True)
class TravelTimeAdjustmentRule:
    affected_stops: frozenset[int]
    extra_minutes_per_leg: float
    description: str


HILLCREST_SMALL_VAN_ONLY_STOPS: tuple[int, ...] = (18, 19, 20, 21, 22)
OAKWOOD_NOISE_ORDINANCE_STOPS: tuple[int, ...] = (10, 11, 12, 13)
NORTHERN_DISTRICT_STOPS: frozenset[int] = frozenset({23, 24, 25, 26, 27, 28})
MAIN_STREET_CONSTRUCTION_STOPS: frozenset[int] = frozenset({1, 2, 3, 4, 5, 6})
MAIN_STREET_CONSTRUCTION_RULE = TravelTimeAdjustmentRule(
    affected_stops=MAIN_STREET_CONSTRUCTION_STOPS,
    extra_minutes_per_leg=20.0,
    description=(
        "Main Street construction: treat every plan as a Monday-Thursday construction "
        "day, and add 20 minutes to any travel leg that goes to or from Stops 1-6."
    ),
)


REDSTONE_PLAZA_PAIR_RULES: tuple[RequiredAdjacentStopPairRule, ...] = (
    RequiredAdjacentStopPairRule(
        stop_a=7,
        stop_b=8,
        combined_service_min=45.0,
        description=(
            "Redstone Plaza mall stores must be handled in one truck trip with an "
            "immediate 7->8 or 8->7 handoff and a combined 45-minute service time."
        ),
    ),
)

HARD_STOP_SERVICE_WINDOW_RULES: tuple[StopServiceWindowRule, ...] = (
    StopServiceWindowRule(
        stop_id=3,
        earliest_min=12.75 * 60.0,
        latest_min=17.25 * 60.0,
        description=(
            "Redstone Medical Center (Stop 3) contractual SLA: service must start "
            "between 12:45 PM and 5:15 PM."
        ),
    ),
    *(
        StopServiceWindowRule(
            stop_id=stop_id,
            earliest_min=8.0 * 60.0,
            latest_min=19.0 * 60.0,
            description=(
                "Oakwood noise ordinance: Stops 10-13 may only be served between "
                "8:00 AM and 7:00 PM."
            ),
        )
        for stop_id in OAKWOOD_NOISE_ORDINANCE_STOPS
    ),
    StopServiceWindowRule(
        stop_id=14,
        earliest_min=10.0 * 60.0,
        latest_min=None,
        description="Sunrise Senior Living (Stop 14) may not be served before 10:00 AM.",
    ),
    StopServiceWindowRule(
        stop_id=27,
        earliest_min=None,
        latest_min=16.0 * 60.0,
        description=(
            "Redstone Elementary School cafeteria (Stop 27) county contract: "
            "service must start before 4:00 PM."
        ),
    ),
)
HARD_STOP_SERVICE_WINDOW_RULE_BY_STOP = {
    rule.stop_id: rule for rule in HARD_STOP_SERVICE_WINDOW_RULES
}

STOP_ALLOWED_VEHICLE_TYPE_RULES: tuple[StopAllowedVehicleTypesRule, ...] = (
    *(
        StopAllowedVehicleTypesRule(
            stop_id=stop_id,
            allowed_vehicle_types=frozenset({"Small Van"}),
            description=(
                "Hillcrest Road five-ton weight limit: Stops 18-22 may only be "
                "served by a Small Van."
            ),
        )
        for stop_id in HILLCREST_SMALL_VAN_ONLY_STOPS
    ),
)
STOP_ALLOWED_VEHICLE_TYPE_RULE_BY_STOP = {
    rule.stop_id: rule for rule in STOP_ALLOWED_VEHICLE_TYPE_RULES
}

VEHICLE_STOP_EXCLUSION_RULES: tuple[VehicleStopExclusionRule, ...] = (
    VehicleStopExclusionRule(
        vehicle_id=5,
        forbidden_stops=NORTHERN_DISTRICT_STOPS,
        description=(
            "Driver 5 northern district restriction: Vehicle 5 may not service "
            "Stops 23-28."
        ),
    ),
)
VEHICLE_STOP_EXCLUSION_RULE_BY_VEHICLE = {
    rule.vehicle_id: rule for rule in VEHICLE_STOP_EXCLUSION_RULES
}

VEHICLE_DEMAND_SERVICE_TIME_ADJUSTMENT_RULES: tuple[
    VehicleDemandServiceTimeAdjustmentRule, ...
] = (
    VehicleDemandServiceTimeAdjustmentRule(
        vehicle_id=3,
        demand_threshold_kg=50.0,
        extra_service_min=15.0,
        description=(
            "Medium Truck #3 faulty liftgate: when Vehicle 3 serves a stop with "
            "demand over 50 kg, add 15 minutes of service time for the two-person "
            "crew workaround."
        ),
    ),
)
VEHICLE_DEMAND_SERVICE_TIME_ADJUSTMENT_RULE_BY_VEHICLE = {
    rule.vehicle_id: rule for rule in VEHICLE_DEMAND_SERVICE_TIME_ADJUSTMENT_RULES
}

MIDDAY_LUNCH_BREAK_RULE = MandatoryLunchBreakRule(
    window_start_min=11.5 * 60.0,
    window_end_min=13.5 * 60.0,
    duration_min=30.0,
    description=(
        "Every used driver route must include one 30-minute lunch break scheduled "
        "between 11:30 AM and 1:30 PM."
    ),
)

CONSTRUCTIVE_ARC_FLOW_SEED_BY_VEHICLE_TYPE: dict[str, tuple[tuple[int, ...], ...]] = {
    "Small Van": (
        (5, 22, 18),
        (20, 19, 21),
    ),
    "Medium Truck": (
        (1, 9, 11, 15, 10, 12, 16, 17, 26),
        (4, 24, 6, 28, 27, 25, 23),
    ),
    "Large Truck": (
        (7, 8, 13, 2, 14, 3),
    ),
}

REQUIRED_ADJACENT_STOP_PARTNER: dict[int, int] = {
    stop_id: partner_id
    for rule in REDSTONE_PLAZA_PAIR_RULES
    for stop_id, partner_id in ((rule.stop_a, rule.stop_b), (rule.stop_b, rule.stop_a))
}
REQUIRED_SAME_ROUTE_PAIRS = frozenset(rule.ordered_pair for rule in REDSTONE_PLAZA_PAIR_RULES)


@dataclass(frozen=True)
class ArcFlowMIPProfile:
    name: str
    description: str
    params: dict[str, float | int | str]


def load_arc_flow_mip_profiles(
    profile_path: Path = DEFAULT_MIP_PROFILE_PATH,
) -> dict[str, ArcFlowMIPProfile]:
    """Load named arc-flow tuning profiles from a repo-local JSON file."""

    try:
        raw_profiles = json.loads(profile_path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ValueError(f"Arc-flow MIP profile file not found: {profile_path}") from exc
    except json.JSONDecodeError as exc:
        raise ValueError(f"Arc-flow MIP profile file is not valid JSON: {profile_path}") from exc

    if not isinstance(raw_profiles, dict):
        raise ValueError(
            f"Arc-flow MIP profile file must contain a JSON object at the top level: {profile_path}"
        )

    profiles: dict[str, ArcFlowMIPProfile] = {}
    for profile_name, payload in raw_profiles.items():
        if not isinstance(profile_name, str) or not profile_name:
            raise ValueError(f"Arc-flow MIP profile names must be non-empty strings: {profile_path}")
        if not isinstance(payload, dict):
            raise ValueError(
                f"Arc-flow MIP profile '{profile_name}' must map to an object in {profile_path}"
            )

        description = payload.get("description")
        params = payload.get("params")
        if not isinstance(description, str) or not description.strip():
            raise ValueError(
                f"Arc-flow MIP profile '{profile_name}' needs a non-empty description in {profile_path}"
            )
        if not isinstance(params, dict) or not params:
            raise ValueError(
                f"Arc-flow MIP profile '{profile_name}' needs a non-empty params object in {profile_path}"
            )

        normalized_params: dict[str, float | int | str] = {}
        for param_name, param_value in params.items():
            if not isinstance(param_name, str) or not param_name:
                raise ValueError(
                    f"Arc-flow MIP profile '{profile_name}' contains an invalid parameter name"
                )
            if not isinstance(param_value, (int, float, str)):
                raise ValueError(
                    f"Arc-flow MIP profile '{profile_name}' uses an unsupported value for "
                    f"parameter '{param_name}'"
                )
            normalized_params[param_name] = param_value

        profiles[profile_name] = ArcFlowMIPProfile(
            name=profile_name,
            description=description.strip(),
            params=normalized_params,
        )

    return profiles


def apply_arc_flow_mip_profile(
    model: gp.Model,
    profile: ArcFlowMIPProfile,
    *,
    skip_params: frozenset[str] = frozenset(),
) -> None:
    """Apply one named tuning profile to the default arc-flow MIP model."""

    for param_name, param_value in profile.params.items():
        if param_name in skip_params:
            continue
        try:
            model.setParam(param_name, param_value)
        except (AttributeError, TypeError, gp.GurobiError) as exc:
            raise ValueError(
                f"Arc-flow MIP profile '{profile.name}' contains unsupported Gurobi parameter "
                f"'{param_name}'"
            ) from exc

# Optional fallback WLS credentials for hosted environments like Google Colab.
# Prefer storing secrets in a .env file beside this script:
#   GUROBI_WLSACCESSID=...
#   GUROBI_WLSSECRET=...
#   GUROBI_LICENSEID=...
# Environment variables override .env, and these hardcoded values are used
# only as a final fallback. Leave them blank to avoid storing secrets here.
GUROBI_WLS_LICENSE: dict[str, str] = {
    "WLSACCESSID": "",
    "WLSSECRET": "",
    "LICENSEID": "",
}

_PRICING_WORKER_INSTANCE: Instance | None = None
_PRICING_WORKER_BRANCH_CTX: BranchContext | None = None
_INTERRUPT_SIGNAL_NAME: str | None = None
_GUROBI_ENV: gp.Env | None = None
_LATEST_INTERRUPT_ARTIFACTS: dict[str, Path] = {}
_MISSING = object()


@dataclass(frozen=True)
class StopData:
    stop_id: int
    customer_name: str
    address: str
    demand_kg: float
    earliest_min: float
    latest_min: float
    service_min: float


@dataclass(frozen=True)
class VehicleData:
    vehicle_id: int
    vehicle_type: str
    capacity_kg: float
    cost_per_mile: float
    fixed_daily_cost: float
    available_from_min: float
    available_until_min: float


@dataclass(frozen=True)
class VehicleTypeData:
    type_id: str
    vehicle_ids: tuple[int, ...]
    vehicle_type: str
    capacity_kg: float
    cost_per_mile: float
    fixed_daily_cost: float
    available_from_min: float
    available_until_min: float
    count: int


@dataclass(frozen=True)
class Instance:
    stops: dict[int, StopData]
    vehicles: dict[int, VehicleData]
    distance_miles: dict[tuple[int, int], float]
    travel_minutes: dict[tuple[int, int], float]
    cost_params: dict[str, float]
    depot_open_min: float
    depot_close_min: float


@dataclass
class RouteColumn:
    type_id: str
    vehicle_type: str
    stop_sequence: tuple[int, ...]
    stop_set: frozenset[int]
    arc_set: frozenset[tuple[int, int]]
    cost: float
    load_kg: float
    distance_mi: float
    active_min: float
    return_min: float
    service_start_min: dict[int, float]
    late_min: dict[int, float]
    lunch_break_start_min: float | None


@dataclass(frozen=True)
class RouteTimingEvaluation:
    service_start_min: dict[int, float]
    late_min: dict[int, float]
    return_min: float
    lunch_break_start_min: float | None


@dataclass(frozen=True)
class BranchState:
    forbidden_arcs: frozenset[tuple[int, int]]
    forced_arcs: frozenset[tuple[int, int]]
    excluded_assignments: frozenset[tuple[int, str]]
    forced_assignments: frozenset[tuple[int, str]]
    same_route_pairs: frozenset[tuple[int, int]]
    different_route_pairs: frozenset[tuple[int, int]]


@dataclass
class BranchContext:
    forbidden_arcs: frozenset[tuple[int, int]]
    forced_arcs: tuple[tuple[int, int], ...]
    forced_successor: dict[int, int]
    forced_predecessor: dict[int, int]
    forced_type_for_stop: dict[int, str]
    excluded_types_by_stop: dict[int, set[str]]
    together_groups: tuple[frozenset[int], ...]
    together_group_by_stop: dict[int, frozenset[int]]
    together_stop_pairs: frozenset[tuple[int, int]]
    separated_stop_pairs: frozenset[tuple[int, int]]
    separated_stops_by_stop: dict[int, set[int]]


@dataclass
class NodeRelaxationResult:
    lower_bound: float
    routes: list[RouteColumn]
    route_values: list[float]
    cg_iterations: int


@dataclass
class NodeMasterLP:
    model: gp.Model
    routes: list[RouteColumn]
    route_vars: list[gp.Var]
    artificial_vars: dict[int, gp.Var]
    cover_constraints: dict[int, gp.Constr]
    type_constraints: dict[str, gp.Constr]


@dataclass
class BranchAndPriceResult:
    status: str
    objective: float | None
    best_bound: float | None
    gap: float | None
    selected_routes: list[RouteColumn]
    vehicle_types: dict[str, VehicleTypeData]
    nodes_processed: int
    nodes_pruned: int
    global_columns: int
    root_lp_objective: float | None


@dataclass
class ArcFlowVehicleSolution:
    vehicle_id: int
    vehicle: VehicleData
    route: tuple[int, ...]
    load_kg: float
    depart_min: float
    return_min: float
    drive_min: float
    service_min: float
    active_min: float
    billed_hours: int
    distance_cost: float
    route_cost: float
    service_start_min: dict[int, float]
    late_min: dict[int, float]
    lunch_break_start_min: float | None


@dataclass
class ArcFlowSolveResult:
    status: str
    objective: float | None
    best_bound: float | None
    gap: float | None
    vehicle_solutions: list[ArcFlowVehicleSolution]
    fixed_arc_count: int
    symmetry_group_count: int
    bound_plot_path: Path | None = None


@dataclass
class MIPBoundPoint:
    elapsed_seconds: float
    upper_bound: float | None
    lower_bound: float | None


@dataclass
class NodeRelaxationCheckpointState:
    iteration: int
    last_lp_objective: float | None
    active_routes: list[RouteColumn]
    stabilized_cover_duals: dict[int, float] | None
    stabilized_type_duals: dict[str, float] | None
    previous_raw_cover_duals: dict[int, float] | None
    previous_raw_type_duals: dict[str, float] | None


@dataclass
class CurrentNodeCheckpointState:
    lower_bound_hint: float
    depth: int
    sequence_id: int
    branch_state: BranchState
    relaxation_state: NodeRelaxationCheckpointState | None


@dataclass
class BranchAndPriceCheckpointState:
    route_pool: list[RouteColumn]
    active_nodes: list[tuple[float, int, int, BranchState]]
    next_sequence_id: int
    incumbent_obj: float | None
    incumbent_routes: list[RouteColumn]
    root_lp_objective: float | None
    nodes_processed: int
    nodes_pruned: int
    current_node: CurrentNodeCheckpointState | None
    status: str


@dataclass(frozen=True, slots=True)
class PricingLabel:
    node: int
    visited_mask: int
    load_kg: float
    time_after_service: float
    lunch_break_taken: bool
    reduced_cost: float
    nonlabor_reduced_cost: float
    predecessor_idx: int | None


@dataclass(slots=True)
class PricingLabelBucket:
    label_indices: list[int]
    min_time_after_service: float
    min_load_kg: float
    min_nonlabor_reduced_cost: float
    max_time_after_service: float
    max_load_kg: float
    max_nonlabor_reduced_cost: float
    union_visited_mask: int
    intersection_visited_mask: int
    cached_index_array: Any | None
    cache_dirty: bool


class ProgressTracker:
    def __init__(self, enabled: bool, interval_seconds: float) -> None:
        self.enabled = enabled
        self.interval_seconds = max(interval_seconds, 0.1)
        self.start_time = perf_counter()
        self.last_emit = self.start_time - self.interval_seconds

    def emit(self, message: str, *, force: bool = False) -> None:
        if not self.enabled:
            return
        now = perf_counter()
        if not force and now - self.last_emit < self.interval_seconds:
            return
        elapsed = now - self.start_time
        print(f"[{elapsed:7.1f}s] {message}", flush=True)
        self.last_emit = now


class MIPBoundLogger:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.points: list[MIPBoundPoint] = []
        self._last_upper_bound: float | None | object = _MISSING
        self._last_lower_bound: float | None | object = _MISSING

    @staticmethod
    def _normalize_bound(value: float) -> float | None:
        if not math.isfinite(value):
            return None
        if abs(value) >= GRB.INFINITY / 2:
            return None
        return float(value)

    def record(self, elapsed_seconds: float, upper_bound: float | None, lower_bound: float | None) -> None:
        if (
            self._last_upper_bound is not _MISSING
            and self._last_lower_bound is not _MISSING
            and optional_float_changed(self._last_upper_bound, upper_bound) is False
            and optional_float_changed(self._last_lower_bound, lower_bound) is False
        ):
            return
        self.points.append(
            MIPBoundPoint(
                elapsed_seconds=max(elapsed_seconds, 0.0),
                upper_bound=upper_bound,
                lower_bound=lower_bound,
            )
        )
        self._last_upper_bound = upper_bound
        self._last_lower_bound = lower_bound

    def record_callback(self, cb_model: gp.Model, where: int) -> None:
        code_map = {
            GRB.Callback.MIP: (GRB.Callback.MIP_OBJBST, GRB.Callback.MIP_OBJBND),
            GRB.Callback.MIPSOL: (GRB.Callback.MIPSOL_OBJBST, GRB.Callback.MIPSOL_OBJBND),
            GRB.Callback.MIPNODE: (GRB.Callback.MIPNODE_OBJBST, GRB.Callback.MIPNODE_OBJBND),
        }
        if where not in code_map:
            return
        upper_code, lower_code = code_map[where]
        elapsed_seconds = cb_model.cbGet(GRB.Callback.RUNTIME)
        upper_bound = self._normalize_bound(cb_model.cbGet(upper_code))
        lower_bound = self._normalize_bound(cb_model.cbGet(lower_code))
        self.record(elapsed_seconds, upper_bound, lower_bound)

    def record_final_model_state(self, model: gp.Model) -> None:
        elapsed_seconds = 0.0
        try:
            elapsed_seconds = max(float(model.Runtime), 0.0)
        except (AttributeError, TypeError, ValueError, gp.GurobiError):
            elapsed_seconds = 0.0

        sol_count = 0
        try:
            sol_count = int(model.SolCount)
        except (AttributeError, TypeError, ValueError, gp.GurobiError):
            sol_count = 0

        upper_bound: float | None = None
        if sol_count > 0:
            try:
                upper_bound = self._normalize_bound(model.ObjVal)
            except (AttributeError, TypeError, ValueError, gp.GurobiError):
                upper_bound = None

        try:
            lower_bound = self._normalize_bound(model.ObjBound)
        except (AttributeError, TypeError, ValueError, gp.GurobiError):
            lower_bound = None
        self.record(elapsed_seconds, upper_bound, lower_bound)

    def has_points(self) -> bool:
        return bool(self.points)

    def latest_upper_bound(self) -> float | None:
        if not self.points:
            return None
        return self.points[-1].upper_bound

    def latest_lower_bound(self) -> float | None:
        if not self.points:
            return None
        return self.points[-1].lower_bound

    def save_plot(self, title: str) -> Path:
        self.path.parent.mkdir(parents=True, exist_ok=True)

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except ImportError as exc:
            raise RuntimeError(
                "Saving the MIP bound plot requires matplotlib to be installed."
            ) from exc

        figure, axis = plt.subplots(figsize=(9, 5))
        has_finite_upper_bounds = any(point.upper_bound is not None for point in self.points)
        has_finite_lower_bounds = any(point.lower_bound is not None for point in self.points)

        if self.points:
            times = [point.elapsed_seconds for point in self.points]
            upper_bounds = [
                math.nan if point.upper_bound is None else point.upper_bound for point in self.points
            ]
            lower_bounds = [
                math.nan if point.lower_bound is None else point.lower_bound for point in self.points
            ]
            if has_finite_lower_bounds:
                axis.step(times, lower_bounds, where="post", label="Lower Bound", linewidth=2.0)
            if has_finite_upper_bounds:
                axis.step(times, upper_bounds, where="post", label="Upper Bound", linewidth=2.0)
        else:
            axis.set_xlim(0.0, 1.0)

        axis.set_xlabel("Elapsed Seconds")
        axis.set_ylabel("Objective Value")
        axis.set_title(title)
        axis.grid(True, alpha=0.3)
        if has_finite_upper_bounds or has_finite_lower_bounds:
            axis.legend()
        else:
            axis.text(
                0.5,
                0.5,
                "No finite MIP bounds were available before the solve stopped.",
                transform=axis.transAxes,
                ha="center",
                va="center",
                fontsize=10,
            )
        figure.tight_layout()
        figure.savefig(self.path, dpi=150)
        plt.close(figure)
        return self.path


def capture_mip_bound_snapshot(model: gp.Model, bound_logger: MIPBoundLogger) -> None:
    try:
        bound_logger.record_final_model_state(model)
    except Exception:
        return


def save_mip_bound_plot_snapshot(
    workbook_path: Path,
    bound_logger: MIPBoundLogger,
    *,
    status_label: str,
) -> Path | None:
    try:
        return bound_logger.save_plot(f"{workbook_path.stem} MIP Bounds ({status_label})")
    except Exception:
        return None


class MIPPersistenceManager:
    VERSION = 1

    def __init__(self, path: Path) -> None:
        self.path = path.resolve()
        self.resume_mip_start_path = self.path / "resume.mst"
        self.resume_solution_path = self.path / "resume.sol"
        self.runs_dir = self.path / "runs"
        self.current_run_dir: Path | None = None
        self.solution_prefix: Path | None = None
        self.state_path = self.path / "state.json"
        self.last_saved_resume_objective: float | None = None

    def ensure_dir(self) -> None:
        self.path.mkdir(parents=True, exist_ok=True)
        self.runs_dir.mkdir(parents=True, exist_ok=True)

    def _autosave_solution_paths(self, root: Path | None = None) -> list[Path]:
        search_root = self.path if root is None else root
        solution_paths: list[tuple[float, str, Path]] = []
        for candidate in search_root.rglob("incumbent_*.sol"):
            suffix = candidate.stem.removeprefix("incumbent_")
            if not suffix.isdigit():
                continue
            try:
                mtime = candidate.stat().st_mtime
            except OSError:
                continue
            solution_paths.append((mtime, str(candidate), candidate))
        return [path for _, _, path in sorted(solution_paths)]

    def latest_autosave_solution_path(self, root: Path | None = None) -> Path | None:
        solution_paths = self._autosave_solution_paths(root)
        if solution_paths:
            return solution_paths[-1]
        return None

    def _cleanup_autosave_files(self, root: Path | None = None) -> None:
        for candidate in self._autosave_solution_paths(root):
            candidate.unlink(missing_ok=True)

    def _clear_saved_starts(self) -> None:
        self._cleanup_autosave_files()
        self.resume_mip_start_path.unlink(missing_ok=True)
        self.resume_solution_path.unlink(missing_ok=True)

    def _read_state(self) -> dict[str, Any] | None:
        if not self.state_path.is_file():
            return None
        with self.state_path.open("r", encoding="utf-8") as handle:
            return json.load(handle)

    def _write_state(self, payload: dict[str, Any]) -> None:
        temp_path = self.state_path.with_suffix(".tmp")
        with temp_path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, sort_keys=True)
            handle.write("\n")
        temp_path.replace(self.state_path)

    def maybe_write_resume_mip_start(
        self,
        *,
        objective: float | None,
        min_improvement: float,
        named_values: list[tuple[str, float]],
        force: bool = False,
    ) -> Path | None:
        if objective is None:
            return None
        threshold = max(float(min_improvement), 0.0)
        if (
            not force
            and self.last_saved_resume_objective is not None
            and self.last_saved_resume_objective - objective < threshold - 1e-9
        ):
            return None

        self.ensure_dir()
        temp_path = self.resume_mip_start_path.with_suffix(".mst.tmp")
        with temp_path.open("w", encoding="utf-8") as handle:
            handle.write("# MIP start\n")
            for var_name, value in named_values:
                handle.write(f"{var_name} {value}\n")
        temp_path.replace(self.resume_mip_start_path)
        self.last_saved_resume_objective = objective
        return self.resume_mip_start_path

    def _start_candidate_paths(self, prior_state: dict[str, Any] | None) -> list[Path]:
        candidates: list[Path] = []
        seen: set[Path] = set()

        def add(candidate: Path | None) -> None:
            if candidate is None:
                return
            resolved = candidate.resolve()
            if resolved in seen or not resolved.is_file():
                return
            seen.add(resolved)
            candidates.append(resolved)

        add(self.resume_mip_start_path)
        add(self.resume_solution_path)
        if prior_state is not None:
            resume_mst_path = prior_state.get("resume_mip_start_path")
            resume_path = prior_state.get("resume_solution_path")
            latest_path = prior_state.get("latest_solution_path")
            if isinstance(resume_mst_path, str):
                add(Path(resume_mst_path))
            if isinstance(resume_path, str):
                add(Path(resume_path))
            if isinstance(latest_path, str):
                add(Path(latest_path))
        add(self.latest_autosave_solution_path())
        return candidates

    def _checkpoint_payload(
        self,
        *,
        workbook_path: Path,
        status: str,
        loaded_start_path: Path | None,
        latest_solution_path: Path | None,
        objective: float | None,
        best_bound: float | None,
        gap: float | None,
        note: str,
    ) -> dict[str, Any]:
        return {
            "version": self.VERSION,
            "formulation_version": DEFAULT_MIP_FORMULATION_VERSION,
            "saved_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "workbook_path": str(workbook_path.resolve()),
            "status": status,
            "loaded_start_path": None if loaded_start_path is None else str(loaded_start_path),
            "resume_mip_start_path": (
                str(self.resume_mip_start_path) if self.resume_mip_start_path.is_file() else None
            ),
            "resume_solution_path": (
                str(self.resume_solution_path) if self.resume_solution_path.is_file() else None
            ),
            "latest_solution_path": (
                None if latest_solution_path is None else str(latest_solution_path.resolve())
            ),
            "autosave_dir": (
                None if self.current_run_dir is None else str(self.current_run_dir.resolve())
            ),
            "objective": objective,
            "best_bound": best_bound,
            "gap": gap,
            "note": note,
            "exact_tree_resume_supported": False,
        }

    def load_start(self, model: gp.Model, workbook_path: Path) -> Path | None:
        self.ensure_dir()

        prior_state = self._read_state()
        skip_saved_start = False
        if prior_state is not None and prior_state.get("workbook_path") is not None:
            saved_workbook_path = Path(str(prior_state["workbook_path"])).resolve()
            if saved_workbook_path != workbook_path.resolve():
                print(
                    "Warning: persistence directory was last used with a different workbook; "
                    "starting fresh in this directory."
                )
                self._clear_saved_starts()
                skip_saved_start = True
        if prior_state is not None and not skip_saved_start:
            saved_formulation_version = int(prior_state.get("formulation_version", -1))
            if saved_formulation_version != DEFAULT_MIP_FORMULATION_VERSION:
                print(
                    "Warning: persistence directory contains a saved start from an older "
                    "default MIP formulation; starting fresh."
                )
                self._clear_saved_starts()
                skip_saved_start = True

        loaded_start_path: Path | None = None
        if not skip_saved_start:
            for candidate in self._start_candidate_paths(prior_state):
                try:
                    model.update()
                    model.read(str(candidate))
                    loaded_start_path = candidate
                    if candidate.suffix == ".mst":
                        if candidate != self.resume_mip_start_path:
                            shutil.copy2(candidate, self.resume_mip_start_path)
                    elif candidate != self.resume_solution_path:
                        shutil.copy2(candidate, self.resume_solution_path)
                    break
                except (OSError, gp.GurobiError) as exc:
                    print(
                        "Warning: could not load saved MIP start from "
                        f"{candidate}; trying next candidate ({exc})."
                    )

        self.current_run_dir = self.runs_dir / datetime.now().strftime("%Y%m%d_%H%M%S")
        self.current_run_dir.mkdir(parents=True, exist_ok=True)
        self.solution_prefix = self.current_run_dir / "incumbent"
        model.Params.SolFiles = str(self.solution_prefix)
        self._write_state(
            self._checkpoint_payload(
                workbook_path=workbook_path,
                status="RUNNING",
                loaded_start_path=loaded_start_path,
                latest_solution_path=None,
                objective=None,
                best_bound=None,
                gap=None,
                note="Autosave enabled for default arc-flow MIP solve",
            )
        )
        return loaded_start_path

    def finalize_run(
        self,
        *,
        workbook_path: Path,
        status: str,
        loaded_start_path: Path | None,
        objective: float | None,
        best_bound: float | None,
        gap: float | None,
        note: str,
        model: gp.Model | None = None,
    ) -> Path | None:
        self.ensure_dir()

        latest_solution_path = self.latest_autosave_solution_path(self.current_run_dir)
        if model is not None:
            try:
                model.write(str(self.resume_mip_start_path))
            except gp.GurobiError:
                pass
        if latest_solution_path is not None:
            shutil.copy2(latest_solution_path, self.resume_solution_path)

        resume_path = None
        if self.resume_mip_start_path.is_file():
            resume_path = self.resume_mip_start_path
        elif self.resume_solution_path.is_file():
            resume_path = self.resume_solution_path
        self._write_state(
            self._checkpoint_payload(
                workbook_path=workbook_path,
                status=status,
                loaded_start_path=loaded_start_path,
                latest_solution_path=latest_solution_path,
                objective=objective,
                best_bound=best_bound,
                gap=gap,
                note=note,
            )
        )
        return resume_path


class ObjectiveTraceLogger:
    FIELDNAMES = (
        "timestamp",
        "elapsed_seconds",
        "event",
        "status",
        "upper_bound",
        "lower_bound",
        "incumbent_objective",
        "best_bound",
        "relative_gap",
        "nodes_processed",
        "nodes_pruned",
        "global_columns",
        "queue_size",
        "node_depth",
        "note",
    )

    def __init__(self, path: Path | None, workbook_path: Path | None) -> None:
        self.path = path
        self.workbook_path = workbook_path
        self.start_time = perf_counter()
        self.latest_incumbent: float | None = None
        self.latest_best_bound: float | None = None
        self.latest_gap: float | None = None
        self.latest_status: str | None = None
        self._file = None
        self._writer: csv.DictWriter[str] | None = None

        if self.path is None:
            return

        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._file = self.path.open("w", encoding="utf-8", newline="")
        self._writer = csv.DictWriter(self._file, fieldnames=self.FIELDNAMES)
        self._writer.writeheader()
        self._flush()

    def _flush(self) -> None:
        if self._file is None:
            return
        self._file.flush()
        os.fsync(self._file.fileno())

    def record(
        self,
        *,
        event: str,
        status: str | None = None,
        incumbent_objective: float | None | object = _MISSING,
        best_bound: float | None | object = _MISSING,
        relative_gap: float | None | object = _MISSING,
        nodes_processed: int | None = None,
        nodes_pruned: int | None = None,
        global_columns: int | None = None,
        queue_size: int | None = None,
        node_depth: int | None = None,
        note: str = "",
    ) -> None:
        if incumbent_objective is not _MISSING:
            self.latest_incumbent = incumbent_objective
        if best_bound is not _MISSING:
            self.latest_best_bound = best_bound
        if relative_gap is not _MISSING:
            self.latest_gap = relative_gap
        if status is not None:
            self.latest_status = status

        if self._writer is None:
            return

        elapsed = perf_counter() - self.start_time
        self._writer.writerow(
            {
                "timestamp": datetime.now().astimezone().isoformat(timespec="seconds"),
                "elapsed_seconds": f"{elapsed:.6f}",
                "event": event,
                "status": self.latest_status or "",
                "upper_bound": (
                    "" if self.latest_incumbent is None else f"{self.latest_incumbent:.10f}"
                ),
                "lower_bound": (
                    "" if self.latest_best_bound is None else f"{self.latest_best_bound:.10f}"
                ),
                "incumbent_objective": (
                    "" if self.latest_incumbent is None else f"{self.latest_incumbent:.10f}"
                ),
                "best_bound": (
                    "" if self.latest_best_bound is None else f"{self.latest_best_bound:.10f}"
                ),
                "relative_gap": "" if self.latest_gap is None else f"{self.latest_gap:.10f}",
                "nodes_processed": "" if nodes_processed is None else nodes_processed,
                "nodes_pruned": "" if nodes_pruned is None else nodes_pruned,
                "global_columns": "" if global_columns is None else global_columns,
                "queue_size": "" if queue_size is None else queue_size,
                "node_depth": "" if node_depth is None else node_depth,
                "note": note,
            }
        )
        self._flush()

    def close(self) -> None:
        if self._file is None:
            return
        self._flush()
        self._file.close()
        self._file = None
        self._writer = None


class IncumbentReportWriter:
    def __init__(
        self,
        path: Path | None,
        *,
        min_improvement: float = DEFAULT_INCUMBENT_REPORT_MIN_IMPROVEMENT,
    ) -> None:
        self.path = None if path is None else path.resolve()
        self.min_improvement = max(float(min_improvement), 0.0)
        self.last_saved_objective: float | None = None

    def should_save(self, objective: float | None, *, force: bool = False) -> bool:
        if self.path is None or objective is None:
            return False
        if force or self.last_saved_objective is None:
            return True
        return self.last_saved_objective - objective >= self.min_improvement - 1e-9

    def maybe_save(self, objective: float | None, text: str, *, force: bool = False) -> Path | None:
        if not self.should_save(objective, force=force):
            return None
        if self.path is None or objective is None:
            return None
        saved_path = write_text_report(self.path, text)
        self.last_saved_objective = objective
        record_interrupt_artifact("incumbent_report", saved_path)
        return saved_path


class ArcFlowProgressLogger:
    FIELDNAMES = (
        "timestamp",
        "elapsed_seconds",
        "marker",
        "explored_nodes",
        "unexplored_nodes",
        "current_node_obj",
        "current_node_depth",
        "current_node_intinf",
        "incumbent_objective",
        "best_bound",
        "gap_percent",
        "simplex_iterations_per_node",
        "display_time_seconds",
        "raw_line",
    )

    _FULL_LINE_RE = re.compile(
        r"^\s*(?P<marker>[H\*]?)\s*"
        r"(?P<explored>\d+)\s+"
        r"(?P<unexplored>\d+)\s+"
        r"(?P<node_obj>(?:[+\-]?\d+(?:\.\d+)?|cutoff|infeasible))\s+"
        r"(?P<depth>\d+)\s+"
        r"(?P<intinf>\d+)\s+"
        r"(?P<incumbent>(?:[+\-]?\d+(?:\.\d+)?|-))\s+"
        r"(?P<bestbd>[+\-]?\d+(?:\.\d+)?)\s+"
        r"(?P<gap>(?:[+\-]?\d+(?:\.\d+)?|-))%?\s+"
        r"(?P<itnode>(?:[+\-]?\d+(?:\.\d+)?|-))\s+"
        r"(?P<time>\d+)s\s*$"
    )
    _INCUMBENT_LINE_RE = re.compile(
        r"^\s*(?P<marker>[H\*])\s*"
        r"(?P<explored>\d+)\s+"
        r"(?P<unexplored>\d+)\s+"
        r"(?P<incumbent>[+\-]?\d+(?:\.\d+)?)\s+"
        r"(?P<bestbd>[+\-]?\d+(?:\.\d+)?)\s+"
        r"(?P<gap>[+\-]?\d+(?:\.\d+)?)%\s+"
        r"(?P<itnode>(?:[+\-]?\d+(?:\.\d+)?|-))\s+"
        r"(?P<time>\d+)s\s*$"
    )

    def __init__(self, path: Path | None) -> None:
        self.path = None if path is None else path.resolve()
        self._file = None
        self._writer: csv.DictWriter[str] | None = None

        if self.path is None:
            return

        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._file = self.path.open("w", encoding="utf-8", newline="")
        self._writer = csv.DictWriter(self._file, fieldnames=self.FIELDNAMES)
        self._writer.writeheader()
        self._flush()

    def _flush(self) -> None:
        if self._file is None:
            return
        self._file.flush()
        os.fsync(self._file.fileno())

    @staticmethod
    def _parse_optional_float(token: str) -> float | None:
        if token in {"-", "cutoff", "infeasible"}:
            return None
        return float(token)

    @staticmethod
    def _parse_optional_int(token: str) -> int | None:
        if token in {"cutoff", "infeasible"}:
            return None
        return int(token)

    def record_message(self, line: str) -> bool:
        if self._writer is None:
            return False
        stripped_line = line.rstrip("\n")
        match = self._FULL_LINE_RE.match(stripped_line)
        row: dict[str, object]
        if match is not None:
            current_node_obj = self._parse_optional_float(match.group("node_obj"))
            incumbent_objective = self._parse_optional_float(match.group("incumbent"))
            gap_percent = self._parse_optional_float(match.group("gap"))
            simplex_iterations_per_node = self._parse_optional_float(match.group("itnode"))
            row = {
                "timestamp": datetime.now().astimezone().isoformat(timespec="seconds"),
                "elapsed_seconds": match.group("time"),
                "marker": match.group("marker") or "",
                "explored_nodes": int(match.group("explored")),
                "unexplored_nodes": int(match.group("unexplored")),
                "current_node_obj": "" if current_node_obj is None else current_node_obj,
                "current_node_depth": int(match.group("depth")),
                "current_node_intinf": int(match.group("intinf")),
                "incumbent_objective": "" if incumbent_objective is None else incumbent_objective,
                "best_bound": float(match.group("bestbd")),
                "gap_percent": "" if gap_percent is None else gap_percent,
                "simplex_iterations_per_node": (
                    "" if simplex_iterations_per_node is None else simplex_iterations_per_node
                ),
                "display_time_seconds": int(match.group("time")),
                "raw_line": stripped_line,
            }
        else:
            match = self._INCUMBENT_LINE_RE.match(stripped_line)
            if match is None:
                return False
            simplex_iterations_per_node = self._parse_optional_float(match.group("itnode"))
            row = {
                "timestamp": datetime.now().astimezone().isoformat(timespec="seconds"),
                "elapsed_seconds": match.group("time"),
                "marker": match.group("marker") or "",
                "explored_nodes": int(match.group("explored")),
                "unexplored_nodes": int(match.group("unexplored")),
                "current_node_obj": "",
                "current_node_depth": "",
                "current_node_intinf": "",
                "incumbent_objective": float(match.group("incumbent")),
                "best_bound": float(match.group("bestbd")),
                "gap_percent": float(match.group("gap")),
                "simplex_iterations_per_node": (
                    "" if simplex_iterations_per_node is None else simplex_iterations_per_node
                ),
                "display_time_seconds": int(match.group("time")),
                "raw_line": stripped_line,
            }
        self._writer.writerow(row)
        self._flush()
        return True

    def close(self) -> None:
        if self._file is None:
            return
        self._flush()
        self._file.close()
        self._file = None
        self._writer = None


def candidate_dotenv_paths() -> list[Path]:
    paths = [Path(__file__).resolve().with_name(".env"), Path.cwd() / ".env"]
    unique_paths: list[Path] = []
    seen: set[Path] = set()
    for path in paths:
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        unique_paths.append(resolved)
    return unique_paths


def parse_dotenv_line(raw_line: str) -> tuple[str, str] | None:
    line = raw_line.strip()
    if not line or line.startswith("#"):
        return None
    if line.startswith("export "):
        line = line[len("export ") :].strip()
    if "=" not in line:
        return None
    key, value = line.split("=", 1)
    key = key.strip()
    value = value.strip()
    if not key:
        return None
    if value and value[0] == value[-1] and value[0] in {"'", '"'}:
        value = value[1:-1]
    return key, value


def load_dotenv_file() -> dict[str, str]:
    for path in candidate_dotenv_paths():
        if not path.is_file():
            continue
        values: dict[str, str] = {}
        for line in path.read_text(encoding="utf-8").splitlines():
            parsed = parse_dotenv_line(line)
            if parsed is None:
                continue
            key, value = parsed
            values[key] = value
        return values
    return {}


def get_license_setting(
    *,
    dotenv_values: dict[str, str],
    env_key: str,
    legacy_env_key: str,
    fallback: str,
) -> str:
    if env_key in os.environ:
        return os.environ[env_key].strip()
    if legacy_env_key in os.environ:
        return os.environ[legacy_env_key].strip()
    if env_key in dotenv_values:
        return dotenv_values[env_key].strip()
    if legacy_env_key in dotenv_values:
        return dotenv_values[legacy_env_key].strip()
    return fallback.strip()


def resolve_gurobi_wls_license() -> dict[str, str | int] | None:
    dotenv_values = load_dotenv_file()

    access_id = get_license_setting(
        dotenv_values=dotenv_values,
        env_key="GUROBI_WLSACCESSID",
        legacy_env_key="WLSACCESSID",
        fallback=GUROBI_WLS_LICENSE["WLSACCESSID"],
    )
    secret = get_license_setting(
        dotenv_values=dotenv_values,
        env_key="GUROBI_WLSSECRET",
        legacy_env_key="WLSSECRET",
        fallback=GUROBI_WLS_LICENSE["WLSSECRET"],
    )
    license_id_text = get_license_setting(
        dotenv_values=dotenv_values,
        env_key="GUROBI_LICENSEID",
        legacy_env_key="LICENSEID",
        fallback=GUROBI_WLS_LICENSE["LICENSEID"],
    )

    configured_fields = [bool(access_id), bool(secret), bool(license_id_text)]
    if not any(configured_fields):
        return None
    if not all(configured_fields):
        raise ValueError(
            "Incomplete Gurobi WLS configuration. Set GUROBI_WLSACCESSID, "
            "GUROBI_WLSSECRET, and GUROBI_LICENSEID in a .env file beside "
            "vrp_gurobi.py, in environment variables, or in GUROBI_WLS_LICENSE."
        )

    try:
        license_id = int(license_id_text)
    except ValueError as exc:
        raise ValueError("GUROBI_LICENSEID must be an integer.") from exc

    return {
        "WLSACCESSID": access_id,
        "WLSSECRET": secret,
        "LICENSEID": license_id,
    }


def get_gurobi_env() -> gp.Env | None:
    global _GUROBI_ENV
    if _GUROBI_ENV is not None:
        return _GUROBI_ENV

    options = resolve_gurobi_wls_license()
    if options is None:
        return None

    _GUROBI_ENV = gp.Env(params=options)
    return _GUROBI_ENV


def dispose_gurobi_env() -> None:
    global _GUROBI_ENV
    if _GUROBI_ENV is None:
        return
    _GUROBI_ENV.dispose()
    _GUROBI_ENV = None


def create_gurobi_model(name: str) -> gp.Model:
    env = get_gurobi_env()
    if env is None:
        return gp.Model(name)
    return gp.Model(name, env=env)


def serialize_route_ref(route: RouteColumn) -> dict[str, Any]:
    return {
        "type_id": route.type_id,
        "stop_sequence": list(route.stop_sequence),
    }


def deserialize_route_refs(
    serialized_routes: list[dict[str, Any]],
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    route_lookup: dict[tuple[str, tuple[int, ...]], RouteColumn] | None = None,
) -> list[RouteColumn]:
    routes: list[RouteColumn] = []
    for item in serialized_routes:
        type_id = str(item["type_id"])
        stop_sequence = tuple(int(stop_id) for stop_id in item["stop_sequence"])
        key = (type_id, stop_sequence)
        route = route_lookup.get(key) if route_lookup is not None else None
        if route is None:
            vehicle_type = vehicle_types.get(type_id)
            if vehicle_type is None:
                raise ValueError(f"Checkpoint refers to unknown vehicle type {type_id!r}.")
            route = evaluate_route_sequence(instance, vehicle_type, stop_sequence)
        if route is None:
            raise ValueError(
                f"Checkpoint route {type_id} {stop_sequence} is infeasible under the workbook data."
            )
        routes.append(route)
    return routes


def serialize_branch_state(branch_state: BranchState) -> dict[str, Any]:
    return {
        "forbidden_arcs": [list(arc) for arc in sorted(branch_state.forbidden_arcs)],
        "forced_arcs": [list(arc) for arc in sorted(branch_state.forced_arcs)],
        "excluded_assignments": [
            [stop_id, type_id]
            for stop_id, type_id in sorted(branch_state.excluded_assignments)
        ],
        "forced_assignments": [
            [stop_id, type_id] for stop_id, type_id in sorted(branch_state.forced_assignments)
        ],
        "same_route_pairs": [list(pair) for pair in sorted(branch_state.same_route_pairs)],
        "different_route_pairs": [
            list(pair) for pair in sorted(branch_state.different_route_pairs)
        ],
    }


def deserialize_branch_state(data: dict[str, Any]) -> BranchState:
    return BranchState(
        forbidden_arcs=frozenset(
            (int(arc[0]), int(arc[1])) for arc in data.get("forbidden_arcs", [])
        ),
        forced_arcs=frozenset(
            (int(arc[0]), int(arc[1])) for arc in data.get("forced_arcs", [])
        ),
        excluded_assignments=frozenset(
            (int(item[0]), str(item[1])) for item in data.get("excluded_assignments", [])
        ),
        forced_assignments=frozenset(
            (int(item[0]), str(item[1])) for item in data.get("forced_assignments", [])
        ),
        same_route_pairs=frozenset(
            (int(pair[0]), int(pair[1])) for pair in data.get("same_route_pairs", [])
        ),
        different_route_pairs=frozenset(
            (int(pair[0]), int(pair[1])) for pair in data.get("different_route_pairs", [])
        ),
    )


def serialize_int_keyed_float_map(values: dict[int, float] | None) -> dict[str, float] | None:
    if values is None:
        return None
    return {str(key): float(values[key]) for key in sorted(values)}


def deserialize_int_keyed_float_map(
    data: dict[str, Any] | None,
) -> dict[int, float] | None:
    if data is None:
        return None
    return {int(key): float(value) for key, value in data.items()}


def serialize_str_keyed_float_map(values: dict[str, float] | None) -> dict[str, float] | None:
    if values is None:
        return None
    return {str(key): float(values[key]) for key in sorted(values)}


def deserialize_str_keyed_float_map(
    data: dict[str, Any] | None,
) -> dict[str, float] | None:
    if data is None:
        return None
    return {str(key): float(value) for key, value in data.items()}


def serialize_node_relaxation_checkpoint(
    checkpoint_state: NodeRelaxationCheckpointState | None,
) -> dict[str, Any] | None:
    if checkpoint_state is None:
        return None
    return {
        "iteration": checkpoint_state.iteration,
        "last_lp_objective": checkpoint_state.last_lp_objective,
        "active_routes": [
            serialize_route_ref(route) for route in checkpoint_state.active_routes
        ],
        "stabilized_cover_duals": serialize_int_keyed_float_map(
            checkpoint_state.stabilized_cover_duals
        ),
        "stabilized_type_duals": serialize_str_keyed_float_map(
            checkpoint_state.stabilized_type_duals
        ),
        "previous_raw_cover_duals": serialize_int_keyed_float_map(
            checkpoint_state.previous_raw_cover_duals
        ),
        "previous_raw_type_duals": serialize_str_keyed_float_map(
            checkpoint_state.previous_raw_type_duals
        ),
    }


def deserialize_node_relaxation_checkpoint(
    data: dict[str, Any] | None,
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    route_lookup: dict[tuple[str, tuple[int, ...]], RouteColumn],
) -> NodeRelaxationCheckpointState | None:
    if data is None:
        return None
    return NodeRelaxationCheckpointState(
        iteration=int(data["iteration"]),
        last_lp_objective=(
            None if data.get("last_lp_objective") is None else float(data["last_lp_objective"])
        ),
        active_routes=deserialize_route_refs(
            data.get("active_routes", []),
            instance=instance,
            vehicle_types=vehicle_types,
            route_lookup=route_lookup,
        ),
        stabilized_cover_duals=deserialize_int_keyed_float_map(
            data.get("stabilized_cover_duals")
        ),
        stabilized_type_duals=deserialize_str_keyed_float_map(
            data.get("stabilized_type_duals")
        ),
        previous_raw_cover_duals=deserialize_int_keyed_float_map(
            data.get("previous_raw_cover_duals")
        ),
        previous_raw_type_duals=deserialize_str_keyed_float_map(
            data.get("previous_raw_type_duals")
        ),
    )


def serialize_current_node_checkpoint(
    current_node: CurrentNodeCheckpointState | None,
) -> dict[str, Any] | None:
    if current_node is None:
        return None
    return {
        "lower_bound_hint": current_node.lower_bound_hint,
        "depth": current_node.depth,
        "sequence_id": current_node.sequence_id,
        "branch_state": serialize_branch_state(current_node.branch_state),
        "relaxation_state": serialize_node_relaxation_checkpoint(current_node.relaxation_state),
    }


def deserialize_current_node_checkpoint(
    data: dict[str, Any] | None,
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    route_lookup: dict[tuple[str, tuple[int, ...]], RouteColumn],
) -> CurrentNodeCheckpointState | None:
    if data is None:
        return None
    return CurrentNodeCheckpointState(
        lower_bound_hint=float(data["lower_bound_hint"]),
        depth=int(data["depth"]),
        sequence_id=int(data["sequence_id"]),
        branch_state=deserialize_branch_state(data["branch_state"]),
        relaxation_state=deserialize_node_relaxation_checkpoint(
            data.get("relaxation_state"),
            instance=instance,
            vehicle_types=vehicle_types,
            route_lookup=route_lookup,
        ),
    )


class CheckpointManager:
    VERSION = 1

    def __init__(self, path: Path | None) -> None:
        self.path = path

    def _write_payload(self, payload: dict[str, Any]) -> None:
        if self.path is None:
            return
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = self.path.with_name(f"{self.path.name}.tmp")
        with tmp_path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, sort_keys=True)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp_path, self.path)

    def save_branch_and_price_state(
        self,
        *,
        workbook_path: Path,
        route_pool: list[RouteColumn],
        active_nodes: list[tuple[float, int, int, BranchState]],
        next_sequence_id: int,
        incumbent_obj: float | None,
        incumbent_routes: list[RouteColumn],
        root_lp_objective: float | None,
        nodes_processed: int,
        nodes_pruned: int,
        current_node: CurrentNodeCheckpointState | None,
        status: str,
        note: str,
    ) -> None:
        if self.path is None:
            return
        payload = {
            "version": self.VERSION,
            "saved_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "workbook_path": str(workbook_path.resolve()),
            "status": status,
            "note": note,
            "route_pool": [serialize_route_ref(route) for route in route_pool],
            "active_nodes": [
                {
                    "bound": bound,
                    "depth": depth,
                    "sequence_id": sequence_id,
                    "branch_state": serialize_branch_state(branch_state),
                }
                for bound, depth, sequence_id, branch_state in active_nodes
            ],
            "next_sequence_id": next_sequence_id,
            "incumbent_obj": incumbent_obj,
            "incumbent_routes": [serialize_route_ref(route) for route in incumbent_routes],
            "root_lp_objective": root_lp_objective,
            "nodes_processed": nodes_processed,
            "nodes_pruned": nodes_pruned,
            "current_node": serialize_current_node_checkpoint(current_node),
        }
        self._write_payload(payload)


def load_branch_and_price_checkpoint(
    checkpoint_path: Path,
    *,
    workbook_path: Path,
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
) -> BranchAndPriceCheckpointState:
    with checkpoint_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)

    if int(payload.get("version", -1)) != CheckpointManager.VERSION:
        raise ValueError(
            f"Checkpoint version {payload.get('version')!r} is not supported by this solver."
        )

    saved_workbook_path = Path(payload["workbook_path"]).resolve()
    if saved_workbook_path != workbook_path.resolve():
        raise ValueError(
            "Checkpoint workbook path does not match the requested workbook. "
            f"Checkpoint: {saved_workbook_path}, requested: {workbook_path.resolve()}"
        )

    route_pool = deserialize_route_refs(
        payload.get("route_pool", []),
        instance=instance,
        vehicle_types=vehicle_types,
    )
    route_lookup = {route_column_key(route): route for route in route_pool}
    active_nodes = [
        (
            float(item["bound"]),
            int(item["depth"]),
            int(item["sequence_id"]),
            deserialize_branch_state(item["branch_state"]),
        )
        for item in payload.get("active_nodes", [])
    ]

    return BranchAndPriceCheckpointState(
        route_pool=route_pool,
        active_nodes=active_nodes,
        next_sequence_id=int(payload.get("next_sequence_id", len(active_nodes))),
        incumbent_obj=(
            None if payload.get("incumbent_obj") is None else float(payload["incumbent_obj"])
        ),
        incumbent_routes=deserialize_route_refs(
            payload.get("incumbent_routes", []),
            instance=instance,
            vehicle_types=vehicle_types,
            route_lookup=route_lookup,
        ),
        root_lp_objective=(
            None
            if payload.get("root_lp_objective") is None
            else float(payload["root_lp_objective"])
        ),
        nodes_processed=int(payload.get("nodes_processed", 0)),
        nodes_pruned=int(payload.get("nodes_pruned", 0)),
        current_node=deserialize_current_node_checkpoint(
            payload.get("current_node"),
            instance=instance,
            vehicle_types=vehicle_types,
            route_lookup=route_lookup,
        ),
        status=str(payload.get("status", "UNKNOWN")),
    )


def optimize_with_heartbeat(
    model: gp.Model,
    progress: ProgressTracker | None,
    label: str,
    bound_logger: MIPBoundLogger | None = None,
    mip_solution_callback: Callable[[gp.Model], None] | None = None,
    message_callback: Callable[[str], None] | None = None,
) -> None:
    if (
        (progress is None or not progress.enabled)
        and bound_logger is None
        and mip_solution_callback is None
        and message_callback is None
    ):
        model.optimize()
        return

    active_callback_points = {
        GRB.Callback.SIMPLEX,
        GRB.Callback.BARRIER,
        GRB.Callback.MIP,
        GRB.Callback.MIPNODE,
        GRB.Callback.MIPSOL,
    }
    heartbeat_interval = (
        progress.interval_seconds if progress is not None and progress.enabled else 0.0
    )
    last_runtime_report = [-heartbeat_interval]

    def callback(cb_model: gp.Model, where: int) -> None:
        if where == GRB.Callback.MESSAGE and message_callback is not None:
            try:
                message = str(cb_model.cbGet(GRB.Callback.MSG_STRING))
            except gp.GurobiError:
                message = ""
            if message:
                message_callback(message)
        if bound_logger is not None:
            bound_logger.record_callback(cb_model, where)
        if where == GRB.Callback.MIPSOL and mip_solution_callback is not None:
            mip_solution_callback(cb_model)
        if where not in active_callback_points:
            return
        if progress is None or not progress.enabled:
            return
        runtime = cb_model.cbGet(GRB.Callback.RUNTIME)
        if runtime < progress.interval_seconds:
            return
        if runtime - last_runtime_report[0] < progress.interval_seconds:
            return
        progress.emit(
            f"{label}: still solving ({runtime:.1f}s in current Gurobi solve)",
            force=True,
        )
        last_runtime_report[0] = runtime

    model.optimize(callback)


def excel_time_to_minutes(value: Any) -> float:
    if value is None:
        raise ValueError("Encountered an empty time value.")
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, dt_time):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, (int, float)):
        if 0 <= value <= 2:
            return float(value) * 24 * 60
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.hour * 60 + parsed.minute + parsed.second / 60
            except ValueError:
                pass
        try:
            return float(text)
        except ValueError as exc:
            raise ValueError(f"Unsupported time format: {value!r}") from exc
    raise TypeError(f"Unsupported time value type: {type(value)!r}")


def minutes_to_clock(minutes: float) -> str:
    total_minutes = int(round(minutes))
    hours, mins = divmod(total_minutes, 60)
    return f"{hours:02d}:{mins:02d}"


def parse_optional_int_limit(value: str) -> int | None:
    text = value.strip().lower()
    if text in {"none", "unlimited", "inf", "infinite"}:
        return None

    parsed = int(text)
    if parsed <= 0:
        return None
    return parsed


def optional_float_changed(previous: float | None, current: float | None, tol: float = 1e-9) -> bool:
    if previous is None or current is None:
        return previous != current
    return abs(previous - current) > tol


def default_objective_trace_path(workbook_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return workbook_path.with_name(f"{workbook_path.stem}_objective_trace_{timestamp}.csv")


def default_arc_flow_progress_log_path(workbook_path: Path, run_timestamp: str) -> Path:
    return workbook_path.parent / f"{workbook_path.stem}_arc_flow_progress_{run_timestamp}.csv"


def default_mip_bound_plot_path(workbook_path: Path, run_timestamp: str) -> Path:
    return workbook_path.parent / "plots" / f"{workbook_path.stem}_mip_bounds_{run_timestamp}.png"


def default_mip_persist_dir(workbook_path: Path) -> Path:
    return workbook_path.parent / ".mip_persist" / workbook_path.stem


def record_interrupt_artifact(label: str, path: Path | None) -> None:
    if path is None:
        return
    _LATEST_INTERRUPT_ARTIFACTS[label] = path.resolve()


def default_interrupt_artifact_report_path(workbook_path: Path) -> Path:
    return workbook_path.parent / f"{workbook_path.stem}_interrupt_artifacts.txt"


def default_incumbent_report_path(workbook_path: Path) -> Path:
    return workbook_path.parent / f"{workbook_path.stem}_incumbent_report.txt"


def write_interrupt_artifact_report(
    report_path: Path,
    *,
    workbook_path: Path,
    interrupt_note: str,
) -> Path:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"saved_at: {datetime.now().astimezone().isoformat(timespec='seconds')}",
        f"workbook_path: {workbook_path.resolve()}",
        f"interrupt_note: {interrupt_note}",
    ]
    for label, path in sorted(_LATEST_INTERRUPT_ARTIFACTS.items()):
        lines.append(f"{label}: {path}")
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report_path


def write_text_report(report_path: Path, text: str) -> Path:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(text, encoding="utf-8")
    return report_path


def install_interrupt_handlers() -> dict[int, Any]:
    previous_handlers: dict[int, Any] = {}

    def handler(signum: int, _frame: Any) -> None:
        global _INTERRUPT_SIGNAL_NAME
        _INTERRUPT_SIGNAL_NAME = signal.Signals(signum).name
        raise KeyboardInterrupt

    for signum in (signal.SIGINT, signal.SIGTERM):
        previous_handlers[signum] = signal.getsignal(signum)
        signal.signal(signum, handler)
    return previous_handlers


def restore_interrupt_handlers(previous_handlers: dict[int, Any]) -> None:
    for signum, previous_handler in previous_handlers.items():
        signal.signal(signum, previous_handler)


def parse_closed_unit_interval(value: str) -> float:
    parsed = float(value)
    if not 0.0 <= parsed <= 1.0:
        raise argparse.ArgumentTypeError("Value must be between 0 and 1, inclusive.")
    return parsed


def parse_positive_int(value: str) -> int:
    parsed = int(value)
    if parsed <= 0:
        raise argparse.ArgumentTypeError("Value must be a positive integer.")
    return parsed


def parse_worker_count(value: str) -> int | None:
    text = value.strip().lower()
    if text == "auto":
        return None
    parsed = int(text)
    if parsed <= 0:
        raise argparse.ArgumentTypeError("Worker count must be a positive integer or 'auto'.")
    return parsed


def remaining_time(deadline: float | None) -> float | None:
    if deadline is None:
        return None
    return max(deadline - perf_counter(), 0.0)


def relative_gap(incumbent: float | None, best_bound: float | None) -> float | None:
    if incumbent is None or best_bound is None:
        return None
    if abs(incumbent) <= INTEGRALITY_TOL:
        return 0.0 if abs(incumbent - best_bound) <= INTEGRALITY_TOL else None
    return max(incumbent - best_bound, 0.0) / abs(incumbent)


def gurobi_status_name(status: int) -> str:
    names = {
        GRB.OPTIMAL: "OPTIMAL",
        GRB.INFEASIBLE: "INFEASIBLE",
        GRB.INF_OR_UNBD: "INF_OR_UNBD",
        GRB.UNBOUNDED: "UNBOUNDED",
        GRB.CUTOFF: "CUTOFF",
        GRB.ITERATION_LIMIT: "ITERATION_LIMIT",
        GRB.NODE_LIMIT: "NODE_LIMIT",
        GRB.TIME_LIMIT: "TIME_LIMIT",
        GRB.SOLUTION_LIMIT: "SOLUTION_LIMIT",
        GRB.INTERRUPTED: "INTERRUPTED",
        GRB.NUMERIC: "NUMERIC",
        GRB.SUBOPTIMAL: "SUBOPTIMAL",
        GRB.USER_OBJ_LIMIT: "USER_OBJ_LIMIT",
        GRB.WORK_LIMIT: "WORK_LIMIT",
        GRB.MEM_LIMIT: "MEM_LIMIT",
    }
    return names.get(status, str(status))


def stabilize_dual_vector(
    raw_duals: dict[Any, float],
    previous_duals: dict[Any, float] | None,
    alpha: float,
) -> dict[Any, float]:
    if previous_duals is None or alpha >= 1.0 - 1e-12:
        return dict(raw_duals)
    if alpha <= 1e-12:
        return dict(previous_duals)
    return {
        key: alpha * raw_duals[key] + (1.0 - alpha) * previous_duals[key]
        for key in raw_duals
    }


def dual_shift_ratio(
    raw_duals: dict[Any, float],
    previous_raw_duals: dict[Any, float] | None,
) -> float:
    if previous_raw_duals is None or not raw_duals:
        return 0.0
    average_shift = sum(
        abs(raw_duals[key] - previous_raw_duals[key]) for key in raw_duals
    ) / len(raw_duals)
    average_magnitude = sum(abs(previous_raw_duals[key]) for key in previous_raw_duals) / len(
        previous_raw_duals
    )
    return average_shift / max(average_magnitude, 1.0)


def adaptive_stabilization_alpha(
    base_alpha: float,
    raw_cover_duals: dict[int, float],
    previous_raw_cover_duals: dict[int, float] | None,
    raw_type_duals: dict[str, float],
    previous_raw_type_duals: dict[str, float] | None,
    mode: str,
) -> float:
    if mode == "fixed" or base_alpha >= 1.0 - 1e-12:
        return base_alpha

    cover_shift = dual_shift_ratio(raw_cover_duals, previous_raw_cover_duals)
    type_shift = dual_shift_ratio(raw_type_duals, previous_raw_type_duals)
    weighted_shift = (
        cover_shift * len(raw_cover_duals) + type_shift * len(raw_type_duals)
    ) / max(len(raw_cover_duals) + len(raw_type_duals), 1)

    spread = min(base_alpha, 1.0 - base_alpha, 0.25)
    min_alpha = max(0.05, base_alpha - spread)
    max_alpha = min(1.0, base_alpha + spread)
    if weighted_shift <= 0.05:
        return max_alpha
    if weighted_shift >= 0.35:
        return min_alpha

    position = (weighted_shift - 0.05) / 0.30
    return max_alpha - position * (max_alpha - min_alpha)


def route_arc_set(stop_sequence: tuple[int, ...]) -> frozenset[tuple[int, int]]:
    if not stop_sequence:
        return frozenset()

    arcs: list[tuple[int, int]] = [(DEPOT_NODE, stop_sequence[0])]
    for idx in range(len(stop_sequence) - 1):
        arcs.append((stop_sequence[idx], stop_sequence[idx + 1]))
    arcs.append((stop_sequence[-1], DEPOT_NODE))
    return frozenset(arcs)


def route_reduced_cost(
    route: RouteColumn,
    cover_duals: dict[int, float],
    type_dual: float,
) -> float:
    return route.cost - sum(cover_duals[stop_id] for stop_id in route.stop_sequence) - type_dual


def route_column_key(route: RouteColumn) -> tuple[str, tuple[int, ...]]:
    return route.type_id, route.stop_sequence


def select_initial_active_routes(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    compatible_routes: list[RouteColumn],
) -> list[RouteColumn]:
    if not compatible_routes:
        return []

    active_keys: set[tuple[str, tuple[int, ...]]] = set()
    routes_by_type: dict[str, list[RouteColumn]] = {type_id: [] for type_id in vehicle_types}
    routes_by_stop: dict[int, list[RouteColumn]] = {stop_id: [] for stop_id in instance.stops}

    for route in compatible_routes:
        routes_by_type[route.type_id].append(route)
        for stop_id in route.stop_set:
            routes_by_stop[stop_id].append(route)

    for stop_id, covering_routes in routes_by_stop.items():
        for route in sorted(
            covering_routes,
            key=lambda candidate: (
                candidate.cost,
                len(candidate.stop_sequence),
                candidate.return_min,
                candidate.stop_sequence,
            ),
        )[:INITIAL_ACTIVE_COVER_CANDIDATES_PER_STOP]:
            active_keys.add(route_column_key(route))

    for type_id, typed_routes in routes_by_type.items():
        cheapest_routes = sorted(
            typed_routes,
            key=lambda candidate: (
                candidate.cost,
                len(candidate.stop_sequence),
                candidate.return_min,
                candidate.stop_sequence,
            ),
        )[:INITIAL_ACTIVE_CHEAPEST_ROUTES_PER_TYPE]
        for route in cheapest_routes:
            active_keys.add(route_column_key(route))

        for route in typed_routes[-INITIAL_ACTIVE_RECENT_ROUTES_PER_TYPE:]:
            active_keys.add(route_column_key(route))

    return [route for route in compatible_routes if route_column_key(route) in active_keys]


def select_inactive_reactivation_routes(
    inactive_routes_by_type: dict[str, dict[tuple[str, tuple[int, ...]], RouteColumn]],
    cover_duals: dict[int, float],
    type_duals: dict[str, float],
    max_per_type: int,
) -> list[RouteColumn]:
    selected_routes: list[RouteColumn] = []
    for type_id, inactive_routes in inactive_routes_by_type.items():
        improving_routes = [
            (
                route_reduced_cost(
                    route,
                    cover_duals=cover_duals,
                    type_dual=type_duals[type_id],
                ),
                route,
            )
            for route in inactive_routes.values()
        ]
        improving_routes = [
            (reduced_cost, route)
            for reduced_cost, route in improving_routes
            if reduced_cost < -REDUCED_COST_TOL
        ]
        improving_routes.sort(
            key=lambda item: (
                item[0],
                item[1].cost,
                len(item[1].stop_sequence),
                item[1].stop_sequence,
            )
        )
        selected_routes.extend(route for _, route in improving_routes[:max_per_type])
    return selected_routes


def root_bootstrap_time_limit(deadline: float | None) -> float | None:
    if deadline is None:
        return ROOT_BOOTSTRAP_MAX_SECONDS
    remaining = remaining_time(deadline)
    if remaining is None or remaining <= 0:
        return 0.0
    return min(
        ROOT_BOOTSTRAP_MAX_SECONDS,
        max(remaining * ROOT_BOOTSTRAP_TIME_FRACTION, 0.0),
    )


def bootstrap_root_route_pool(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    route_pool: list[RouteColumn],
    route_key_set: set[tuple[str, tuple[int, ...]]],
    branch_state: BranchState,
    deadline: float | None,
    progress: ProgressTracker | None = None,
) -> int:
    bootstrap_time_limit = root_bootstrap_time_limit(deadline)
    if (
        bootstrap_time_limit is None
        or bootstrap_time_limit <= 0.0
        or ROOT_BOOTSTRAP_MAX_ITERATIONS <= 0
        or ROOT_BOOTSTRAP_ROUTES_PER_TYPE <= 0
    ):
        return 0

    branch_ctx = build_branch_context(branch_state, instance, vehicle_types)
    if branch_ctx is None:
        return 0

    compatible_routes = [route for route in route_pool if route_is_compatible(route, branch_ctx)]
    if not compatible_routes:
        return 0

    route_sequences_by_type: dict[str, set[tuple[int, ...]]] = {
        type_id: set() for type_id in vehicle_types
    }
    for type_id, stop_sequence in route_key_set:
        if type_id in route_sequences_by_type:
            route_sequences_by_type[type_id].add(stop_sequence)

    bootstrap_deadline = perf_counter() + bootstrap_time_limit
    node_master = build_incremental_node_master(
        instance=instance,
        vehicle_types=vehicle_types,
        routes=compatible_routes,
    )
    added_route_count = 0

    if progress is not None:
        progress.emit(
            "root bootstrap: starting heuristic column harvest over "
            f"{len(compatible_routes)} columns for up to {bootstrap_time_limit:.1f}s",
            force=True,
        )

    for bootstrap_iteration in range(1, ROOT_BOOTSTRAP_MAX_ITERATIONS + 1):
        bootstrap_remaining = remaining_time(bootstrap_deadline)
        overall_remaining = remaining_time(deadline)
        effective_time_limit = bootstrap_remaining
        if overall_remaining is not None:
            effective_time_limit = (
                min(overall_remaining, bootstrap_remaining)
                if bootstrap_remaining is not None
                else overall_remaining
            )
        if effective_time_limit is not None and effective_time_limit <= 0:
            break

        if effective_time_limit is not None:
            node_master.model.Params.TimeLimit = effective_time_limit

        optimize_with_heartbeat(
            node_master.model,
            progress=progress,
            label=f"root bootstrap LP iteration {bootstrap_iteration}",
        )
        if node_master.model.Status != GRB.OPTIMAL:
            break

        cover_duals = {
            stop_id: node_master.cover_constraints[stop_id].Pi
            for stop_id in sorted(instance.stops)
        }
        type_duals = {
            type_id: node_master.type_constraints[type_id].Pi for type_id in vehicle_types
        }

        new_routes: list[RouteColumn] = []
        bootstrap_timed_out = False
        for vehicle_type in vehicle_types.values():
            try:
                candidate_routes = solve_pricing_subproblem(
                    instance=instance,
                    vehicle_type=vehicle_type,
                    cover_duals=cover_duals,
                    type_dual=type_duals[vehicle_type.type_id],
                    branch_ctx=branch_ctx,
                    max_routes=ROOT_BOOTSTRAP_ROUTES_PER_TYPE,
                    excluded_stop_sequences=route_sequences_by_type[vehicle_type.type_id],
                    pricing_time_limit=None,
                    deadline=bootstrap_deadline,
                    progress=progress,
                    progress_label=f"root bootstrap pricing {vehicle_type.type_id}",
                    exact_fallback_max_routes=ROOT_EXACT_PRICING_FALLBACK_MAX_ROUTES,
                )
            except TimeoutError:
                bootstrap_timed_out = True
                break
            for route in candidate_routes:
                if (
                    route_reduced_cost(
                        route,
                        cover_duals=cover_duals,
                        type_dual=type_duals[vehicle_type.type_id],
                    )
                    >= -REDUCED_COST_TOL
                ):
                    continue
                key = route_column_key(route)
                if key in route_key_set:
                    continue
                route_key_set.add(key)
                route_sequences_by_type[route.type_id].add(route.stop_sequence)
                route_pool.append(route)
                new_routes.append(route)
                add_route_column_to_master(node_master, route)

        if bootstrap_timed_out:
            if progress is not None:
                progress.emit(
                    "root bootstrap: stopped early after hitting the bootstrap time budget",
                    force=True,
                )
            break

        if not new_routes:
            if progress is not None:
                progress.emit(
                    f"root bootstrap: converged after {bootstrap_iteration} LP rounds "
                    "with no new heuristic routes",
                    force=True,
                )
            break

        added_route_count += len(new_routes)
        if progress is not None:
            progress.emit(
                f"root bootstrap: iteration {bootstrap_iteration} added {len(new_routes)} routes; "
                f"route pool now has {len(route_pool)} columns",
                force=True,
            )

    return added_route_count


def init_pricing_worker(instance: Instance, branch_ctx: BranchContext) -> None:
    global _PRICING_WORKER_INSTANCE, _PRICING_WORKER_BRANCH_CTX
    _PRICING_WORKER_INSTANCE = instance
    _PRICING_WORKER_BRANCH_CTX = branch_ctx


def run_pricing_worker(
    vehicle_type: VehicleTypeData,
    cover_duals: dict[int, float],
    type_dual: float,
    max_routes: int,
    excluded_stop_sequences: set[tuple[int, ...]],
    pricing_time_limit: float | None,
    deadline: float | None,
) -> tuple[str, list[RouteColumn]]:
    if _PRICING_WORKER_INSTANCE is None or _PRICING_WORKER_BRANCH_CTX is None:
        raise RuntimeError("Pricing worker was not initialized before solving pricing tasks.")
    return (
        vehicle_type.type_id,
        solve_pricing_subproblem(
            instance=_PRICING_WORKER_INSTANCE,
            vehicle_type=vehicle_type,
            cover_duals=cover_duals,
            type_dual=type_dual,
            branch_ctx=_PRICING_WORKER_BRANCH_CTX,
            max_routes=max_routes,
            excluded_stop_sequences=excluded_stop_sequences,
            pricing_time_limit=pricing_time_limit,
            deadline=deadline,
            progress=None,
            progress_label=None,
        ),
    )


def heuristic_pricing_seed_routes(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    cover_duals: dict[int, float],
    type_dual: float,
    branch_ctx: BranchContext,
    excluded_stop_sequences: set[tuple[int, ...]] | None,
) -> list[RouteColumn]:
    excluded_stop_sequences = excluded_stop_sequences or set()
    seeds: list[tuple[float, RouteColumn]] = []
    seen_sequences: set[tuple[int, ...]] = set()

    def maybe_add_seed(route: RouteColumn | None) -> None:
        if route is None:
            return
        if route.stop_sequence in excluded_stop_sequences:
            return
        if route.stop_sequence in seen_sequences:
            return
        if not route_is_compatible(route, branch_ctx):
            return
        seen_sequences.add(route.stop_sequence)
        seeds.append(
            (
                route_reduced_cost(
                    route,
                    cover_duals=cover_duals,
                    type_dual=type_dual,
                ),
                route,
            )
        )

    for candidate_type_name, route_sequences in CONSTRUCTIVE_ARC_FLOW_SEED_BY_VEHICLE_TYPE.items():
        if candidate_type_name != vehicle_type.vehicle_type:
            continue
        for stop_sequence in route_sequences:
            maybe_add_seed(evaluate_route_sequence(instance, vehicle_type, stop_sequence))

    for rule in REDSTONE_PLAZA_PAIR_RULES:
        for stop_sequence in ((rule.stop_a, rule.stop_b), (rule.stop_b, rule.stop_a)):
            maybe_add_seed(evaluate_route_sequence(instance, vehicle_type, stop_sequence))

    singleton_routes: list[tuple[float, RouteColumn]] = []
    for stop_id in sorted(
        instance.stops,
        key=lambda candidate_stop_id: (
            -cover_duals[candidate_stop_id],
            instance.stops[candidate_stop_id].latest_min,
            -instance.stops[candidate_stop_id].demand_kg,
            candidate_stop_id,
        ),
    ):
        route = evaluate_route_sequence(instance, vehicle_type, (stop_id,))
        if route is None or not route_is_compatible(route, branch_ctx):
            continue
        singleton_routes.append(
            (
                route_reduced_cost(
                    route,
                    cover_duals=cover_duals,
                    type_dual=type_dual,
                ),
                route,
            )
        )
    singleton_routes.sort(
        key=lambda item: (
            item[0],
            item[1].cost,
            len(item[1].stop_sequence),
            item[1].stop_sequence,
        )
    )
    for _, route in singleton_routes[:HEURISTIC_PRICING_SINGLETON_SEED_LIMIT]:
        maybe_add_seed(route)

    seeds.sort(
        key=lambda item: (
            item[0],
            item[1].cost,
            len(item[1].stop_sequence),
            item[1].stop_sequence,
        )
    )
    return [route for _, route in seeds]


def solve_heuristic_pricing_subproblem(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    cover_duals: dict[int, float],
    type_dual: float,
    branch_ctx: BranchContext,
    max_routes: int,
    excluded_stop_sequences: set[tuple[int, ...]] | None,
) -> list[RouteColumn]:
    if max_routes <= 0:
        return []

    excluded_stop_sequences = excluded_stop_sequences or set()
    candidate_stops = sorted(
        instance.stops,
        key=lambda stop_id: (
            -cover_duals[stop_id],
            instance.stops[stop_id].latest_min,
            -instance.stops[stop_id].demand_kg,
            stop_id,
        ),
    )
    seed_routes = heuristic_pricing_seed_routes(
        instance=instance,
        vehicle_type=vehicle_type,
        cover_duals=cover_duals,
        type_dual=type_dual,
        branch_ctx=branch_ctx,
        excluded_stop_sequences=excluded_stop_sequences,
    )
    if not seed_routes:
        return []

    improving_routes: list[tuple[float, RouteColumn]] = []
    seen_sequences: set[tuple[int, ...]] = set(excluded_stop_sequences)

    for seed_route in seed_routes:
        frontier: list[tuple[float, RouteColumn]] = [
            (
                route_reduced_cost(
                    seed_route,
                    cover_duals=cover_duals,
                    type_dual=type_dual,
                ),
                seed_route,
            )
        ]

        while frontier:
            next_frontier: list[tuple[float, RouteColumn]] = []
            for current_rc, current_route in frontier:
                current_sequence = current_route.stop_sequence
                if current_sequence not in seen_sequences and current_rc < -REDUCED_COST_TOL:
                    seen_sequences.add(current_sequence)
                    improving_routes.append((current_rc, current_route))
                    if len(improving_routes) >= max_routes:
                        break

                extension_candidates: list[tuple[float, RouteColumn]] = []
                for stop_id in candidate_stops:
                    if stop_id in current_route.stop_set:
                        continue
                    for pos in range(len(current_sequence) + 1):
                        candidate_sequence = (
                            current_sequence[:pos] + (stop_id,) + current_sequence[pos:]
                        )
                        if candidate_sequence in seen_sequences:
                            continue
                        candidate_route = evaluate_route_sequence(
                            instance,
                            vehicle_type,
                            candidate_sequence,
                        )
                        if candidate_route is None:
                            continue
                        if not route_is_compatible(candidate_route, branch_ctx):
                            continue
                        candidate_rc = route_reduced_cost(
                            candidate_route,
                            cover_duals=cover_duals,
                            type_dual=type_dual,
                        )
                        if candidate_rc < current_rc - REDUCED_COST_TOL:
                            extension_candidates.append((candidate_rc, candidate_route))

                extension_candidates.sort(
                    key=lambda item: (
                        item[0],
                        item[1].cost,
                        len(item[1].stop_sequence),
                        item[1].stop_sequence,
                    )
                )
                next_frontier.extend(extension_candidates[:HEURISTIC_PRICING_BEAM_WIDTH])

            if len(improving_routes) >= max_routes:
                break
            if not next_frontier:
                break

            frontier = []
            frontier_seen: set[tuple[int, ...]] = set()
            for candidate_rc, candidate_route in sorted(
                next_frontier,
                key=lambda item: (
                    item[0],
                    item[1].cost,
                    len(item[1].stop_sequence),
                    item[1].stop_sequence,
                ),
            ):
                if candidate_route.stop_sequence in frontier_seen:
                    continue
                frontier_seen.add(candidate_route.stop_sequence)
                frontier.append((candidate_rc, candidate_route))
                if len(frontier) >= HEURISTIC_PRICING_BEAM_WIDTH:
                    break

    improving_routes.sort(
        key=lambda item: (
            item[0],
            item[1].cost,
            len(item[1].stop_sequence),
            item[1].stop_sequence,
        )
    )
    return [route for _, route in improving_routes[:max_routes]]


def load_instance(workbook_path: Path) -> Instance:
    wb = load_workbook(workbook_path, data_only=True)

    ws_stops = wb["Delivery Stops"]
    ws_vehicles = wb["Vehicle Fleet"]
    ws_depot = wb["Depot"]
    ws_dist = wb["Distance Matrix (miles)"]
    ws_time = wb["Travel Time Matrix (min)"]
    ws_cost = wb["Cost Parameters"]

    stops: dict[int, StopData] = {}
    for row in range(2, ws_stops.max_row + 1):
        stop_id = int(ws_stops[f"A{row}"].value)
        stops[stop_id] = StopData(
            stop_id=stop_id,
            customer_name=str(ws_stops[f"B{row}"].value),
            address=str(ws_stops[f"C{row}"].value),
            demand_kg=float(ws_stops[f"D{row}"].value),
            earliest_min=excel_time_to_minutes(ws_stops[f"E{row}"].value),
            latest_min=excel_time_to_minutes(ws_stops[f"F{row}"].value),
            service_min=float(ws_stops[f"G{row}"].value),
        )
    vehicles: dict[int, VehicleData] = {}
    for row in range(2, ws_vehicles.max_row + 1):
        vehicle_id = int(ws_vehicles[f"A{row}"].value)
        vehicles[vehicle_id] = VehicleData(
            vehicle_id=vehicle_id,
            vehicle_type=str(ws_vehicles[f"B{row}"].value),
            capacity_kg=float(ws_vehicles[f"C{row}"].value),
            cost_per_mile=float(ws_vehicles[f"D{row}"].value),
            fixed_daily_cost=float(ws_vehicles[f"E{row}"].value),
            available_from_min=excel_time_to_minutes(ws_vehicles[f"F{row}"].value),
            available_until_min=excel_time_to_minutes(ws_vehicles[f"G{row}"].value),
        )

    depot_open_min = excel_time_to_minutes(ws_depot["C2"].value)
    depot_close_min = excel_time_to_minutes(ws_depot["D2"].value)

    node_ids = [DEPOT_NODE] + sorted(stops)
    expected_size = len(node_ids)
    if ws_dist.max_row != expected_size + 1 or ws_dist.max_column != expected_size + 1:
        raise ValueError(
            "Distance matrix dimensions do not match the number of nodes in the workbook."
        )
    if ws_time.max_row != expected_size + 1 or ws_time.max_column != expected_size + 1:
        raise ValueError(
            "Travel-time matrix dimensions do not match the number of nodes in the workbook."
        )

    distance_miles: dict[tuple[int, int], float] = {}
    travel_minutes: dict[tuple[int, int], float] = {}
    for row_offset, from_node in enumerate(node_ids, start=2):
        for col_offset, to_node in enumerate(node_ids, start=2):
            distance_miles[from_node, to_node] = float(ws_dist.cell(row_offset, col_offset).value)
            travel_minutes[from_node, to_node] = float(ws_time.cell(row_offset, col_offset).value)
    travel_minutes = apply_business_rule_travel_time_adjustments(travel_minutes)

    cost_params: dict[str, float] = {}
    for row in range(2, ws_cost.max_row + 1):
        key = str(ws_cost[f"A{row}"].value)
        cost_params[key] = float(ws_cost[f"B{row}"].value)

    return Instance(
        stops=stops,
        vehicles=vehicles,
        distance_miles=distance_miles,
        travel_minutes=travel_minutes,
        cost_params=cost_params,
        depot_open_min=depot_open_min,
        depot_close_min=depot_close_min,
    )


def format_stop_reference(stop: StopData) -> str:
    if stop.address:
        return f"Stop {stop.stop_id}: {stop.customer_name} | {stop.address}"
    return f"Stop {stop.stop_id}: {stop.customer_name}"


def format_stop_detail_label(stop: StopData) -> str:
    if stop.address:
        return f"Stop {stop.stop_id} ({stop.customer_name}, {stop.address})"
    return f"Stop {stop.stop_id} ({stop.customer_name})"


def format_route_stop_ids(route: tuple[int, ...]) -> str:
    stop_ids = [str(node) for node in route if node != DEPOT_NODE]
    return ", ".join(stop_ids) if stop_ids else "(empty)"


def format_route_stop_details(instance: Instance, route: tuple[int, ...]) -> str:
    stops = [
        format_stop_detail_label(instance.stops[node])
        for node in route
        if node != DEPOT_NODE
    ]
    return " -> ".join(stops) if stops else "(empty)"


def print_stop_reference(instance: Instance) -> None:
    print("Stop reference:")
    for stop_id in sorted(instance.stops):
        print(f"  {format_stop_reference(instance.stops[stop_id])}")
    print()


def validate_shared_cost_assumptions(instance: Instance) -> None:
    early_penalty = instance.cost_params.get("Early_Delivery_Penalty", 0.0)
    if abs(early_penalty) > 1e-9:
        raise NotImplementedError(
            "This solver currently assumes zero early-delivery penalty. "
            "That matches the workbook scenario, where Early_Delivery_Penalty = 0."
        )


def vehicle_business_rule_signature(vehicle_id: int) -> tuple[Any, ...]:
    """Split otherwise identical vehicles when hard driver/equipment rules differ."""

    exclusion_rule = VEHICLE_STOP_EXCLUSION_RULE_BY_VEHICLE.get(vehicle_id)
    service_adjustment_rule = VEHICLE_DEMAND_SERVICE_TIME_ADJUSTMENT_RULE_BY_VEHICLE.get(
        vehicle_id
    )
    return (
        None
        if exclusion_rule is None
        else tuple(sorted(exclusion_rule.forbidden_stops)),
        None
        if service_adjustment_rule is None
        else (
            service_adjustment_rule.demand_threshold_kg,
            service_adjustment_rule.extra_service_min,
        ),
    )


def build_vehicle_types(instance: Instance) -> dict[str, VehicleTypeData]:
    grouped: dict[tuple[Any, ...], list[int]] = defaultdict(list)
    for vehicle_id, vehicle in sorted(instance.vehicles.items()):
        key = (
            vehicle.vehicle_type,
            vehicle.capacity_kg,
            vehicle.cost_per_mile,
            vehicle.fixed_daily_cost,
            vehicle.available_from_min,
            vehicle.available_until_min,
            vehicle_business_rule_signature(vehicle_id),
        )
        grouped[key].append(vehicle_id)

    vehicle_types: dict[str, VehicleTypeData] = {}
    for idx, (key, vehicle_ids) in enumerate(
        sorted(grouped.items(), key=lambda item: min(item[1])), start=1
    ):
        (
            vehicle_type,
            capacity_kg,
            cost_per_mile,
            fixed_daily_cost,
            available_from_min,
            available_until_min,
            _business_rule_signature,
        ) = key
        type_id = f"T{idx}"
        vehicle_types[type_id] = VehicleTypeData(
            type_id=type_id,
            vehicle_ids=tuple(vehicle_ids),
            vehicle_type=str(vehicle_type),
            capacity_kg=float(capacity_kg),
            cost_per_mile=float(cost_per_mile),
            fixed_daily_cost=float(fixed_daily_cost),
            available_from_min=float(available_from_min),
            available_until_min=float(available_until_min),
            count=len(vehicle_ids),
        )
    return vehicle_types


def vehicle_start_limit(instance: Instance, vehicle_type: VehicleTypeData) -> float:
    return max(vehicle_type.available_from_min, instance.depot_open_min)


def vehicle_end_limit(instance: Instance, vehicle_type: VehicleTypeData) -> float:
    return min(vehicle_type.available_until_min, instance.depot_close_min)


def route_distance_unit_cost(instance: Instance, vehicle_type: VehicleTypeData) -> float:
    return vehicle_type.cost_per_mile + (
        instance.cost_params["Fuel_Cost_per_Liter"]
        * instance.cost_params["Avg_Fuel_Consumption_L_per_mile"]
    )


def billed_operating_hours(active_min: float) -> int:
    if active_min <= 1e-9:
        return 0
    return math.ceil(max(active_min - 1e-9, 0.0) / 60.0)


def operating_labor_cost(instance: Instance, active_min: float) -> float:
    return billed_operating_hours(active_min) * instance.cost_params["Driver_Hourly_Wage"]


def incremental_labor_cost(
    instance: Instance,
    elapsed_before_min: float,
    delta_min: float,
) -> float:
    if delta_min <= 0:
        return 0.0

    elapsed_after_min = max(elapsed_before_min + delta_min, 0.0)
    return operating_labor_cost(instance, elapsed_after_min) - operating_labor_cost(
        instance,
        max(elapsed_before_min, 0.0),
    )


def active_business_rule_descriptions() -> list[str]:
    descriptions: list[str] = []
    seen: set[str] = set()

    def add(description: str) -> None:
        if description in seen:
            return
        seen.add(description)
        descriptions.append(description)

    for rule in HARD_STOP_SERVICE_WINDOW_RULES:
        add(rule.description)
    for rule in REDSTONE_PLAZA_PAIR_RULES:
        add(rule.description)
    for rule in STOP_ALLOWED_VEHICLE_TYPE_RULES:
        add(rule.description)
    for rule in VEHICLE_STOP_EXCLUSION_RULES:
        add(rule.description)
    for rule in VEHICLE_DEMAND_SERVICE_TIME_ADJUSTMENT_RULES:
        add(rule.description)
    add(MAIN_STREET_CONSTRUCTION_RULE.description)
    if MIDDAY_LUNCH_BREAK_RULE is not None:
        add(MIDDAY_LUNCH_BREAK_RULE.description)
    return descriptions


def apply_business_rule_travel_time_adjustments(
    travel_minutes: dict[tuple[int, int], float]
) -> dict[tuple[int, int], float]:
    adjusted_travel_minutes = dict(travel_minutes)
    for (from_node, to_node), base_minutes in travel_minutes.items():
        if from_node == to_node:
            continue
        if (
            from_node in MAIN_STREET_CONSTRUCTION_RULE.affected_stops
            or to_node in MAIN_STREET_CONSTRUCTION_RULE.affected_stops
        ):
            adjusted_travel_minutes[from_node, to_node] = (
                base_minutes + MAIN_STREET_CONSTRUCTION_RULE.extra_minutes_per_leg
            )
    return adjusted_travel_minutes


def hard_service_start_lb(stop_id: int) -> float | None:
    rule = HARD_STOP_SERVICE_WINDOW_RULE_BY_STOP.get(stop_id)
    if rule is None:
        return None
    return rule.earliest_min


def hard_service_start_ub(stop_id: int) -> float | None:
    rule = HARD_STOP_SERVICE_WINDOW_RULE_BY_STOP.get(stop_id)
    if rule is None:
        return None
    return rule.latest_min


def hard_service_start_after_arrival(stop_id: int, arrival_min: float) -> float:
    earliest_min = hard_service_start_lb(stop_id)
    if earliest_min is None:
        return arrival_min
    return max(arrival_min, earliest_min)


def hard_service_start_is_feasible(stop_id: int, service_start_min: float) -> bool:
    latest_min = hard_service_start_ub(stop_id)
    return latest_min is None or service_start_min <= latest_min + 1e-9


def hard_service_finish_after_arrival(
    stop: StopData,
    arrival_min: float,
    service_min: float,
) -> float | None:
    service_start_min = hard_service_start_after_arrival(stop.stop_id, arrival_min)
    if not hard_service_start_is_feasible(stop.stop_id, service_start_min):
        return None
    return service_start_min + service_min


def vehicle_type_can_serve_stop(stop_id: int, vehicle_type_name: str) -> bool:
    rule = STOP_ALLOWED_VEHICLE_TYPE_RULE_BY_STOP.get(stop_id)
    return rule is None or vehicle_type_name in rule.allowed_vehicle_types


def vehicle_id_can_serve_stop(stop_id: int, vehicle_id: int) -> bool:
    rule = VEHICLE_STOP_EXCLUSION_RULE_BY_VEHICLE.get(vehicle_id)
    return rule is None or stop_id not in rule.forbidden_stops


def vehicle_data_can_serve_stop(stop_id: int, vehicle_id: int, vehicle: VehicleData) -> bool:
    return vehicle_type_can_serve_stop(stop_id, vehicle.vehicle_type) and vehicle_id_can_serve_stop(
        stop_id, vehicle_id
    )


def vehicle_type_data_can_serve_stop(stop_id: int, vehicle_type: VehicleTypeData) -> bool:
    return vehicle_type_can_serve_stop(stop_id, vehicle_type.vehicle_type) and all(
        vehicle_id_can_serve_stop(stop_id, vehicle_id) for vehicle_id in vehicle_type.vehicle_ids
    )


def extra_service_min_for_vehicle_id(
    instance: Instance,
    vehicle_id: int,
    stop_id: int,
) -> float:
    rule = VEHICLE_DEMAND_SERVICE_TIME_ADJUSTMENT_RULE_BY_VEHICLE.get(vehicle_id)
    if rule is None:
        return 0.0
    if instance.stops[stop_id].demand_kg <= rule.demand_threshold_kg + 1e-9:
        return 0.0
    return rule.extra_service_min


def extra_service_min_for_vehicle_type(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    stop_id: int,
) -> float:
    extras = {
        extra_service_min_for_vehicle_id(instance, vehicle_id, stop_id)
        for vehicle_id in vehicle_type.vehicle_ids
    }
    if len(extras) != 1:
        raise ValueError(
            "Vehicle type grouping mixed different service-time adjustment rules; "
            f"type {vehicle_type.type_id} contains vehicles {vehicle_type.vehicle_ids}."
        )
    return next(iter(extras))


def adjusted_stop_service_min_for_vehicle_id(
    instance: Instance,
    vehicle_id: int,
    stop_id: int,
) -> float:
    return instance.stops[stop_id].service_min + extra_service_min_for_vehicle_id(
        instance, vehicle_id, stop_id
    )


def adjusted_stop_service_min_for_vehicle_type(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    stop_id: int,
) -> float:
    return instance.stops[stop_id].service_min + extra_service_min_for_vehicle_type(
        instance, vehicle_type, stop_id
    )


def paired_stop_rule_for_arc(
    stop_id: int,
    next_node: int,
) -> RequiredAdjacentStopPairRule | None:
    for rule in REDSTONE_PLAZA_PAIR_RULES:
        if (stop_id, next_node) in rule.directed_arcs:
            return rule
    return None


def service_min_before_arc_for_vehicle_id(
    instance: Instance,
    vehicle_id: int,
    stop_id: int,
    next_node: int,
) -> float:
    rule = paired_stop_rule_for_arc(stop_id, next_node)
    if rule is None:
        return adjusted_stop_service_min_for_vehicle_id(instance, vehicle_id, stop_id)
    paired_total = (
        rule.combined_service_min
        + extra_service_min_for_vehicle_id(instance, vehicle_id, rule.stop_a)
        + extra_service_min_for_vehicle_id(instance, vehicle_id, rule.stop_b)
    )
    next_service_min = adjusted_stop_service_min_for_vehicle_id(instance, vehicle_id, next_node)
    return max(paired_total - next_service_min, 0.0)


def service_min_before_arc_for_vehicle_type(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    stop_id: int,
    next_node: int,
) -> float:
    rule = paired_stop_rule_for_arc(stop_id, next_node)
    if rule is None:
        return adjusted_stop_service_min_for_vehicle_type(instance, vehicle_type, stop_id)
    paired_total = (
        rule.combined_service_min
        + extra_service_min_for_vehicle_type(instance, vehicle_type, rule.stop_a)
        + extra_service_min_for_vehicle_type(instance, vehicle_type, rule.stop_b)
    )
    next_service_min = adjusted_stop_service_min_for_vehicle_type(
        instance, vehicle_type, next_node
    )
    return max(paired_total - next_service_min, 0.0)


def route_service_min_for_vehicle_id(
    instance: Instance,
    vehicle_id: int,
    route: tuple[int, ...],
) -> float:
    served_stops = [stop_id for stop_id in route if stop_id != DEPOT_NODE]
    if not served_stops:
        return 0.0
    total_service_min = 0.0
    for idx, stop_id in enumerate(served_stops):
        next_node = served_stops[idx + 1] if idx + 1 < len(served_stops) else DEPOT_NODE
        total_service_min += service_min_before_arc_for_vehicle_id(
            instance, vehicle_id, stop_id, next_node
        )
    return total_service_min


def lunch_break_interval_after(current_time_min: float) -> tuple[float, float] | None:
    if MIDDAY_LUNCH_BREAK_RULE is None:
        return None
    break_start = max(current_time_min, MIDDAY_LUNCH_BREAK_RULE.window_start_min)
    break_end = break_start + MIDDAY_LUNCH_BREAK_RULE.duration_min
    if break_end > MIDDAY_LUNCH_BREAK_RULE.window_end_min + 1e-9:
        return None
    return break_start, break_end


def route_overlaps_interval(
    route_start_min: float,
    route_end_min: float,
    interval_start_min: float,
    interval_end_min: float,
) -> bool:
    return (
        route_start_min < interval_end_min - 1e-9
        and route_end_min > interval_start_min + 1e-9
    )


def stop_sequence_respects_required_adjacencies(stop_sequence: tuple[int, ...]) -> bool:
    stop_positions = {stop_id: idx for idx, stop_id in enumerate(stop_sequence)}
    for rule in REDSTONE_PLAZA_PAIR_RULES:
        idx_a = stop_positions.get(rule.stop_a)
        idx_b = stop_positions.get(rule.stop_b)
        if (idx_a is None) != (idx_b is None):
            return False
        if idx_a is not None and abs(idx_a - idx_b) != 1:
            return False
    return True


def route_respects_required_adjacencies(route: RouteColumn) -> bool:
    for rule in REDSTONE_PLAZA_PAIR_RULES:
        has_a = rule.stop_a in route.stop_set
        has_b = rule.stop_b in route.stop_set
        if has_a != has_b:
            return False
        if has_a and not any(arc in route.arc_set for arc in rule.directed_arcs):
            return False
    return True


def simulate_stop_sequence_timing(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    start_limit: float,
    stop_sequence: tuple[int, ...],
    lunch_break_after_stop_count: int | None,
) -> RouteTimingEvaluation | None:
    current_time = start_limit
    current_node = DEPOT_NODE
    service_start_min: dict[int, float] = {}
    late_min: dict[int, float] = {}
    lunch_break_start_min: float | None = None

    def maybe_take_lunch(completed_stops: int) -> bool:
        nonlocal current_time, lunch_break_start_min
        if lunch_break_after_stop_count != completed_stops:
            return True
        lunch_window = lunch_break_interval_after(current_time)
        if lunch_window is None:
            return False
        lunch_break_start_min, current_time = lunch_window
        return True

    if not maybe_take_lunch(0):
        return None

    for idx, stop_id in enumerate(stop_sequence):
        stop = instance.stops[stop_id]
        next_node = stop_sequence[idx + 1] if idx + 1 < len(stop_sequence) else DEPOT_NODE
        service_min = service_min_before_arc_for_vehicle_type(
            instance, vehicle_type, stop_id, next_node
        )
        arrival_min = current_time + instance.travel_minutes[current_node, stop_id]
        service_start_min[stop_id] = hard_service_start_after_arrival(stop_id, arrival_min)
        if not hard_service_start_is_feasible(stop_id, service_start_min[stop_id]):
            return None
        late_min[stop_id] = max(service_start_min[stop_id] - stop.latest_min, 0.0)
        current_time = service_start_min[stop_id] + service_min
        current_node = stop_id
        if not maybe_take_lunch(idx + 1):
            return None

    current_time += instance.travel_minutes[current_node, DEPOT_NODE]
    return RouteTimingEvaluation(
        service_start_min=service_start_min,
        late_min=late_min,
        return_min=current_time,
        lunch_break_start_min=lunch_break_start_min,
    )


def evaluate_route_sequence(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    stop_sequence: tuple[int, ...],
) -> RouteColumn | None:
    if not stop_sequence:
        return None
    if not stop_sequence_respects_required_adjacencies(stop_sequence):
        return None
    if any(
        not vehicle_type_data_can_serve_stop(stop_id, vehicle_type)
        for stop_id in stop_sequence
    ):
        return None

    start_limit = vehicle_start_limit(instance, vehicle_type)
    end_limit = vehicle_end_limit(instance, vehicle_type)
    load_kg = sum(instance.stops[s].demand_kg for s in stop_sequence)
    if load_kg > vehicle_type.capacity_kg + 1e-9:
        return None

    current_node = DEPOT_NODE
    distance_mi = 0.0
    for stop_id in stop_sequence:
        distance_mi += instance.distance_miles[current_node, stop_id]
        current_node = stop_id

    distance_mi += instance.distance_miles[current_node, DEPOT_NODE]
    route_distance_cost = distance_mi * route_distance_unit_cost(instance, vehicle_type)

    baseline_timing = simulate_stop_sequence_timing(
        instance=instance,
        vehicle_type=vehicle_type,
        start_limit=start_limit,
        stop_sequence=stop_sequence,
        lunch_break_after_stop_count=None,
    )
    if baseline_timing is None:
        return None

    candidate_timings: list[RouteTimingEvaluation] = []
    lunch_break_required = MIDDAY_LUNCH_BREAK_RULE is not None
    if lunch_break_required:
        for break_after_stop_count in range(len(stop_sequence) + 1):
            timing = simulate_stop_sequence_timing(
                instance=instance,
                vehicle_type=vehicle_type,
                start_limit=start_limit,
                stop_sequence=stop_sequence,
                lunch_break_after_stop_count=break_after_stop_count,
            )
            if timing is None or timing.return_min > end_limit + 1e-9:
                continue
            candidate_timings.append(timing)
    elif baseline_timing.return_min <= end_limit + 1e-9:
        candidate_timings.append(baseline_timing)

    if not candidate_timings:
        return None

    best_route: RouteColumn | None = None
    best_score: tuple[float, float, float] | None = None
    for timing in candidate_timings:
        active_min = timing.return_min - start_limit
        cost = vehicle_type.fixed_daily_cost + route_distance_cost
        cost += operating_labor_cost(instance, active_min)
        cost += (
            sum(timing.late_min.values())
            / 60.0
            * instance.cost_params["Late_Delivery_Penalty_per_Hour"]
        )
        score = (cost, timing.return_min, sum(timing.late_min.values()))
        if best_score is not None and score >= best_score:
            continue
        best_score = score
        best_route = RouteColumn(
            type_id=vehicle_type.type_id,
            vehicle_type=vehicle_type.vehicle_type,
            stop_sequence=stop_sequence,
            stop_set=frozenset(stop_sequence),
            arc_set=route_arc_set(stop_sequence),
            cost=cost,
            load_kg=load_kg,
            distance_mi=distance_mi,
            active_min=active_min,
            return_min=timing.return_min,
            service_start_min=timing.service_start_min,
            late_min=timing.late_min,
            lunch_break_start_min=timing.lunch_break_start_min,
        )

    return best_route


def build_constructive_arc_flow_seed(
    instance: Instance,
) -> list[tuple[int, RouteColumn, VehicleTypeData]] | None:
    vehicle_types = build_vehicle_types(instance)
    vehicle_types_by_name: dict[str, list[VehicleTypeData]] = defaultdict(list)
    for vehicle_type in vehicle_types.values():
        vehicle_types_by_name[vehicle_type.vehicle_type].append(vehicle_type)
    for candidate_types in vehicle_types_by_name.values():
        candidate_types.sort(key=lambda candidate: min(candidate.vehicle_ids))

    available_vehicle_ids: dict[str, list[int]] = {
        vehicle_type.type_id: list(vehicle_type.vehicle_ids)
        for vehicle_type in vehicle_types.values()
    }
    assignments: list[tuple[int, RouteColumn, VehicleTypeData]] = []
    covered_stops: list[int] = []

    for vehicle_type_name, route_sequences in CONSTRUCTIVE_ARC_FLOW_SEED_BY_VEHICLE_TYPE.items():
        candidate_types = vehicle_types_by_name.get(vehicle_type_name, [])
        if not candidate_types:
            return None
        for stop_sequence in route_sequences:
            chosen_route: RouteColumn | None = None
            chosen_vehicle_type: VehicleTypeData | None = None
            for candidate_type in candidate_types:
                if not available_vehicle_ids[candidate_type.type_id]:
                    continue
                route = evaluate_route_sequence(instance, candidate_type, stop_sequence)
                if route is None:
                    continue
                if sum(route.late_min.values()) > 1e-6:
                    continue
                chosen_route = route
                chosen_vehicle_type = candidate_type
                break
            if chosen_route is None or chosen_vehicle_type is None:
                return None
            vehicle_id = available_vehicle_ids[chosen_vehicle_type.type_id].pop(0)
            assignments.append((vehicle_id, chosen_route, chosen_vehicle_type))
            covered_stops.extend(stop_sequence)

    if len(covered_stops) != len(instance.stops):
        return None
    if set(covered_stops) != set(instance.stops):
        return None
    return assignments


def apply_constructive_arc_flow_mip_start(
    instance: Instance,
    data: dict[str, Any],
    *,
    model: gp.Model | None = None,
    start_number: int | None = None,
) -> bool:
    assignments = build_constructive_arc_flow_seed(instance)
    if assignments is None:
        return False

    x = data["x"]
    visit = data["visit"]
    use_vehicle = data["use_vehicle"]
    depart_min = data["depart_min"]
    return_min = data["return_min"]
    service_start_min = data["service_start_min"]
    late_min = data["late_min"]
    billed_hours = data["billed_hours"]
    load_after_kg = data["load_after_kg"]
    lunch_required = data["lunch_required"]
    lunch_start_min = data["lunch_start_min"]
    served_after_lunch = data["served_after_lunch"]
    served_before_lunch = data["served_before_lunch"]
    lunch_before_start_arc = data["lunch_before_start_arc"]
    lunch_before_stop_arc = data["lunch_before_stop_arc"]
    stop_ids: list[int] = data["stop_ids"]
    previous_start_number: int | None = None

    if start_number is not None:
        if model is None:
            raise ValueError("model is required when applying a constructive MIP start slot")
        previous_start_number = int(model.Params.StartNumber)
        model.Params.StartNumber = start_number

    for var in x.values():
        var.Start = 0.0
    for var in visit.values():
        var.Start = 0.0
    for var in use_vehicle.values():
        var.Start = 0.0
    for var in depart_min.values():
        var.Start = 0.0
    for var in return_min.values():
        var.Start = 0.0
    for var in service_start_min.values():
        var.Start = 0.0
    for var in late_min.values():
        var.Start = 0.0
    for var in billed_hours.values():
        var.Start = 0.0
    for var in load_after_kg.values():
        var.Start = 0.0
    for var in lunch_required.values():
        var.Start = 0.0
    for var in lunch_start_min.values():
        var.Start = 0.0
    for var in served_after_lunch.values():
        var.Start = 0.0
    for var in served_before_lunch.values():
        var.Start = 0.0
    for var in lunch_before_start_arc.values():
        var.Start = 0.0
    for var in lunch_before_stop_arc.values():
        var.Start = 0.0

    assigned_service_starts: dict[int, float] = {}
    assigned_late: dict[int, float] = {}

    for vehicle_id, route, vehicle_type in assignments:
        start_min = vehicle_start_limit(instance, vehicle_type)
        route_with_depot = (DEPOT_NODE,) + route.stop_sequence + (DEPOT_NODE,)
        use_vehicle[vehicle_id].Start = 1.0
        depart_min[vehicle_id].Start = start_min
        return_min[vehicle_id].Start = route.return_min
        billed_hours[vehicle_id].Start = float(
            billed_operating_hours(route.return_min - start_min)
        )
        if route.lunch_break_start_min is not None:
            lunch_required[vehicle_id].Start = 1.0
            lunch_start_min[vehicle_id].Start = route.lunch_break_start_min
        cumulative_load_kg = 0.0
        for stop_id in route.stop_sequence:
            visit[vehicle_id, stop_id].Start = 1.0
            assigned_service_starts[stop_id] = route.service_start_min[stop_id]
            assigned_late[stop_id] = route.late_min[stop_id]
            cumulative_load_kg += instance.stops[stop_id].demand_kg
            load_after_kg[vehicle_id, stop_id].Start = cumulative_load_kg
        lunch_end_min = (
            None
            if route.lunch_break_start_min is None
            else route.lunch_break_start_min + MIDDAY_LUNCH_BREAK_RULE.duration_min
        )
        served_after_lunch_stops: set[int] = set()
        for stop_id in route.stop_sequence:
            if lunch_end_min is not None and route.service_start_min[stop_id] >= lunch_end_min - 1e-9:
                served_after_lunch[vehicle_id, stop_id].Start = 1.0
                served_after_lunch_stops.add(stop_id)
            else:
                served_before_lunch[vehicle_id, stop_id].Start = 1.0
        for i, j in zip(route_with_depot, route_with_depot[1:]):
            arc_key = (vehicle_id, i, j)
            arc_var = x.get(arc_key)
            if arc_var is None:
                return False
            arc_var.Start = 1.0
            if j in served_after_lunch_stops:
                if i == DEPOT_NODE:
                    lunch_before_start_arc[vehicle_id, j].Start = 1.0
                elif j != DEPOT_NODE:
                    lunch_before_stop_arc[vehicle_id, i, j].Start = 1.0

    if set(assigned_service_starts) != set(stop_ids):
        if model is not None and previous_start_number is not None:
            model.Params.StartNumber = previous_start_number
        return False
    for stop_id in stop_ids:
        service_start_min[stop_id].Start = assigned_service_starts[stop_id]
        late_min[stop_id].Start = assigned_late[stop_id]

    if model is not None and previous_start_number is not None:
        model.Params.StartNumber = previous_start_number
    return True


def build_arc_flow_model(instance: Instance) -> tuple[gp.Model, dict[str, Any]]:
    validate_shared_cost_assumptions(instance)

    model = create_gurobi_model("arc_flow_vrp")

    stop_ids = sorted(instance.stops)
    vehicle_ids = sorted(instance.vehicles)
    node_ids = [DEPOT_NODE] + stop_ids
    node_pairs = [(i, j) for i in node_ids for j in node_ids if i != j]
    total_arc_count = len(vehicle_ids) * len(node_pairs)
    identical_vehicle_groups: dict[tuple[Any, ...], list[int]] = {}
    for vehicle_id in vehicle_ids:
        vehicle = instance.vehicles[vehicle_id]
        key = (
            vehicle.vehicle_type,
            vehicle.capacity_kg,
            vehicle.cost_per_mile,
            vehicle.fixed_daily_cost,
            vehicle.available_from_min,
            vehicle.available_until_min,
            vehicle_business_rule_signature(vehicle_id),
        )
        identical_vehicle_groups.setdefault(key, []).append(vehicle_id)

    start_limit = {
        vehicle_id: max(instance.vehicles[vehicle_id].available_from_min, instance.depot_open_min)
        for vehicle_id in vehicle_ids
    }
    end_limit = {
        vehicle_id: min(instance.vehicles[vehicle_id].available_until_min, instance.depot_close_min)
        for vehicle_id in vehicle_ids
    }
    route_horizon = {
        vehicle_id: max(end_limit[vehicle_id] - start_limit[vehicle_id], 0.0)
        for vehicle_id in vehicle_ids
    }
    lunch_duration_min = (
        MIDDAY_LUNCH_BREAK_RULE.duration_min if MIDDAY_LUNCH_BREAK_RULE is not None else 0.0
    )
    lunch_required_by_vehicle = {
        vehicle_id: (
            MIDDAY_LUNCH_BREAK_RULE is not None
            and start_limit[vehicle_id] < MIDDAY_LUNCH_BREAK_RULE.window_end_min - 1e-9
        )
        for vehicle_id in vehicle_ids
    }
    lunch_start_ub_by_vehicle = {
        vehicle_id: (
            MIDDAY_LUNCH_BREAK_RULE.latest_start_min
            if lunch_required_by_vehicle[vehicle_id]
            else 0.0
        )
        for vehicle_id in vehicle_ids
    }
    vehicle_stop_lb: dict[tuple[int, int], float] = {}
    vehicle_stop_ub: dict[tuple[int, int], float] = {}
    feasible_vehicle_stop: dict[tuple[int, int], bool] = {}
    min_round_trip: dict[int, float] = {}

    for vehicle_id in vehicle_ids:
        vehicle = instance.vehicles[vehicle_id]
        feasible_round_trips: list[float] = []
        for stop_id in stop_ids:
            stop = instance.stops[stop_id]
            service_to_depot_min = service_min_before_arc_for_vehicle_id(
                instance, vehicle_id, stop_id, DEPOT_NODE
            )
            lb = start_limit[vehicle_id] + instance.travel_minutes[DEPOT_NODE, stop_id]
            hard_lb = hard_service_start_lb(stop_id)
            if hard_lb is not None:
                lb = max(lb, hard_lb)
            ub = (
                end_limit[vehicle_id]
                - service_to_depot_min
                - instance.travel_minutes[stop_id, DEPOT_NODE]
            )
            hard_ub = hard_service_start_ub(stop_id)
            if hard_ub is not None:
                ub = min(ub, hard_ub)
            feasible = (
                vehicle_data_can_serve_stop(stop_id, vehicle_id, vehicle)
                and stop.demand_kg <= vehicle.capacity_kg + 1e-9
                and lb <= ub + 1e-9
            )
            vehicle_stop_lb[vehicle_id, stop_id] = lb
            vehicle_stop_ub[vehicle_id, stop_id] = ub
            feasible_vehicle_stop[vehicle_id, stop_id] = feasible
            if feasible:
                feasible_round_trips.append(
                    (lb - start_limit[vehicle_id])
                    + service_to_depot_min
                    + instance.travel_minutes[stop_id, DEPOT_NODE]
                )
        min_round_trip[vehicle_id] = min(feasible_round_trips) if feasible_round_trips else 0.0

    service_start_lb: dict[int, float] = {}
    service_start_ub: dict[int, float] = {}
    late_ub: dict[int, float] = {}
    for stop_id in stop_ids:
        feasible_vehicles = [
            vehicle_id for vehicle_id in vehicle_ids if feasible_vehicle_stop[vehicle_id, stop_id]
        ]
        if not feasible_vehicles:
            raise ValueError(
                f"Stop {stop_id} cannot be served by any vehicle under the workbook data."
            )
        service_start_lb[stop_id] = min(
            vehicle_stop_lb[vehicle_id, stop_id] for vehicle_id in feasible_vehicles
        )
        service_start_ub[stop_id] = max(
            vehicle_stop_ub[vehicle_id, stop_id] for vehicle_id in feasible_vehicles
        )
        late_ub[stop_id] = max(
            service_start_ub[stop_id] - instance.stops[stop_id].latest_min,
            0.0,
        )

    feasible_arc: dict[tuple[int, int, int], bool] = {}
    for vehicle_id in vehicle_ids:
        for stop_id in stop_ids:
            can_serve = feasible_vehicle_stop[vehicle_id, stop_id]
            feasible_arc[vehicle_id, DEPOT_NODE, stop_id] = can_serve
            feasible_arc[vehicle_id, stop_id, DEPOT_NODE] = can_serve
            for next_stop in stop_ids:
                if stop_id == next_stop:
                    continue
                if not can_serve or not feasible_vehicle_stop[vehicle_id, next_stop]:
                    feasible_arc[vehicle_id, stop_id, next_stop] = False
                    continue
                stop_to_next_service_min = service_min_before_arc_for_vehicle_id(
                    instance, vehicle_id, stop_id, next_stop
                )
                feasible_arc[vehicle_id, stop_id, next_stop] = (
                    vehicle_stop_lb[vehicle_id, stop_id]
                    + stop_to_next_service_min
                    + instance.travel_minutes[stop_id, next_stop]
                    <= vehicle_stop_ub[vehicle_id, next_stop] + 1e-9
                )

    feasible_arc_keys = [
        (vehicle_id, i, j)
        for vehicle_id in vehicle_ids
        for i, j in node_pairs
        if feasible_arc.get((vehicle_id, i, j), False)
    ]
    feasible_arc_keys_by_vehicle: dict[int, list[tuple[int, int]]] = {
        vehicle_id: [] for vehicle_id in vehicle_ids
    }
    incoming_arc_keys_by_stop: dict[tuple[int, int], list[tuple[int, int, int]]] = {}
    outgoing_arc_keys_by_stop: dict[tuple[int, int], list[tuple[int, int, int]]] = {}
    start_arc_keys_by_vehicle: dict[int, list[tuple[int, int, int]]] = {
        vehicle_id: [] for vehicle_id in vehicle_ids
    }
    return_arc_keys_by_vehicle: dict[int, list[tuple[int, int, int]]] = {
        vehicle_id: [] for vehicle_id in vehicle_ids
    }
    for arc_key in feasible_arc_keys:
        vehicle_id, i, j = arc_key
        feasible_arc_keys_by_vehicle[vehicle_id].append((i, j))
        if i == DEPOT_NODE:
            start_arc_keys_by_vehicle[vehicle_id].append(arc_key)
        else:
            outgoing_arc_keys_by_stop.setdefault((vehicle_id, i), []).append(arc_key)
        if j == DEPOT_NODE:
            return_arc_keys_by_vehicle[vehicle_id].append(arc_key)
        else:
            incoming_arc_keys_by_stop.setdefault((vehicle_id, j), []).append(arc_key)
    lunch_before_start_arc_keys = [
        (vehicle_id, stop_id)
        for vehicle_id, from_node, stop_id in feasible_arc_keys
        if MIDDAY_LUNCH_BREAK_RULE is not None and from_node == DEPOT_NODE
    ]
    lunch_before_stop_arc_keys = [
        (vehicle_id, stop_id, next_stop)
        for vehicle_id, stop_id, next_stop in feasible_arc_keys
        if MIDDAY_LUNCH_BREAK_RULE is not None
        and stop_id != DEPOT_NODE
        and next_stop != DEPOT_NODE
    ]

    x = model.addVars(feasible_arc_keys, vtype=GRB.BINARY, name="x")
    visit = model.addVars(vehicle_ids, stop_ids, vtype=GRB.BINARY, name="visit")
    use_vehicle = model.addVars(vehicle_ids, vtype=GRB.BINARY, name="use_vehicle")
    depart_min = model.addVars(vehicle_ids, lb=0.0, ub=start_limit, name="depart_min")
    return_min = model.addVars(vehicle_ids, lb=0.0, ub=end_limit, name="return_min")
    service_start_min = model.addVars(
        stop_ids,
        lb=service_start_lb,
        ub=service_start_ub,
        name="service_start_min",
    )
    late_min = model.addVars(stop_ids, lb=0.0, ub=late_ub, name="late_min")
    billed_hours = model.addVars(
        vehicle_ids,
        lb=0.0,
        ub={vehicle_id: billed_operating_hours(route_horizon[vehicle_id]) for vehicle_id in vehicle_ids},
        vtype=GRB.INTEGER,
        name="billed_hours",
    )
    load_after_kg = model.addVars(
        vehicle_ids,
        stop_ids,
        lb=0.0,
        ub={
            (vehicle_id, stop_id): (
                instance.vehicles[vehicle_id].capacity_kg
                if feasible_vehicle_stop[vehicle_id, stop_id]
                else 0.0
            )
            for vehicle_id in vehicle_ids
            for stop_id in stop_ids
        },
        name="load_after_kg",
    )
    lunch_required = model.addVars(vehicle_ids, vtype=GRB.BINARY, name="lunch_required")
    lunch_start_min = model.addVars(
        vehicle_ids,
        lb=0.0,
        ub=lunch_start_ub_by_vehicle,
        name="lunch_start_min",
    )
    served_after_lunch = model.addVars(
        vehicle_ids,
        stop_ids,
        vtype=GRB.BINARY,
        name="served_after_lunch",
    )
    served_before_lunch = model.addVars(
        vehicle_ids,
        stop_ids,
        vtype=GRB.BINARY,
        name="served_before_lunch",
    )
    lunch_before_start_arc = model.addVars(
        lunch_before_start_arc_keys,
        vtype=GRB.BINARY,
        name="lunch_before_start_arc",
    )
    lunch_before_stop_arc = model.addVars(
        lunch_before_stop_arc_keys,
        vtype=GRB.BINARY,
        name="lunch_before_stop_arc",
    )

    fixed_arc_count = total_arc_count - len(feasible_arc_keys)
    for vehicle_id in vehicle_ids:
        for stop_id in stop_ids:
            if not feasible_vehicle_stop[vehicle_id, stop_id]:
                visit[vehicle_id, stop_id].ub = 0.0
                served_before_lunch[vehicle_id, stop_id].ub = 0.0
                served_after_lunch[vehicle_id, stop_id].ub = 0.0

    incoming = {
        (vehicle_id, stop_id): gp.quicksum(
            x[arc_key] for arc_key in incoming_arc_keys_by_stop.get((vehicle_id, stop_id), [])
        )
        for vehicle_id in vehicle_ids
        for stop_id in stop_ids
    }
    outgoing = {
        (vehicle_id, stop_id): gp.quicksum(
            x[arc_key] for arc_key in outgoing_arc_keys_by_stop.get((vehicle_id, stop_id), [])
        )
        for vehicle_id in vehicle_ids
        for stop_id in stop_ids
    }
    starts = {
        vehicle_id: gp.quicksum(x[arc_key] for arc_key in start_arc_keys_by_vehicle[vehicle_id])
        for vehicle_id in vehicle_ids
    }
    returns = {
        vehicle_id: gp.quicksum(x[arc_key] for arc_key in return_arc_keys_by_vehicle[vehicle_id])
        for vehicle_id in vehicle_ids
    }
    first_customer_index_expr = {
        vehicle_id: gp.quicksum(
            stop_id * x[vehicle_id, DEPOT_NODE, stop_id]
            for _, _, stop_id in start_arc_keys_by_vehicle[vehicle_id]
        )
        for vehicle_id in vehicle_ids
    }
    drive_min_expr = {
        vehicle_id: gp.quicksum(
            instance.travel_minutes[i, j] * x[vehicle_id, i, j]
            for i, j in feasible_arc_keys_by_vehicle[vehicle_id]
        )
        for vehicle_id in vehicle_ids
    }
    service_after_stop_expr = {
        (vehicle_id, stop_id): gp.quicksum(
            service_min_before_arc_for_vehicle_id(instance, vehicle_id, stop_id, j)
            * x[vehicle_id, stop_id, j]
            for _, _, j in outgoing_arc_keys_by_stop.get((vehicle_id, stop_id), [])
        )
        for vehicle_id in vehicle_ids
        for stop_id in stop_ids
    }
    service_after_stop_ub = {
        (vehicle_id, stop_id): max(
            (
                service_min_before_arc_for_vehicle_id(instance, vehicle_id, stop_id, j)
                for _, _, j in outgoing_arc_keys_by_stop.get((vehicle_id, stop_id), [])
            ),
            default=0.0,
        )
        for vehicle_id in vehicle_ids
        for stop_id in stop_ids
    }
    service_after_stop_lb = {
        (vehicle_id, stop_id): min(
            (
                service_min_before_arc_for_vehicle_id(instance, vehicle_id, stop_id, j)
                for _, _, j in outgoing_arc_keys_by_stop.get((vehicle_id, stop_id), [])
            ),
            default=0.0,
        )
        for vehicle_id in vehicle_ids
        for stop_id in stop_ids
    }
    service_min_expr = {
        vehicle_id: gp.quicksum(
            service_after_stop_expr[vehicle_id, stop_id]
            for stop_id in stop_ids
        )
        for vehicle_id in vehicle_ids
    }
    served_load_expr = {
        vehicle_id: gp.quicksum(
            instance.stops[stop_id].demand_kg * visit[vehicle_id, stop_id]
            for stop_id in stop_ids
        )
        for vehicle_id in vehicle_ids
    }
    total_demand_kg = sum(instance.stops[stop_id].demand_kg for stop_id in stop_ids)
    active_min_expr = {
        vehicle_id: return_min[vehicle_id] - depart_min[vehicle_id] for vehicle_id in vehicle_ids
    }
    distance_cost_expr = {
        vehicle_id: gp.quicksum(
            instance.distance_miles[i, j]
            * (
                instance.vehicles[vehicle_id].cost_per_mile
                + instance.cost_params["Fuel_Cost_per_Liter"]
                * instance.cost_params["Avg_Fuel_Consumption_L_per_mile"]
            )
            * x[vehicle_id, i, j]
            for i, j in feasible_arc_keys_by_vehicle[vehicle_id]
        )
        for vehicle_id in vehicle_ids
    }

    for stop_id in stop_ids:
        model.addConstr(
            gp.quicksum(visit[vehicle_id, stop_id] for vehicle_id in vehicle_ids) == 1,
            name=f"visit_once[{stop_id}]",
        )
        model.addConstr(
            late_min[stop_id] >= service_start_min[stop_id] - instance.stops[stop_id].latest_min,
            name=f"late_lb[{stop_id}]",
        )

    model.addConstr(
        gp.quicksum(
            instance.vehicles[vehicle_id].capacity_kg * use_vehicle[vehicle_id]
            for vehicle_id in vehicle_ids
        )
        >= total_demand_kg,
        name="global_capacity_cover",
    )

    # Hard operational rule: the Redstone Plaza deliveries must stay on the
    # same truck and be served consecutively, but the model may choose 7->8 or
    # 8->7. These assignment/arc constraints are much tighter than trying to
    # infer the rule from timing alone.
    for rule in REDSTONE_PLAZA_PAIR_RULES:
        if rule.stop_a not in stop_ids or rule.stop_b not in stop_ids:
            continue
        for vehicle_id in vehicle_ids:
            model.addConstr(
                visit[vehicle_id, rule.stop_a] == visit[vehicle_id, rule.stop_b],
                name=f"paired_stop_same_vehicle[{vehicle_id},{rule.stop_a},{rule.stop_b}]",
            )
            model.addConstr(
                x.get((vehicle_id, rule.stop_a, rule.stop_b), 0.0)
                + x.get((vehicle_id, rule.stop_b, rule.stop_a), 0.0)
                == visit[vehicle_id, rule.stop_a],
                name=f"paired_stop_adjacent[{vehicle_id},{rule.stop_a},{rule.stop_b}]",
            )

    symmetry_group_count = 0
    for group in identical_vehicle_groups.values():
        sorted_group = sorted(group)
        if len(sorted_group) <= 1:
            continue
        symmetry_group_count += 1
        for prev_vehicle_id, next_vehicle_id in zip(sorted_group, sorted_group[1:]):
            model.addConstr(
                use_vehicle[prev_vehicle_id] >= use_vehicle[next_vehicle_id],
                name=f"symmetry_use_order[{prev_vehicle_id},{next_vehicle_id}]",
            )
            model.addConstr(
                first_customer_index_expr[prev_vehicle_id]
                - first_customer_index_expr[next_vehicle_id]
                <= max(stop_ids) * (1 - use_vehicle[next_vehicle_id]),
                name=f"symmetry_first_customer[{prev_vehicle_id},{next_vehicle_id}]",
            )
            model.addConstr(
                served_load_expr[prev_vehicle_id] >= served_load_expr[next_vehicle_id],
                name=f"symmetry_load[{prev_vehicle_id},{next_vehicle_id}]",
            )

    for vehicle_id in vehicle_ids:
        vehicle = instance.vehicles[vehicle_id]

        model.addConstr(starts[vehicle_id] == use_vehicle[vehicle_id], name=f"use_vehicle_start[{vehicle_id}]")
        model.addConstr(returns[vehicle_id] == use_vehicle[vehicle_id], name=f"use_vehicle_return[{vehicle_id}]")
        model.addConstr(
            active_min_expr[vehicle_id] >= min_round_trip[vehicle_id] * use_vehicle[vehicle_id],
            name=f"minimum_route_time[{vehicle_id}]",
        )
        model.addConstr(
            active_min_expr[vehicle_id] >= drive_min_expr[vehicle_id] + service_min_expr[vehicle_id],
            name=f"drive_service_duration_lb[{vehicle_id}]",
        )

        for i, j in feasible_arc_keys_by_vehicle[vehicle_id]:
            model.addConstr(
                x[vehicle_id, i, j] <= use_vehicle[vehicle_id],
                name=f"arc_use_link[{vehicle_id},{i},{j}]",
            )

        for stop_id in stop_ids:
            model.addConstr(
                service_after_stop_expr[vehicle_id, stop_id]
                >= service_after_stop_lb[vehicle_id, stop_id] * visit[vehicle_id, stop_id],
                name=f"service_after_stop_lb[{vehicle_id},{stop_id}]",
            )
            model.addConstr(
                service_after_stop_expr[vehicle_id, stop_id]
                <= service_after_stop_ub[vehicle_id, stop_id] * visit[vehicle_id, stop_id],
                name=f"service_after_stop_ub[{vehicle_id},{stop_id}]",
            )
            model.addConstr(
                incoming[vehicle_id, stop_id] == visit[vehicle_id, stop_id],
                name=f"incoming_visit[{vehicle_id},{stop_id}]",
            )
            model.addConstr(
                outgoing[vehicle_id, stop_id] == visit[vehicle_id, stop_id],
                name=f"outgoing_visit[{vehicle_id},{stop_id}]",
            )
            model.addConstr(
                visit[vehicle_id, stop_id] <= use_vehicle[vehicle_id],
                name=f"visit_use_link[{vehicle_id},{stop_id}]",
            )

        model.addConstr(
            served_load_expr[vehicle_id] <= vehicle.capacity_kg * use_vehicle[vehicle_id],
            name=f"capacity[{vehicle_id}]",
        )
        model.addConstr(
            depart_min[vehicle_id] == start_limit[vehicle_id] * use_vehicle[vehicle_id],
            name=f"depart_fixed[{vehicle_id}]",
        )
        model.addConstr(
            return_min[vehicle_id] >= depart_min[vehicle_id],
            name=f"nonnegative_route_duration[{vehicle_id}]",
        )
        model.addConstr(
            return_min[vehicle_id] <= end_limit[vehicle_id] * use_vehicle[vehicle_id],
            name=f"return_ub[{vehicle_id}]",
        )
        model.addConstr(
            active_min_expr[vehicle_id] <= 60.0 * billed_hours[vehicle_id],
            name=f"billed_hours_lb[{vehicle_id}]",
        )
        model.addConstr(
            billed_hours[vehicle_id]
            <= billed_operating_hours(route_horizon[vehicle_id]) * use_vehicle[vehicle_id],
            name=f"billed_hours_cap[{vehicle_id}]",
        )
        model.addConstr(
            lunch_required[vehicle_id] == use_vehicle[vehicle_id],
            name=f"lunch_required_for_used_driver[{vehicle_id}]",
        )

        if (
            MIDDAY_LUNCH_BREAK_RULE is not None
            and start_limit[vehicle_id] < MIDDAY_LUNCH_BREAK_RULE.window_end_min - 1e-9
        ):
            model.addConstr(
                lunch_start_min[vehicle_id]
                >= MIDDAY_LUNCH_BREAK_RULE.window_start_min * lunch_required[vehicle_id],
                name=f"lunch_window_lb[{vehicle_id}]",
            )
            model.addConstr(
                lunch_start_min[vehicle_id]
                <= MIDDAY_LUNCH_BREAK_RULE.latest_start_min * lunch_required[vehicle_id],
                name=f"lunch_window_ub[{vehicle_id}]",
            )
            model.addConstr(
                lunch_start_min[vehicle_id]
                >= depart_min[vehicle_id],
                name=f"lunch_after_depart[{vehicle_id}]",
            )
            model.addConstr(
                lunch_start_min[vehicle_id]
                <= return_min[vehicle_id]
                - MIDDAY_LUNCH_BREAK_RULE.duration_min * lunch_required[vehicle_id],
                name=f"lunch_before_return[{vehicle_id}]",
            )
        else:
            lunch_required[vehicle_id].ub = 0.0

        for stop_id in stop_ids:
            stop = instance.stops[stop_id]
            model.addConstr(
                load_after_kg[vehicle_id, stop_id] >= stop.demand_kg * visit[vehicle_id, stop_id],
                name=f"load_lb[{vehicle_id},{stop_id}]",
            )
            model.addConstr(
                load_after_kg[vehicle_id, stop_id] <= vehicle.capacity_kg * visit[vehicle_id, stop_id],
                name=f"load_ub[{vehicle_id},{stop_id}]",
            )
            model.addConstr(
                served_before_lunch[vehicle_id, stop_id]
                + served_after_lunch[vehicle_id, stop_id]
                == visit[vehicle_id, stop_id],
                name=f"lunch_state_partition[{vehicle_id},{stop_id}]",
            )
            model.addConstr(
                served_after_lunch[vehicle_id, stop_id] <= lunch_required[vehicle_id],
                name=f"lunch_state_required_link[{vehicle_id},{stop_id}]",
            )
            if MIDDAY_LUNCH_BREAK_RULE is not None and feasible_vehicle_stop[vehicle_id, stop_id]:
                before_lunch_service_start_ub = (
                    MIDDAY_LUNCH_BREAK_RULE.latest_start_min
                    - service_after_stop_lb[vehicle_id, stop_id]
                )
                after_lunch_service_start_lb = (
                    MIDDAY_LUNCH_BREAK_RULE.window_start_min
                    + MIDDAY_LUNCH_BREAK_RULE.duration_min
                )
                if vehicle_stop_lb[vehicle_id, stop_id] > before_lunch_service_start_ub + 1e-9:
                    served_before_lunch[vehicle_id, stop_id].ub = 0.0
                else:
                    before_lunch_visit_tightening = (
                        service_start_ub[stop_id] - vehicle_stop_ub[vehicle_id, stop_id]
                    )
                    before_lunch_state_tightening = max(
                        vehicle_stop_ub[vehicle_id, stop_id]
                        - before_lunch_service_start_ub,
                        0.0,
                    )
                    if (
                        before_lunch_visit_tightening > 1e-9
                        or before_lunch_state_tightening > 1e-9
                    ):
                        model.addConstr(
                            service_start_min[stop_id]
                            <= service_start_ub[stop_id]
                            - before_lunch_visit_tightening * visit[vehicle_id, stop_id]
                            - before_lunch_state_tightening
                            * served_before_lunch[vehicle_id, stop_id],
                            name=f"before_lunch_service_start_ub[{vehicle_id},{stop_id}]",
                        )
                if vehicle_stop_ub[vehicle_id, stop_id] < after_lunch_service_start_lb - 1e-9:
                    served_after_lunch[vehicle_id, stop_id].ub = 0.0
                else:
                    after_lunch_visit_tightening = (
                        vehicle_stop_lb[vehicle_id, stop_id] - service_start_lb[stop_id]
                    )
                    after_lunch_state_tightening = max(
                        after_lunch_service_start_lb
                        - vehicle_stop_lb[vehicle_id, stop_id],
                        0.0,
                    )
                    if (
                        after_lunch_visit_tightening > 1e-9
                        or after_lunch_state_tightening > 1e-9
                    ):
                        model.addConstr(
                            service_start_min[stop_id]
                            >= service_start_lb[stop_id]
                            + after_lunch_visit_tightening * visit[vehicle_id, stop_id]
                            + after_lunch_state_tightening
                            * served_after_lunch[vehicle_id, stop_id],
                            name=f"after_lunch_service_start_lb[{vehicle_id},{stop_id}]",
                        )
            if MIDDAY_LUNCH_BREAK_RULE is not None:
                before_lunch_expr = (
                    service_start_min[stop_id]
                    + service_after_stop_expr[vehicle_id, stop_id]
                    - lunch_start_min[vehicle_id]
                )
                before_lunch_no_visit_ub = service_start_ub[stop_id]
                before_lunch_after_visit_ub = (
                    vehicle_stop_ub[vehicle_id, stop_id]
                    + service_after_stop_ub[vehicle_id, stop_id]
                    - MIDDAY_LUNCH_BREAK_RULE.window_start_min
                )
                model.addConstr(
                    before_lunch_expr
                    <= before_lunch_no_visit_ub * (1 - visit[vehicle_id, stop_id])
                    + before_lunch_after_visit_ub
                    * served_after_lunch[vehicle_id, stop_id],
                    name=f"served_before_lunch[{vehicle_id},{stop_id}]",
                )
                after_lunch_expr = (
                    service_start_min[stop_id]
                    - lunch_start_min[vehicle_id]
                    - MIDDAY_LUNCH_BREAK_RULE.duration_min
                )
                after_lunch_no_visit_lb = (
                    service_start_lb[stop_id]
                    - lunch_start_ub_by_vehicle[vehicle_id]
                    - MIDDAY_LUNCH_BREAK_RULE.duration_min
                )
                after_lunch_before_visit_lb = (
                    vehicle_stop_lb[vehicle_id, stop_id]
                    - lunch_start_ub_by_vehicle[vehicle_id]
                    - MIDDAY_LUNCH_BREAK_RULE.duration_min
                )
                model.addConstr(
                    after_lunch_expr
                    >= after_lunch_no_visit_lb * (1 - visit[vehicle_id, stop_id])
                    + after_lunch_before_visit_lb
                    * served_before_lunch[vehicle_id, stop_id],
                    name=f"served_after_lunch[{vehicle_id},{stop_id}]",
                )
            if feasible_vehicle_stop[vehicle_id, stop_id]:
                visit_lb_tightening = (
                    vehicle_stop_lb[vehicle_id, stop_id] - service_start_lb[stop_id]
                )
                if visit_lb_tightening > 1e-9:
                    model.addConstr(
                        service_start_min[stop_id]
                        >= service_start_lb[stop_id]
                        + visit_lb_tightening * visit[vehicle_id, stop_id],
                        name=f"visit_time_lb[{vehicle_id},{stop_id}]",
                    )
                visit_ub_tightening = (
                    service_start_ub[stop_id] - vehicle_stop_ub[vehicle_id, stop_id]
                )
                if visit_ub_tightening > 1e-9:
                    model.addConstr(
                        service_start_min[stop_id]
                        <= service_start_ub[stop_id]
                        - visit_ub_tightening * visit[vehicle_id, stop_id],
                        name=f"visit_time_ub[{vehicle_id},{stop_id}]",
                    )
            if feasible_arc[vehicle_id, DEPOT_NODE, stop_id]:
                time_from_depot_expr = (
                    service_start_min[stop_id]
                    - depart_min[vehicle_id]
                    - instance.travel_minutes[DEPOT_NODE, stop_id]
                )
                if MIDDAY_LUNCH_BREAK_RULE is not None:
                    time_from_depot_expr -= (
                        MIDDAY_LUNCH_BREAK_RULE.duration_min
                        * served_after_lunch[vehicle_id, stop_id]
                    )
                time_from_depot_lb = (
                    service_start_lb[stop_id]
                    - start_limit[vehicle_id]
                    - instance.travel_minutes[DEPOT_NODE, stop_id]
                    - lunch_duration_min
                )
                model.addConstr(
                    time_from_depot_expr
                    >= time_from_depot_lb
                    * (1 - x[vehicle_id, DEPOT_NODE, stop_id]),
                    name=f"time_from_depot[{vehicle_id},{stop_id}]",
                )
                load_from_depot_expr = load_after_kg[vehicle_id, stop_id] - stop.demand_kg
                model.addConstr(
                    load_from_depot_expr
                    <= (vehicle.capacity_kg - stop.demand_kg)
                    * (visit[vehicle_id, stop_id] - x[vehicle_id, DEPOT_NODE, stop_id])
                    - stop.demand_kg * (1 - visit[vehicle_id, stop_id]),
                    name=f"load_from_depot[{vehicle_id},{stop_id}]",
                )
                if MIDDAY_LUNCH_BREAK_RULE is not None:
                    model.addConstr(
                        lunch_before_start_arc[vehicle_id, stop_id]
                        <= x[vehicle_id, DEPOT_NODE, stop_id],
                        name=f"lunch_before_start_arc_x_ub[{vehicle_id},{stop_id}]",
                    )
                    model.addConstr(
                        lunch_before_start_arc[vehicle_id, stop_id]
                        <= served_after_lunch[vehicle_id, stop_id],
                        name=f"lunch_before_start_arc_state_ub[{vehicle_id},{stop_id}]",
                    )
                    model.addConstr(
                        lunch_before_start_arc[vehicle_id, stop_id]
                        >= x[vehicle_id, DEPOT_NODE, stop_id]
                        + served_after_lunch[vehicle_id, stop_id]
                        - 1,
                        name=f"lunch_before_start_arc_lb[{vehicle_id},{stop_id}]",
                    )
                    lunch_on_first_arc_expr = (
                        lunch_start_min[vehicle_id]
                        - service_start_min[stop_id]
                        + instance.travel_minutes[DEPOT_NODE, stop_id]
                        + MIDDAY_LUNCH_BREAK_RULE.duration_min
                    )
                    lunch_on_first_arc_no_after_ub = (
                        lunch_start_ub_by_vehicle[vehicle_id]
                        - service_start_lb[stop_id]
                        + instance.travel_minutes[DEPOT_NODE, stop_id]
                        + MIDDAY_LUNCH_BREAK_RULE.duration_min
                    )
                    lunch_on_first_arc_after_not_first_ub = (
                        lunch_start_ub_by_vehicle[vehicle_id]
                        - max(
                            vehicle_stop_lb[vehicle_id, stop_id],
                            MIDDAY_LUNCH_BREAK_RULE.window_start_min
                            + MIDDAY_LUNCH_BREAK_RULE.duration_min,
                        )
                        + instance.travel_minutes[DEPOT_NODE, stop_id]
                        + MIDDAY_LUNCH_BREAK_RULE.duration_min
                    )
                    model.addConstr(
                        lunch_on_first_arc_expr
                        <= lunch_on_first_arc_no_after_ub
                        * (1 - served_after_lunch[vehicle_id, stop_id])
                        + lunch_on_first_arc_after_not_first_ub
                        * (
                            served_after_lunch[vehicle_id, stop_id]
                            - lunch_before_start_arc[vehicle_id, stop_id]
                        ),
                        name=f"lunch_on_first_arc_ub[{vehicle_id},{stop_id}]",
                    )
            if feasible_arc[vehicle_id, stop_id, DEPOT_NODE]:
                service_to_depot_min = service_min_before_arc_for_vehicle_id(
                    instance,
                    vehicle_id,
                    stop_id,
                    DEPOT_NODE,
                )
                time_to_depot_expr = (
                    return_min[vehicle_id]
                    - service_start_min[stop_id]
                    - service_to_depot_min
                    - instance.travel_minutes[stop_id, DEPOT_NODE]
                )
                if MIDDAY_LUNCH_BREAK_RULE is not None:
                    time_to_depot_expr -= (
                        MIDDAY_LUNCH_BREAK_RULE.duration_min
                        * (
                            lunch_required[vehicle_id]
                            - served_after_lunch[vehicle_id, stop_id]
                        )
                    )
                time_to_depot_lb = min(
                    -service_start_ub[stop_id]
                    - service_to_depot_min
                    - instance.travel_minutes[stop_id, DEPOT_NODE],
                    start_limit[vehicle_id]
                    - vehicle_stop_ub[vehicle_id, stop_id]
                    - service_to_depot_min
                    - instance.travel_minutes[stop_id, DEPOT_NODE]
                    - lunch_duration_min,
                )
                model.addConstr(
                    time_to_depot_expr
                    >= time_to_depot_lb
                    * (1 - x[vehicle_id, stop_id, DEPOT_NODE]),
                    name=f"time_to_depot[{vehicle_id},{stop_id}]",
                )

        for stop_id in stop_ids:
            for next_stop in stop_ids:
                if stop_id == next_stop or not feasible_arc[vehicle_id, stop_id, next_stop]:
                    continue
                model.addConstr(
                    served_after_lunch[vehicle_id, next_stop]
                    - served_after_lunch[vehicle_id, stop_id]
                    >= -(1 - x[vehicle_id, stop_id, next_stop]),
                    name=f"lunch_state_monotone[{vehicle_id},{stop_id},{next_stop}]",
                )
                stop_to_next_service_min = service_min_before_arc_for_vehicle_id(
                    instance,
                    vehicle_id,
                    stop_id,
                    next_stop,
                )
                time_between_stops_expr = (
                    service_start_min[next_stop]
                    - service_start_min[stop_id]
                    - stop_to_next_service_min
                    - instance.travel_minutes[stop_id, next_stop]
                )
                if MIDDAY_LUNCH_BREAK_RULE is not None:
                    time_between_stops_expr -= (
                        MIDDAY_LUNCH_BREAK_RULE.duration_min
                        * (
                            served_after_lunch[vehicle_id, next_stop]
                            - served_after_lunch[vehicle_id, stop_id]
                        )
                    )
                time_between_stops_lb = (
                    service_start_lb[next_stop]
                    - service_start_ub[stop_id]
                    - stop_to_next_service_min
                    - instance.travel_minutes[stop_id, next_stop]
                    - lunch_duration_min
                )
                model.addConstr(
                    time_between_stops_expr
                    >= time_between_stops_lb
                    * (1 - x[vehicle_id, stop_id, next_stop]),
                    name=f"time_between_stops[{vehicle_id},{stop_id},{next_stop}]",
                )
                if MIDDAY_LUNCH_BREAK_RULE is not None:
                    model.addConstr(
                        lunch_before_stop_arc[vehicle_id, stop_id, next_stop]
                        <= x[vehicle_id, stop_id, next_stop],
                        name=f"lunch_before_stop_arc_x_ub[{vehicle_id},{stop_id},{next_stop}]",
                    )
                    model.addConstr(
                        lunch_before_stop_arc[vehicle_id, stop_id, next_stop]
                        <= served_after_lunch[vehicle_id, next_stop],
                        name=f"lunch_before_stop_arc_state_ub[{vehicle_id},{stop_id},{next_stop}]",
                    )
                    model.addConstr(
                        lunch_before_stop_arc[vehicle_id, stop_id, next_stop]
                        >= x[vehicle_id, stop_id, next_stop]
                        + served_after_lunch[vehicle_id, next_stop]
                        - 1,
                        name=f"lunch_before_stop_arc_lb[{vehicle_id},{stop_id},{next_stop}]",
                    )
                    lunch_between_stops_expr = (
                        lunch_start_min[vehicle_id]
                        - service_start_min[next_stop]
                        + instance.travel_minutes[stop_id, next_stop]
                        + MIDDAY_LUNCH_BREAK_RULE.duration_min
                    )
                    lunch_between_stops_no_after_ub = (
                        lunch_start_ub_by_vehicle[vehicle_id]
                        - service_start_lb[next_stop]
                        + instance.travel_minutes[stop_id, next_stop]
                        + MIDDAY_LUNCH_BREAK_RULE.duration_min
                    )
                    lunch_between_stops_after_not_arc_ub = (
                        lunch_start_ub_by_vehicle[vehicle_id]
                        - max(
                            vehicle_stop_lb[vehicle_id, next_stop],
                            MIDDAY_LUNCH_BREAK_RULE.window_start_min
                            + MIDDAY_LUNCH_BREAK_RULE.duration_min,
                        )
                        + instance.travel_minutes[stop_id, next_stop]
                        + MIDDAY_LUNCH_BREAK_RULE.duration_min
                    )
                    model.addConstr(
                        lunch_between_stops_expr
                        <= lunch_between_stops_no_after_ub
                        * (1 - served_after_lunch[vehicle_id, next_stop])
                        + lunch_between_stops_after_not_arc_ub
                        * (
                            served_after_lunch[vehicle_id, next_stop]
                            - lunch_before_stop_arc[vehicle_id, stop_id, next_stop]
                        ),
                        name=f"lunch_between_stops_ub[{vehicle_id},{stop_id},{next_stop}]",
                    )
                load_progress_expr = (
                    load_after_kg[vehicle_id, next_stop]
                    - load_after_kg[vehicle_id, stop_id]
                    - instance.stops[next_stop].demand_kg
                )
                model.addConstr(
                    load_progress_expr
                    >= -(vehicle.capacity_kg + instance.stops[next_stop].demand_kg)
                    * (visit[vehicle_id, stop_id] - x[vehicle_id, stop_id, next_stop])
                    - instance.stops[next_stop].demand_kg
                    * (1 - visit[vehicle_id, stop_id]),
                    name=f"load_progress_lb[{vehicle_id},{stop_id},{next_stop}]",
                )
                model.addConstr(
                    load_progress_expr
                    <= (vehicle.capacity_kg - instance.stops[next_stop].demand_kg)
                    * (visit[vehicle_id, next_stop] - x[vehicle_id, stop_id, next_stop])
                    - instance.stops[next_stop].demand_kg
                    * (1 - visit[vehicle_id, next_stop]),
                    name=f"load_progress_ub[{vehicle_id},{stop_id},{next_stop}]",
                )

    objective = gp.quicksum(
        distance_cost_expr[vehicle_id]
        + instance.vehicles[vehicle_id].fixed_daily_cost * use_vehicle[vehicle_id]
        + instance.cost_params["Driver_Hourly_Wage"] * billed_hours[vehicle_id]
        for vehicle_id in vehicle_ids
    ) + gp.quicksum(
        instance.cost_params["Late_Delivery_Penalty_per_Hour"] / 60.0 * late_min[stop_id]
        for stop_id in stop_ids
    )
    model.setObjective(objective, GRB.MINIMIZE)

    model._data = {
        "x": x,
        "visit": visit,
        "use_vehicle": use_vehicle,
        "depart_min": depart_min,
        "return_min": return_min,
        "service_start_min": service_start_min,
        "late_min": late_min,
        "billed_hours": billed_hours,
        "load_after_kg": load_after_kg,
        "lunch_required": lunch_required,
        "lunch_start_min": lunch_start_min,
        "served_after_lunch": served_after_lunch,
        "served_before_lunch": served_before_lunch,
        "lunch_before_start_arc": lunch_before_start_arc,
        "lunch_before_stop_arc": lunch_before_stop_arc,
        "incoming": incoming,
        "starts": starts,
        "returns": returns,
        "drive_min_expr": drive_min_expr,
        "service_min_expr": service_min_expr,
        "served_load_expr": served_load_expr,
        "active_min_expr": active_min_expr,
        "distance_cost_expr": distance_cost_expr,
        "fixed_arc_count": fixed_arc_count,
        "symmetry_group_count": symmetry_group_count,
        "node_ids": node_ids,
        "stop_ids": stop_ids,
        "vehicle_ids": vehicle_ids,
    }
    return model, model._data


def extract_arc_flow_route_from_solution(
    data: dict[str, Any],
    vehicle_id: int,
    value_of: Callable[[gp.Var], float],
) -> tuple[int, ...]:
    x = data["x"]
    stop_ids = data["stop_ids"]
    route = [DEPOT_NODE]
    current = DEPOT_NODE
    seen_stops: set[int] = set()

    while True:
        next_nodes = [
            node
            for node in stop_ids + [DEPOT_NODE]
            if node != current
            and (vehicle_id, current, node) in x
            and value_of(x[vehicle_id, current, node]) > 0.5
        ]
        if not next_nodes:
            break
        next_node = next_nodes[0]
        route.append(next_node)
        if next_node == DEPOT_NODE:
            break
        if next_node in seen_stops:
            break
        seen_stops.add(next_node)
        current = next_node

    return tuple(route)


def extract_arc_flow_vehicle_solutions(
    instance: Instance,
    data: dict[str, Any],
) -> list[ArcFlowVehicleSolution]:
    return extract_arc_flow_vehicle_solutions_from_solution(
        instance,
        data,
        lambda var: float(var.X),
    )


def extract_arc_flow_vehicle_solutions_from_solution(
    instance: Instance,
    data: dict[str, Any],
    value_of: Callable[[gp.Var], float],
) -> list[ArcFlowVehicleSolution]:
    vehicle_solutions: list[ArcFlowVehicleSolution] = []
    for vehicle_id in data["vehicle_ids"]:
        if value_of(data["use_vehicle"][vehicle_id]) < 0.5:
            continue
        vehicle = instance.vehicles[vehicle_id]
        route = extract_arc_flow_route_from_solution(data, vehicle_id, value_of)
        served_stops = [stop_id for stop_id in route if stop_id != DEPOT_NODE]
        service_start_min = {stop_id: value_of(data["service_start_min"][stop_id]) for stop_id in served_stops}
        late_min = {stop_id: value_of(data["late_min"][stop_id]) for stop_id in served_stops}
        lunch_break_start_min = (
            value_of(data["lunch_start_min"][vehicle_id])
            if value_of(data["lunch_required"][vehicle_id]) > 0.5
            else None
        )
        distance_cost = sum(
            instance.distance_miles[i, j]
            * (
                vehicle.cost_per_mile
                + instance.cost_params["Fuel_Cost_per_Liter"]
                * instance.cost_params["Avg_Fuel_Consumption_L_per_mile"]
            )
            for i, j in zip(route, route[1:])
        )
        billed_hours = int(round(value_of(data["billed_hours"][vehicle_id])))
        drive_min = sum(instance.travel_minutes[i, j] for i, j in zip(route, route[1:]))
        service_min = route_service_min_for_vehicle_id(instance, vehicle_id, route)
        load_kg = sum(instance.stops[stop_id].demand_kg for stop_id in served_stops)
        depart_min = value_of(data["depart_min"][vehicle_id])
        return_min = value_of(data["return_min"][vehicle_id])
        route_cost = (
            vehicle.fixed_daily_cost
            + distance_cost
            + billed_hours * instance.cost_params["Driver_Hourly_Wage"]
            + sum(late_min.values()) / 60.0 * instance.cost_params["Late_Delivery_Penalty_per_Hour"]
        )
        vehicle_solutions.append(
            ArcFlowVehicleSolution(
                vehicle_id=vehicle_id,
                vehicle=vehicle,
                route=route,
                load_kg=load_kg,
                depart_min=depart_min,
                return_min=return_min,
                drive_min=drive_min,
                service_min=service_min,
                active_min=return_min - depart_min,
                billed_hours=billed_hours,
                distance_cost=distance_cost,
                route_cost=route_cost,
                service_start_min=service_start_min,
                late_min=late_min,
                lunch_break_start_min=lunch_break_start_min,
            )
        )
    return vehicle_solutions


def render_arc_flow_vehicle_solution_report(
    instance: Instance,
    vehicle_solutions: list[ArcFlowVehicleSolution],
    *,
    status: str,
    objective: float | None,
    best_bound: float | None,
    gap: float | None,
) -> str:
    lines = [
        f"Status: {status}",
        "Formulation: Arc-Flow MIP with Vehicle Symmetry Breaking",
    ]
    if best_bound is not None:
        lines.append(f"Best bound: {best_bound:.2f}")
    if objective is not None:
        lines.append(f"Objective value: {objective:.2f}")
    if gap is not None:
        lines.append(f"Relative gap: {100.0 * gap:.2f}%")
    lines.append("")

    if not vehicle_solutions:
        lines.append("No feasible incumbent route set found.")
        return "\n".join(lines) + "\n"

    for vehicle_solution in vehicle_solutions:
        lines.append(f"Vehicle {vehicle_solution.vehicle_id} ({vehicle_solution.vehicle.vehicle_type})")
        lines.append(
            "  "
            f"Depart {minutes_to_clock(vehicle_solution.depart_min)}, "
            f"Return {minutes_to_clock(vehicle_solution.return_min)}, "
            f"Load {vehicle_solution.load_kg:.0f} / {vehicle_solution.vehicle.capacity_kg:.0f} kg"
        )
        lines.append(
            "  "
            f"Active {vehicle_solution.active_min:.1f} min, Drive {vehicle_solution.drive_min:.1f} min, "
            f"Service {vehicle_solution.service_min:.1f} min, Billed {vehicle_solution.billed_hours} h"
        )
        if vehicle_solution.lunch_break_start_min is not None:
            lunch_end = vehicle_solution.lunch_break_start_min + MIDDAY_LUNCH_BREAK_RULE.duration_min
            lines.append(
                "  "
                f"Lunch break {minutes_to_clock(vehicle_solution.lunch_break_start_min)}-"
                f"{minutes_to_clock(lunch_end)}"
            )
        lines.append(
            "  "
            f"Distance cost {vehicle_solution.distance_cost:.2f}, "
            f"Route cost {vehicle_solution.route_cost:.2f}"
        )
        lines.append(f"  Route IDs: {format_route_stop_ids(vehicle_solution.route)}")
        lines.append(
            "  Route detail: "
            f"{format_route_stop_details(instance, vehicle_solution.route)}"
        )
        for stop_id in vehicle_solution.route:
            if stop_id == DEPOT_NODE:
                continue
            lines.append(
                "    "
                f"{format_stop_detail_label(instance.stops[stop_id])}: "
                f"start {minutes_to_clock(vehicle_solution.service_start_min[stop_id])}, "
                f"late {vehicle_solution.late_min[stop_id]:.1f} min"
            )
        lines.append("")

    return "\n".join(lines)


def render_branch_and_price_route_report(
    instance: Instance,
    *,
    status: str,
    objective: float | None,
    best_bound: float | None,
    gap: float | None,
    selected_routes: list[RouteColumn],
    vehicle_types: dict[str, VehicleTypeData],
    nodes_processed: int | None = None,
    nodes_pruned: int | None = None,
    root_lp_objective: float | None = None,
    global_columns: int | None = None,
) -> str:
    lines = [
        f"Status: {status}",
        "Formulation: Branch-and-Price Set Partitioning",
    ]
    if global_columns is not None:
        lines.append(f"Columns generated: {global_columns}")
    if nodes_processed is not None:
        lines.append(f"Nodes processed: {nodes_processed}")
    if nodes_pruned is not None:
        lines.append(f"Nodes pruned: {nodes_pruned}")
    if root_lp_objective is not None:
        lines.append(f"Root LP objective: {root_lp_objective:.2f}")
    if best_bound is not None:
        lines.append(f"Best bound: {best_bound:.2f}")
    if objective is not None:
        lines.append(f"Objective value: {objective:.2f}")
    if gap is not None:
        lines.append(f"Relative gap: {100.0 * gap:.2f}%")
    lines.append("")

    if not selected_routes:
        lines.append("No feasible incumbent route set found.")
        return "\n".join(lines) + "\n"

    usage_counter: dict[str, int] = defaultdict(int)
    for route in sorted(
        selected_routes,
        key=lambda candidate: (candidate.type_id, candidate.return_min, candidate.stop_sequence),
    ):
        vehicle_type = vehicle_types[route.type_id]
        usage_counter[route.type_id] += 1
        vehicle_position = usage_counter[route.type_id] - 1
        vehicle_id = vehicle_type.vehicle_ids[vehicle_position]

        lines.append(f"Vehicle {vehicle_id} ({vehicle_type.vehicle_type})")
        lines.append(
            "  "
            f"Depart {minutes_to_clock(vehicle_start_limit(instance, vehicle_type))}, "
            f"Return {minutes_to_clock(route.return_min)}, "
            f"Load {route.load_kg:.0f} / {vehicle_type.capacity_kg:.0f} kg"
        )
        lines.append(
            "  "
            f"Active {route.active_min:.1f} min, Distance {route.distance_mi:.2f} mi, "
            f"Route cost {route.cost:.2f}"
        )
        if route.lunch_break_start_min is not None:
            lunch_end = route.lunch_break_start_min + MIDDAY_LUNCH_BREAK_RULE.duration_min
            lines.append(
                "  "
                f"Lunch break {minutes_to_clock(route.lunch_break_start_min)}-"
                f"{minutes_to_clock(lunch_end)}"
            )
        route_with_depot = (DEPOT_NODE,) + route.stop_sequence + (DEPOT_NODE,)
        lines.append(f"  Route IDs: {format_route_stop_ids(route_with_depot)}")
        lines.append(
            "  Route detail: "
            f"{format_route_stop_details(instance, route_with_depot)}"
        )
        for stop_id in route.stop_sequence:
            start = route.service_start_min[stop_id]
            early = max(instance.stops[stop_id].earliest_min - start, 0.0)
            late = route.late_min[stop_id]
            lines.append(
                "    "
                f"{format_stop_detail_label(instance.stops[stop_id])}: "
                f"start {minutes_to_clock(start)}, "
                f"early {early:.1f} min, late {late:.1f} min"
            )
        lines.append("")

    return "\n".join(lines)


def solve_arc_flow_model(
    workbook_path: Path,
    time_limit: float | None,
    mip_gap: float | None,
    mip_profile: ArcFlowMIPProfile | None = None,
    skip_mip_profile_params: frozenset[str] = frozenset(),
    progress: ProgressTracker | None = None,
    trace_logger: ObjectiveTraceLogger | None = None,
    persistence_manager: MIPPersistenceManager | None = None,
    incumbent_report_writer: IncumbentReportWriter | None = None,
) -> ArcFlowSolveResult:
    run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    progress_log_path = default_arc_flow_progress_log_path(workbook_path, run_timestamp)
    progress_logger = ArcFlowProgressLogger(progress_log_path)
    bound_plot_path = default_mip_bound_plot_path(workbook_path, run_timestamp)
    bound_logger = MIPBoundLogger(bound_plot_path)
    saved_bound_plot_path: Path | None = None
    loaded_start_path: Path | None = None
    model: gp.Model | None = None

    try:
        if progress_logger.path is not None:
            record_interrupt_artifact("arc_flow_progress_log", progress_logger.path)
            print(f"Arc-flow progress log: {progress_logger.path}")
        instance = load_instance(workbook_path)
        print_stop_reference(instance)
        model, data = build_arc_flow_model(instance)

        if mip_profile is not None:
            print(f"Arc-flow MIP profile: {mip_profile.name}")
            print(f"  {mip_profile.description}")
            apply_arc_flow_mip_profile(model, mip_profile, skip_params=skip_mip_profile_params)
        if time_limit is not None:
            model.Params.TimeLimit = time_limit
        if mip_gap is not None:
            model.Params.MIPGap = mip_gap
        if persistence_manager is not None:
            loaded_start_path = persistence_manager.load_start(model, workbook_path)
        applied_constructive_seed = False
        if loaded_start_path is None:
            applied_constructive_seed = apply_constructive_arc_flow_mip_start(
                instance,
                data,
                model=model,
            )
            if applied_constructive_seed:
                model.Params.StartNodeLimit = max(int(getattr(model.Params, "StartNodeLimit", 0)), 5000)
                model.update()
                if progress is not None:
                    progress.emit(
                        "arc-flow master MIP: applied built-in constructive MIP start with elevated start repair effort",
                        force=True,
                    )
                else:
                    print(
                        "Applied built-in constructive MIP start with elevated start repair effort.",
                        flush=True,
                    )
        else:
            # Keep the persisted resume as start 0 and append the built-in constructive
            # seed as start 1 so Gurobi can fall back immediately if the resume start is
            # stale, partial, or otherwise fails to yield an incumbent.
            model.NumStart = max(int(model.NumStart), 2)
            applied_constructive_seed = apply_constructive_arc_flow_mip_start(
                instance,
                data,
                model=model,
                start_number=1,
            )
            if applied_constructive_seed:
                model.Params.StartNodeLimit = max(int(getattr(model.Params, "StartNodeLimit", 0)), 5000)
                model.update()
                print(
                    "Loaded saved MIP start from persistence and added the built-in "
                    "constructive fallback start.",
                    flush=True,
                )
            else:
                print(
                    "Loaded saved MIP start from persistence; built-in constructive "
                    "fallback start could not be built.",
                    flush=True,
                )
        mip_start_vars = [
            (var.VarName, var, var.VType)
            for var in model.getVars()
            if var.VType in {GRB.BINARY, GRB.INTEGER, GRB.SEMIINT}
        ]

        def maybe_save_mip_incumbent_report(cb_model: gp.Model) -> None:
            try:
                objective = float(cb_model.cbGet(GRB.Callback.MIPSOL_OBJ))
                report_should_save = (
                    incumbent_report_writer is not None
                    and incumbent_report_writer.should_save(objective)
                )
                if persistence_manager is not None:
                    saved_mst_path = persistence_manager.maybe_write_resume_mip_start(
                        objective=objective,
                        min_improvement=(
                            incumbent_report_writer.min_improvement
                            if incumbent_report_writer is not None
                            else DEFAULT_INCUMBENT_REPORT_MIN_IMPROVEMENT
                        ),
                        named_values=[
                            (
                                var_name,
                                float(int(round(cb_model.cbGetSolution(var))))
                                if var_type in {GRB.BINARY, GRB.INTEGER, GRB.SEMIINT}
                                else float(cb_model.cbGetSolution(var)),
                            )
                            for var_name, var, var_type in mip_start_vars
                        ],
                    )
                    if saved_mst_path is not None and progress is not None:
                        progress.emit(
                            f"arc-flow master MIP: refreshed resume MIP start at {saved_mst_path}",
                            force=True,
                        )
                if not report_should_save:
                    return
                best_bound = MIPBoundLogger._normalize_bound(cb_model.cbGet(GRB.Callback.MIPSOL_OBJBND))
                gap = relative_gap(objective, best_bound)
                vehicle_solutions = extract_arc_flow_vehicle_solutions_from_solution(
                    instance,
                    data,
                    lambda var: float(cb_model.cbGetSolution(var)),
                )
                report_text = render_arc_flow_vehicle_solution_report(
                    instance,
                    vehicle_solutions,
                    status="RUNNING",
                    objective=objective,
                    best_bound=best_bound,
                    gap=gap,
                )
                saved_path = incumbent_report_writer.maybe_save(objective, report_text)
                if saved_path is not None and progress is not None:
                    progress.emit(
                        f"arc-flow master MIP: saved incumbent route report to {saved_path}",
                        force=True,
                    )
            except Exception:
                return

        try:
            optimize_with_heartbeat(
                model,
                progress=progress,
                label="arc-flow master MIP",
                bound_logger=bound_logger,
                mip_solution_callback=maybe_save_mip_incumbent_report,
                message_callback=progress_logger.record_message,
            )
            capture_mip_bound_snapshot(model, bound_logger)
        except KeyboardInterrupt:
            capture_mip_bound_snapshot(model, bound_logger)
            interrupted_objective = bound_logger.latest_upper_bound()
            interrupted_best_bound = bound_logger.latest_lower_bound()
            interrupted_gap = relative_gap(interrupted_objective, interrupted_best_bound)
            saved_bound_plot_path = save_mip_bound_plot_snapshot(
                workbook_path,
                bound_logger,
                status_label="INTERRUPTED",
            )
            resume_solution_path: Path | None = None
            if persistence_manager is not None:
                resume_solution_path = persistence_manager.finalize_run(
                    workbook_path=workbook_path,
                    status="INTERRUPTED",
                    loaded_start_path=loaded_start_path,
                    objective=interrupted_objective,
                    best_bound=interrupted_best_bound,
                    gap=interrupted_gap,
                    note="Arc-flow MIP interrupted",
                    model=model,
                )
            if incumbent_report_writer is not None and incumbent_report_writer.path is not None:
                record_interrupt_artifact("incumbent_report", incumbent_report_writer.path)
            if saved_bound_plot_path is not None:
                record_interrupt_artifact("partial_mip_bound_plot", saved_bound_plot_path)
                print(f"\nPartial MIP bound plot saved to {saved_bound_plot_path}")
            if resume_solution_path is not None:
                record_interrupt_artifact("resumable_mip_start", resume_solution_path)
                print(f"Saved resumable MIP start to {resume_solution_path}")
            raise

        status = gurobi_status_name(model.Status)

        best_bound: float | None
        try:
            best_bound = model.ObjBound
        except gp.GurobiError:
            best_bound = None

        objective = model.ObjVal if model.SolCount > 0 else None
        gap = relative_gap(objective, best_bound)

        vehicle_solutions = (
            extract_arc_flow_vehicle_solutions(instance, data) if model.SolCount > 0 else []
        )

        saved_bound_plot_path = bound_logger.save_plot(f"{workbook_path.stem} MIP Bounds ({status})")
        result = ArcFlowSolveResult(
            status=status,
            objective=objective,
            best_bound=best_bound,
            gap=gap,
            vehicle_solutions=vehicle_solutions,
            fixed_arc_count=data["fixed_arc_count"],
            symmetry_group_count=data["symmetry_group_count"],
            bound_plot_path=saved_bound_plot_path,
        )

        resume_solution_path: Path | None = None
        if persistence_manager is not None:
            resume_solution_path = persistence_manager.finalize_run(
                workbook_path=workbook_path,
                status=result.status,
                loaded_start_path=loaded_start_path,
                objective=result.objective,
                best_bound=result.best_bound,
                gap=result.gap,
                note="Arc-flow MIP finished",
                model=model,
            )

        if trace_logger is not None:
            trace_logger.record(
                event="solve_end",
                status=result.status,
                incumbent_objective=result.objective,
                best_bound=result.best_bound,
                relative_gap=result.gap,
                note="Arc-flow MIP finished",
            )
        if incumbent_report_writer is not None and result.objective is not None:
            incumbent_report_writer.maybe_save(
                result.objective,
                render_arc_flow_vehicle_solution_report(
                    instance,
                    result.vehicle_solutions,
                    status=result.status,
                    objective=result.objective,
                    best_bound=result.best_bound,
                    gap=result.gap,
                ),
                force=True,
            )

        print(f"Status: {result.status}")
        print("Formulation: Arc-Flow MIP with Vehicle Symmetry Breaking")
        print(f"Fixed impossible arcs: {result.fixed_arc_count}")
        print(f"Symmetry groups: {result.symmetry_group_count}")
        if loaded_start_path is not None:
            print(f"Loaded saved MIP start: {loaded_start_path}")
        if persistence_manager is not None:
            print(f"MIP persistence dir: {persistence_manager.path}")
        if resume_solution_path is not None:
            print(f"Saved resumable MIP start: {resume_solution_path}")
        if result.bound_plot_path is not None:
            print(f"MIP bound plot: {result.bound_plot_path}")
        if result.best_bound is not None:
            print(f"Best bound: {result.best_bound:.2f}")
        if result.objective is not None:
            print(f"Objective value: {result.objective:.2f}")
        if result.gap is not None:
            print(f"Relative gap: {100.0 * result.gap:.2f}%")
        print()

        if not result.vehicle_solutions:
            print("No feasible solution found.")
            return result

        for vehicle_solution in result.vehicle_solutions:
            print(f"Vehicle {vehicle_solution.vehicle_id} ({vehicle_solution.vehicle.vehicle_type})")
            print(
                "  "
                f"Depart {minutes_to_clock(vehicle_solution.depart_min)}, "
                f"Return {minutes_to_clock(vehicle_solution.return_min)}, "
                f"Load {vehicle_solution.load_kg:.0f} / {vehicle_solution.vehicle.capacity_kg:.0f} kg"
            )
            print(
                "  "
                f"Active {vehicle_solution.active_min:.1f} min, Drive {vehicle_solution.drive_min:.1f} min, "
                f"Service {vehicle_solution.service_min:.1f} min, Billed {vehicle_solution.billed_hours} h"
            )
            if vehicle_solution.lunch_break_start_min is not None:
                lunch_end = vehicle_solution.lunch_break_start_min + MIDDAY_LUNCH_BREAK_RULE.duration_min
                print(
                    "  "
                    f"Lunch break {minutes_to_clock(vehicle_solution.lunch_break_start_min)}-"
                    f"{minutes_to_clock(lunch_end)}"
                )
            print(
                "  "
                f"Distance cost {vehicle_solution.distance_cost:.2f}, "
                f"Route cost {vehicle_solution.route_cost:.2f}"
            )
            print(f"  Route IDs: {format_route_stop_ids(vehicle_solution.route)}")
            print(
                "  Route detail: "
                f"{format_route_stop_details(instance, vehicle_solution.route)}"
            )
            for stop_id in vehicle_solution.route:
                if stop_id == DEPOT_NODE:
                    continue
                print(
                    "    "
                    f"{format_stop_detail_label(instance.stops[stop_id])}: "
                    f"start {minutes_to_clock(vehicle_solution.service_start_min[stop_id])}, "
                    f"late {vehicle_solution.late_min[stop_id]:.1f} min"
                )
            print()

        return result
    except BaseException as exc:
        if saved_bound_plot_path is None:
            if model is not None:
                capture_mip_bound_snapshot(model, bound_logger)
            saved_bound_plot_path = save_mip_bound_plot_snapshot(
                workbook_path,
                bound_logger,
                status_label=("INTERRUPTED" if isinstance(exc, KeyboardInterrupt) else "ABORTED"),
            )
            if saved_bound_plot_path is not None and isinstance(exc, KeyboardInterrupt):
                record_interrupt_artifact("partial_mip_bound_plot", saved_bound_plot_path)
                print(f"\nPartial MIP bound plot saved to {saved_bound_plot_path}")
        raise
    finally:
        progress_logger.close()


def best_stop_insertion(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    current_sequence: tuple[int, ...],
    stop_id: int,
) -> RouteColumn | None:
    current_route = (
        evaluate_route_sequence(instance, vehicle_type, current_sequence)
        if current_sequence
        else None
    )
    current_cost = current_route.cost if current_route is not None else 0.0

    best_route: RouteColumn | None = None
    best_delta = float("inf")
    for pos in range(len(current_sequence) + 1):
        candidate_sequence = current_sequence[:pos] + (stop_id,) + current_sequence[pos:]
        candidate_route = evaluate_route_sequence(instance, vehicle_type, candidate_sequence)
        if candidate_route is None:
            continue
        delta = candidate_route.cost - current_cost
        if delta < best_delta - 1e-9:
            best_delta = delta
            best_route = candidate_route
    return best_route


def construct_seed_partition_routes(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
) -> list[RouteColumn]:
    slot_types: list[VehicleTypeData] = []
    for vehicle_type in sorted(
        vehicle_types.values(),
        key=lambda vt: (-vt.capacity_kg, vt.fixed_daily_cost, vt.type_id),
    ):
        slot_types.extend([vehicle_type] * vehicle_type.count)

    if not slot_types:
        return []

    stop_ids = sorted(instance.stops)
    round_trip_minutes = {
        stop_id: (
            instance.travel_minutes[DEPOT_NODE, stop_id]
            + instance.travel_minutes[stop_id, DEPOT_NODE]
        )
        for stop_id in stop_ids
    }
    orderings = [
        sorted(
            stop_ids,
            key=lambda stop_id: (
                -instance.stops[stop_id].demand_kg,
                -round_trip_minutes[stop_id],
                instance.stops[stop_id].latest_min,
            ),
        ),
        sorted(
            stop_ids,
            key=lambda stop_id: (
                -round_trip_minutes[stop_id],
                -instance.stops[stop_id].demand_kg,
                instance.stops[stop_id].latest_min,
            ),
        ),
        sorted(
            stop_ids,
            key=lambda stop_id: (
                instance.stops[stop_id].latest_min,
                -instance.stops[stop_id].demand_kg,
                -round_trip_minutes[stop_id],
            ),
        ),
    ]

    for ordering in orderings:
        slot_routes: list[RouteColumn | None] = [None] * len(slot_types)
        feasible = True

        for stop_id in ordering:
            best_slot: int | None = None
            best_candidate: RouteColumn | None = None
            best_delta = float("inf")

            for slot_idx, vehicle_type in enumerate(slot_types):
                current_route = slot_routes[slot_idx]
                current_sequence = (
                    current_route.stop_sequence if current_route is not None else tuple()
                )
                candidate = best_stop_insertion(
                    instance=instance,
                    vehicle_type=vehicle_type,
                    current_sequence=current_sequence,
                    stop_id=stop_id,
                )
                if candidate is None:
                    continue

                current_cost = current_route.cost if current_route is not None else 0.0
                delta = candidate.cost - current_cost
                if delta < best_delta - 1e-9:
                    best_delta = delta
                    best_slot = slot_idx
                    best_candidate = candidate

            if best_slot is None or best_candidate is None:
                feasible = False
                break

            slot_routes[best_slot] = best_candidate

        if feasible:
            return [route for route in slot_routes if route is not None]

    return []


def construct_known_feasible_seed_routes(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
) -> list[RouteColumn]:
    vehicle_types_by_name: dict[str, list[VehicleTypeData]] = defaultdict(list)
    for vehicle_type in vehicle_types.values():
        vehicle_types_by_name[vehicle_type.vehicle_type].append(vehicle_type)
    for candidate_types in vehicle_types_by_name.values():
        candidate_types.sort(key=lambda candidate: min(candidate.vehicle_ids))

    remaining_by_type: dict[str, int] = {
        type_id: vehicle_type.count for type_id, vehicle_type in vehicle_types.items()
    }
    seed_routes: list[RouteColumn] = []
    covered_stops: list[int] = []

    for vehicle_type_name, route_sequences in CONSTRUCTIVE_ARC_FLOW_SEED_BY_VEHICLE_TYPE.items():
        candidate_types = vehicle_types_by_name.get(vehicle_type_name, [])
        if not candidate_types:
            return []
        for stop_sequence in route_sequences:
            chosen_route: RouteColumn | None = None
            chosen_vehicle_type: VehicleTypeData | None = None
            for candidate_type in candidate_types:
                if remaining_by_type[candidate_type.type_id] <= 0:
                    continue
                route = evaluate_route_sequence(instance, candidate_type, stop_sequence)
                if route is None:
                    continue
                if sum(route.late_min.values()) > 1e-6:
                    continue
                chosen_route = route
                chosen_vehicle_type = candidate_type
                break
            if chosen_route is None or chosen_vehicle_type is None:
                return []
            remaining_by_type[chosen_vehicle_type.type_id] -= 1
            seed_routes.append(chosen_route)
            covered_stops.extend(stop_sequence)

    if len(covered_stops) != len(instance.stops):
        return []
    if set(covered_stops) != set(instance.stops):
        return []
    return seed_routes


def initial_routes(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
) -> list[RouteColumn]:
    routes: list[RouteColumn] = []
    for vehicle_type in vehicle_types.values():
        for rule in REDSTONE_PLAZA_PAIR_RULES:
            for stop_sequence in ((rule.stop_a, rule.stop_b), (rule.stop_b, rule.stop_a)):
                route = evaluate_route_sequence(instance, vehicle_type, stop_sequence)
                if route is not None:
                    routes.append(route)
        for stop_id in sorted(instance.stops):
            route = evaluate_route_sequence(instance, vehicle_type, (stop_id,))
            if route is not None:
                routes.append(route)

    route_keys = {(route.type_id, route.stop_sequence) for route in routes}
    for route in construct_known_feasible_seed_routes(instance, vehicle_types):
        key = (route.type_id, route.stop_sequence)
        if key not in route_keys:
            route_keys.add(key)
            routes.append(route)
    for route in construct_seed_partition_routes(instance, vehicle_types):
        key = (route.type_id, route.stop_sequence)
        if key not in route_keys:
            route_keys.add(key)
            routes.append(route)
    return routes


def build_branch_context(
    branch_state: BranchState,
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
) -> BranchContext | None:
    forced_successor: dict[int, int] = {}
    forced_predecessor: dict[int, int] = {}
    forced_type_for_stop: dict[int, str] = {}
    excluded_types_by_stop: dict[int, set[str]] = defaultdict(set)

    for arc in branch_state.forced_arcs:
        if arc in branch_state.forbidden_arcs:
            return None
        i, j = arc
        if i == DEPOT_NODE and j == DEPOT_NODE:
            return None
        if i != DEPOT_NODE:
            existing = forced_successor.get(i)
            if existing is not None and existing != j:
                return None
            forced_successor[i] = j
        if j != DEPOT_NODE:
            existing = forced_predecessor.get(j)
            if existing is not None and existing != i:
                return None
            forced_predecessor[j] = i

    for start_node in forced_successor:
        seen: set[int] = set()
        current = start_node
        while current != DEPOT_NODE and current in forced_successor:
            if current in seen:
                return None
            seen.add(current)
            next_node = forced_successor[current]
            if next_node == DEPOT_NODE:
                break
            current = next_node

    for stop_id, type_id in branch_state.forced_assignments:
        if type_id not in vehicle_types:
            return None
        existing = forced_type_for_stop.get(stop_id)
        if existing is not None and existing != type_id:
            return None
        forced_type_for_stop[stop_id] = type_id

    for stop_id, type_id in branch_state.excluded_assignments:
        if type_id not in vehicle_types:
            return None
        excluded_types_by_stop[stop_id].add(type_id)

    all_type_ids = set(vehicle_types)
    for stop_id, type_id in forced_type_for_stop.items():
        if type_id in excluded_types_by_stop[stop_id]:
            return None
    for stop_id in instance.stops:
        if stop_id not in forced_type_for_stop and excluded_types_by_stop.get(stop_id) == all_type_ids:
            return None

    stop_ids = sorted(instance.stops)
    parent = {stop_id: stop_id for stop_id in stop_ids}

    def find(stop_id: int) -> int:
        while parent[stop_id] != stop_id:
            parent[stop_id] = parent[parent[stop_id]]
            stop_id = parent[stop_id]
        return stop_id

    def union(stop_a: int, stop_b: int) -> None:
        root_a = find(stop_a)
        root_b = find(stop_b)
        if root_a != root_b:
            parent[root_b] = root_a

    for stop_a, stop_b in branch_state.same_route_pairs:
        if stop_a not in instance.stops or stop_b not in instance.stops or stop_a == stop_b:
            return None
        union(stop_a, stop_b)

    group_members: dict[int, set[int]] = defaultdict(set)
    for stop_id in stop_ids:
        group_members[find(stop_id)].add(stop_id)

    together_groups = tuple(
        sorted((frozenset(group) for group in group_members.values() if len(group) > 1), key=min)
    )
    together_group_by_stop = {
        stop_id: group for group in together_groups for stop_id in group
    }

    together_stop_pairs: set[tuple[int, int]] = set()
    for group in together_groups:
        ordered_group = sorted(group)
        for idx, stop_a in enumerate(ordered_group):
            for stop_b in ordered_group[idx + 1 :]:
                together_stop_pairs.add((stop_a, stop_b))

    separated_stop_pairs: set[tuple[int, int]] = set()
    separated_stops_by_stop: dict[int, set[int]] = defaultdict(set)
    for stop_a, stop_b in branch_state.different_route_pairs:
        if stop_a not in instance.stops or stop_b not in instance.stops or stop_a == stop_b:
            return None
        root_a = find(stop_a)
        root_b = find(stop_b)
        if root_a == root_b:
            return None
        for group_stop_a in group_members[root_a]:
            for group_stop_b in group_members[root_b]:
                pair = tuple(sorted((group_stop_a, group_stop_b)))
                separated_stop_pairs.add(pair)
                separated_stops_by_stop[group_stop_a].add(group_stop_b)
                separated_stops_by_stop[group_stop_b].add(group_stop_a)

    return BranchContext(
        forbidden_arcs=branch_state.forbidden_arcs,
        forced_arcs=tuple(sorted(branch_state.forced_arcs)),
        forced_successor=forced_successor,
        forced_predecessor=forced_predecessor,
        forced_type_for_stop=forced_type_for_stop,
        excluded_types_by_stop=excluded_types_by_stop,
        together_groups=together_groups,
        together_group_by_stop=together_group_by_stop,
        together_stop_pairs=frozenset(together_stop_pairs),
        separated_stop_pairs=frozenset(separated_stop_pairs),
        separated_stops_by_stop=separated_stops_by_stop,
    )


def route_is_compatible(route: RouteColumn, branch_ctx: BranchContext) -> bool:
    if not route_respects_required_adjacencies(route):
        return False
    if any(
        not vehicle_type_can_serve_stop(stop_id, route.vehicle_type)
        for stop_id in route.stop_set
    ):
        return False
    if route.arc_set & branch_ctx.forbidden_arcs:
        return False

    for group in branch_ctx.together_groups:
        if route.stop_set & group and not group.issubset(route.stop_set):
            return False

    for stop_id in route.stop_set:
        if branch_ctx.separated_stops_by_stop.get(stop_id, set()) & route.stop_set:
            return False

    for stop_id, forced_type in branch_ctx.forced_type_for_stop.items():
        if stop_id in route.stop_set and route.type_id != forced_type:
            return False

    for stop_id, excluded_types in branch_ctx.excluded_types_by_stop.items():
        if route.type_id in excluded_types and stop_id in route.stop_set:
            return False

    for i, j in branch_ctx.forced_arcs:
        if i == DEPOT_NODE:
            if j in route.stop_set and (i, j) not in route.arc_set:
                return False
        elif j == DEPOT_NODE:
            if i in route.stop_set and (i, j) not in route.arc_set:
                return False
        else:
            if (i in route.stop_set or j in route.stop_set) and (i, j) not in route.arc_set:
                return False

    return True


def solve_pricing_subproblem(
    instance: Instance,
    vehicle_type: VehicleTypeData,
    cover_duals: dict[int, float],
    type_dual: float,
    branch_ctx: BranchContext,
    max_routes: int,
    excluded_stop_sequences: set[tuple[int, ...]] | None,
    pricing_time_limit: float | None,
    deadline: float | None,
    progress: ProgressTracker | None = None,
    progress_label: str | None = None,
    exact_fallback_max_routes: int | None = None,
) -> list[RouteColumn]:
    if max_routes <= 0:
        return []
    heuristic_routes = solve_heuristic_pricing_subproblem(
        instance=instance,
        vehicle_type=vehicle_type,
        cover_duals=cover_duals,
        type_dual=type_dual,
        branch_ctx=branch_ctx,
        max_routes=max_routes,
        excluded_stop_sequences=excluded_stop_sequences,
    )
    if heuristic_routes:
        if progress is not None:
            progress.emit(
                f"{progress_label or f'pricing {vehicle_type.type_id}'}: heuristic pricing "
                f"found {len(heuristic_routes)} improving route(s); skipping exact labeling"
            )
        return heuristic_routes
    fallback_route_limit = (
        EXACT_PRICING_FALLBACK_MAX_ROUTES
        if exact_fallback_max_routes is None
        else exact_fallback_max_routes
    )
    exact_max_routes = min(max_routes, fallback_route_limit)
    if progress is not None and exact_max_routes < max_routes:
        progress.emit(
            f"{progress_label or f'pricing {vehicle_type.type_id}'}: heuristic pricing found "
            "no improving routes; exact fallback will search for the first certified "
            "improving route"
        )
    max_routes = exact_max_routes
    stop_ids = sorted(instance.stops)
    start_limit = vehicle_start_limit(instance, vehicle_type)
    end_limit = vehicle_end_limit(instance, vehicle_type)
    route_horizon = max(end_limit - start_limit, 0.0)
    if route_horizon <= 0:
        return []

    pricing_start = perf_counter()
    distance_unit_cost = route_distance_unit_cost(instance, vehicle_type)
    last_forced_stops = {
        i for i, j in branch_ctx.forced_arcs if i != DEPOT_NODE and j == DEPOT_NODE
    }
    excluded_stop_sequences = excluded_stop_sequences or set()

    def check_time_limit() -> None:
        now = perf_counter()
        if deadline is not None and now >= deadline - 1e-9:
            raise TimeoutError("Overall time limit reached during pricing.")
        if pricing_time_limit is not None and now >= pricing_start + pricing_time_limit - 1e-9:
            raise RuntimeError(
                "A pricing subproblem hit the user pricing time limit before optimality. "
                "Exact branch-and-price requires complete label-setting pricing."
            )

    stop_to_pos = {stop_id: idx for idx, stop_id in enumerate(stop_ids)}
    stop_mask = {stop_id: 1 << stop_to_pos[stop_id] for stop_id in stop_ids}
    together_group_masks = {
        sum(stop_mask[group_stop] for group_stop in group if group_stop in stop_mask)
        for group in branch_ctx.together_groups
    }
    separated_mask_by_stop = {
        stop_id: sum(
            stop_mask[other_stop]
            for other_stop in branch_ctx.separated_stops_by_stop.get(stop_id, set())
            if other_stop in stop_mask
        )
        for stop_id in stop_ids
    }

    feasible_stop: dict[int, bool] = {}
    potential_successors: dict[int, list[int]] = {DEPOT_NODE: []}
    for stop_id in stop_ids:
        potential_successors[stop_id] = []

    for stop_id in stop_ids:
        stop = instance.stops[stop_id]
        service_to_depot_min = service_min_before_arc_for_vehicle_type(
            instance, vehicle_type, stop_id, DEPOT_NODE
        )
        finish_time = hard_service_finish_after_arrival(
            stop,
            start_limit + instance.travel_minutes[DEPOT_NODE, stop_id],
            service_to_depot_min,
        )
        assignment_allowed = True
        forced_type = branch_ctx.forced_type_for_stop.get(stop_id)
        if forced_type is not None and forced_type != vehicle_type.type_id:
            assignment_allowed = False
        if vehicle_type.type_id in branch_ctx.excluded_types_by_stop.get(stop_id, set()):
            assignment_allowed = False

        feasible = (
            assignment_allowed
            and finish_time is not None
            and vehicle_type_data_can_serve_stop(stop_id, vehicle_type)
            and stop.demand_kg <= vehicle_type.capacity_kg + 1e-9
            and finish_time + instance.travel_minutes[stop_id, DEPOT_NODE] <= end_limit + 1e-9
        )
        feasible_stop[stop_id] = feasible
        if feasible and (DEPOT_NODE, stop_id) not in branch_ctx.forbidden_arcs:
            potential_successors[DEPOT_NODE].append(stop_id)

    for i in stop_ids:
        if not feasible_stop[i]:
            continue
        earliest_service_start_i = hard_service_start_after_arrival(
            i, start_limit + instance.travel_minutes[DEPOT_NODE, i]
        )
        if not hard_service_start_is_feasible(i, earliest_service_start_i):
            continue
        for j in stop_ids:
            if i == j or not feasible_stop[j]:
                continue
            if (i, j) in branch_ctx.forbidden_arcs:
                continue
            earliest_depart_i = earliest_service_start_i + service_min_before_arc_for_vehicle_type(
                instance, vehicle_type, i, j
            )
            service_j_to_depot_min = service_min_before_arc_for_vehicle_type(
                instance, vehicle_type, j, DEPOT_NODE
            )
            earliest_finish_j = hard_service_finish_after_arrival(
                instance.stops[j],
                earliest_depart_i + instance.travel_minutes[i, j],
                service_j_to_depot_min,
            )
            if (
                earliest_finish_j is not None
                and earliest_finish_j + instance.travel_minutes[j, DEPOT_NODE]
                <= end_limit + 1e-9
            ):
                potential_successors[i].append(j)

    positive_cover_duals = {
        stop_id: max(cover_duals[stop_id], 0.0) for stop_id in stop_ids
    }
    positive_dual_stops = [
        stop_id for stop_id in stop_ids if positive_cover_duals[stop_id] > REDUCED_COST_TOL
    ]
    positive_dual_stop_items = [
        (
            stop_id,
            stop_mask[stop_id],
            positive_cover_duals[stop_id],
            instance.stops[stop_id].demand_kg,
            adjusted_stop_service_min_for_vehicle_type(instance, vehicle_type, stop_id),
        )
        for stop_id in positive_dual_stops
    ]
    positive_dual_index_by_stop_id = {
        stop_id: idx for idx, stop_id in enumerate(positive_dual_stops)
    }
    positive_dual_density_order = sorted(
        positive_dual_stop_items,
        key=lambda item: (
            item[2] / max(item[3], 1e-9),
            item[2],
            -instance.stops[item[0]].latest_min,
            item[0],
        ),
        reverse=True,
    )
    positive_dual_service_order = sorted(
        positive_dual_stop_items,
        key=lambda item: (
            item[2] / max(item[4], 1e-9),
            item[2],
            -instance.stops[item[0]].latest_min,
            item[0],
        ),
        reverse=True,
    )
    group_mask_list = sorted(together_group_masks)
    stop_bit_pairs = [(stop_id, stop_mask[stop_id]) for stop_id in stop_ids]
    mask_resource_cache: dict[int, tuple[float, float]] = {}
    partial_group_requirement_cache: dict[int, int] = {}
    forced_chain_cache: dict[int, tuple[int, float] | None] = {}
    mandatory_remainder_cache: dict[tuple[int, int], tuple[int, float, float] | None] = {}
    optimistic_reward_cache: dict[tuple[int, int], float] = {}

    def mask_resources(mask: int) -> tuple[float, float]:
        cached = mask_resource_cache.get(mask)
        if cached is not None:
            return cached
        demand = 0.0
        service = 0.0
        for stop_id, stop_bit in stop_bit_pairs:
            if mask & stop_bit:
                stop = instance.stops[stop_id]
                demand += stop.demand_kg
                service += adjusted_stop_service_min_for_vehicle_type(
                    instance, vehicle_type, stop_id
                )
        mask_resource_cache[mask] = (demand, service)
        return mask_resource_cache[mask]

    def partial_group_requirement_mask(visited_mask: int) -> int:
        cached = partial_group_requirement_cache.get(visited_mask)
        if cached is not None:
            return cached
        requirement_mask = 0
        for group_mask in group_mask_list:
            visited_group_mask = visited_mask & group_mask
            if visited_group_mask not in {0, group_mask}:
                requirement_mask |= group_mask & ~visited_mask
        partial_group_requirement_cache[visited_mask] = requirement_mask
        return requirement_mask

    def forced_chain_requirement(node: int) -> tuple[int, float] | None:
        if node in forced_chain_cache:
            return forced_chain_cache[node]
        if node == DEPOT_NODE:
            forced_chain_cache[node] = (0, 0.0)
            return forced_chain_cache[node]

        requirement_mask = 0
        minimum_extra_time = 0.0
        current = node
        seen: set[int] = set()
        while current in branch_ctx.forced_successor:
            next_node = branch_ctx.forced_successor[current]
            minimum_extra_time += instance.travel_minutes[current, next_node]
            if next_node == DEPOT_NODE:
                forced_chain_cache[node] = (requirement_mask, minimum_extra_time)
                return forced_chain_cache[node]
            if next_node in seen or not feasible_stop.get(next_node, False):
                forced_chain_cache[node] = None
                return None
            seen.add(next_node)
            requirement_mask |= stop_mask[next_node]
            minimum_extra_time += adjusted_stop_service_min_for_vehicle_type(
                instance, vehicle_type, next_node
            )
            current = next_node

        forced_chain_cache[node] = (requirement_mask, minimum_extra_time)
        return forced_chain_cache[node]

    label_store: list[PricingLabel] = []
    label_alive: list[bool] = []
    labels_by_state: dict[
        tuple[int, bool, int, int], list[PricingLabelBucket | None]
    ] = {}
    numpy_acceleration_enabled = np is not None
    native_dominance_helper = native_pricing_helper() if numpy_acceleration_enabled else None
    native_reward_helper = (
        native_pricing_reward_helper() if numpy_acceleration_enabled else None
    )
    native_acceleration_enabled = native_dominance_helper is not None
    native_reward_enabled = native_reward_helper is not None and bool(positive_dual_stop_items)
    if numpy_acceleration_enabled:
        label_capacity = 1024
        label_time_array = np.empty(label_capacity, dtype=np.float64)
        label_load_array = np.empty(label_capacity, dtype=np.float64)
        label_nonlabor_array = np.empty(label_capacity, dtype=np.float64)
        label_visited_mask_array = np.empty(label_capacity, dtype=np.uint64)
        label_alive_array = np.zeros(label_capacity, dtype=np.uint8)
        survivor_index_buffer = np.empty(label_capacity, dtype=np.int32)
        label_time_ptr = label_time_array.ctypes.data_as(ctypes.POINTER(ctypes.c_double))
        label_load_ptr = label_load_array.ctypes.data_as(ctypes.POINTER(ctypes.c_double))
        label_nonlabor_ptr = label_nonlabor_array.ctypes.data_as(ctypes.POINTER(ctypes.c_double))
        label_mask_ptr = label_visited_mask_array.ctypes.data_as(ctypes.POINTER(ctypes.c_uint64))
        label_alive_ptr = label_alive_array.ctypes.data_as(ctypes.POINTER(ctypes.c_uint8))
        survivor_index_buffer_ptr = survivor_index_buffer.ctypes.data_as(
            ctypes.POINTER(ctypes.c_int32)
        )
        native_survivor_count = ctypes.c_size_t(0)
        native_label_mask = ctypes.c_uint64(0)
        positive_dual_stop_bits_array = np.asarray(
            [item[1] for item in positive_dual_stop_items],
            dtype=np.uint64,
        )
        positive_dual_rewards_array = np.asarray(
            [item[2] for item in positive_dual_stop_items],
            dtype=np.float64,
        )
        positive_dual_demands_array = np.asarray(
            [item[3] for item in positive_dual_stop_items],
            dtype=np.float64,
        )
        positive_dual_services_array = np.asarray(
            [item[4] for item in positive_dual_stop_items],
            dtype=np.float64,
        )
        positive_dual_separated_masks_array = np.asarray(
            [separated_mask_by_stop[item[0]] for item in positive_dual_stop_items],
            dtype=np.uint64,
        )
        positive_dual_forced_pred_nodes_array = np.asarray(
            [
                -1
                if branch_ctx.forced_predecessor.get(item[0]) is None
                else int(branch_ctx.forced_predecessor[item[0]])
                for item in positive_dual_stop_items
            ],
            dtype=np.int32,
        )
        positive_dual_forced_pred_bits_array = np.asarray(
            [
                0
                if (
                    branch_ctx.forced_predecessor.get(item[0]) is None
                    or branch_ctx.forced_predecessor[item[0]] not in stop_mask
                )
                else stop_mask[branch_ctx.forced_predecessor[item[0]]]
                for item in positive_dual_stop_items
            ],
            dtype=np.uint64,
        )
        density_order_array = np.asarray(
            [positive_dual_index_by_stop_id[item[0]] for item in positive_dual_density_order],
            dtype=np.int32,
        )
        service_order_array = np.asarray(
            [positive_dual_index_by_stop_id[item[0]] for item in positive_dual_service_order],
            dtype=np.int32,
        )
        positive_dual_stop_bits_ptr = positive_dual_stop_bits_array.ctypes.data_as(
            ctypes.POINTER(ctypes.c_uint64)
        )
        positive_dual_rewards_ptr = positive_dual_rewards_array.ctypes.data_as(
            ctypes.POINTER(ctypes.c_double)
        )
        positive_dual_demands_ptr = positive_dual_demands_array.ctypes.data_as(
            ctypes.POINTER(ctypes.c_double)
        )
        positive_dual_services_ptr = positive_dual_services_array.ctypes.data_as(
            ctypes.POINTER(ctypes.c_double)
        )
        positive_dual_separated_masks_ptr = positive_dual_separated_masks_array.ctypes.data_as(
            ctypes.POINTER(ctypes.c_uint64)
        )
        positive_dual_forced_pred_nodes_ptr = (
            positive_dual_forced_pred_nodes_array.ctypes.data_as(ctypes.POINTER(ctypes.c_int32))
        )
        positive_dual_forced_pred_bits_ptr = positive_dual_forced_pred_bits_array.ctypes.data_as(
            ctypes.POINTER(ctypes.c_uint64)
        )
        density_order_ptr = density_order_array.ctypes.data_as(ctypes.POINTER(ctypes.c_int32))
        service_order_ptr = service_order_array.ctypes.data_as(ctypes.POINTER(ctypes.c_int32))

        def ensure_label_capacity(required_size: int) -> None:
            nonlocal label_capacity, label_time_array, label_load_array
            nonlocal label_nonlabor_array, label_visited_mask_array, label_alive_array
            nonlocal survivor_index_buffer
            nonlocal label_time_ptr, label_load_ptr, label_nonlabor_ptr
            nonlocal label_mask_ptr, label_alive_ptr, survivor_index_buffer_ptr
            if required_size <= label_capacity:
                return
            new_capacity = label_capacity
            while new_capacity < required_size:
                new_capacity *= 2
            new_time_array = np.empty(new_capacity, dtype=np.float64)
            new_time_array[: len(label_store)] = label_time_array[: len(label_store)]
            label_time_array = new_time_array
            new_load_array = np.empty(new_capacity, dtype=np.float64)
            new_load_array[: len(label_store)] = label_load_array[: len(label_store)]
            label_load_array = new_load_array
            new_nonlabor_array = np.empty(new_capacity, dtype=np.float64)
            new_nonlabor_array[: len(label_store)] = label_nonlabor_array[: len(label_store)]
            label_nonlabor_array = new_nonlabor_array
            new_mask_array = np.empty(new_capacity, dtype=np.uint64)
            new_mask_array[: len(label_store)] = label_visited_mask_array[: len(label_store)]
            label_visited_mask_array = new_mask_array
            new_alive_array = np.zeros(new_capacity, dtype=np.uint8)
            new_alive_array[: len(label_store)] = label_alive_array[: len(label_store)]
            label_alive_array = new_alive_array
            new_survivor_buffer = np.empty(new_capacity, dtype=np.int32)
            new_survivor_buffer[: len(label_store)] = survivor_index_buffer[: len(label_store)]
            survivor_index_buffer = new_survivor_buffer
            label_time_ptr = label_time_array.ctypes.data_as(ctypes.POINTER(ctypes.c_double))
            label_load_ptr = label_load_array.ctypes.data_as(ctypes.POINTER(ctypes.c_double))
            label_nonlabor_ptr = label_nonlabor_array.ctypes.data_as(
                ctypes.POINTER(ctypes.c_double)
            )
            label_mask_ptr = label_visited_mask_array.ctypes.data_as(
                ctypes.POINTER(ctypes.c_uint64)
            )
            label_alive_ptr = label_alive_array.ctypes.data_as(ctypes.POINTER(ctypes.c_uint8))
            survivor_index_buffer_ptr = survivor_index_buffer.ctypes.data_as(
                ctypes.POINTER(ctypes.c_int32)
            )
            label_capacity = new_capacity

        def bucket_index_array(bucket: PricingLabelBucket) -> Any:
            if bucket.cached_index_array is None or bucket.cache_dirty:
                if bucket.label_indices:
                    bucket.cached_index_array = np.asarray(bucket.label_indices, dtype=np.int32)
                else:
                    bucket.cached_index_array = np.empty(0, dtype=np.int32)
                bucket.cache_dirty = False
            return bucket.cached_index_array

        def refresh_bucket_from_indices(
            bucket: PricingLabelBucket,
            survivor_indices: Any,
        ) -> bool:
            if survivor_indices.size == 0:
                return False
            time_values = label_time_array[survivor_indices]
            load_values = label_load_array[survivor_indices]
            nonlabor_values = label_nonlabor_array[survivor_indices]
            mask_values = label_visited_mask_array[survivor_indices]
            bucket.label_indices = survivor_indices.tolist()
            bucket.cached_index_array = survivor_indices
            bucket.cache_dirty = False
            bucket.min_time_after_service = float(time_values.min())
            bucket.min_load_kg = float(load_values.min())
            bucket.min_nonlabor_reduced_cost = float(nonlabor_values.min())
            bucket.max_time_after_service = float(time_values.max())
            bucket.max_load_kg = float(load_values.max())
            bucket.max_nonlabor_reduced_cost = float(nonlabor_values.max())
            bucket.union_visited_mask = int(np.bitwise_or.reduce(mask_values))
            bucket.intersection_visited_mask = int(np.bitwise_and.reduce(mask_values))
            return True
    frontier: list[tuple[float, float, int]] = []
    explored_labels = 0

    def pending_required_adjacent_partner(label: PricingLabel) -> int | None:
        partner = REQUIRED_ADJACENT_STOP_PARTNER.get(label.node)
        if partner is None or partner not in stop_mask:
            return None
        if label.visited_mask & stop_mask[partner]:
            return None
        return partner

    def completion_delta_to_depot(label: PricingLabel) -> float | None:
        travel_to_depot = instance.travel_minutes[label.node, DEPOT_NODE]
        if MIDDAY_LUNCH_BREAK_RULE is None or label.lunch_break_taken:
            return travel_to_depot
        lunch_window = lunch_break_interval_after(label.time_after_service)
        if lunch_window is None:
            return None
        _, break_end = lunch_window
        return (break_end - label.time_after_service) + travel_to_depot

    def lunch_can_fit_during_wait(arrival_min: float, service_start_min: float) -> bool:
        if MIDDAY_LUNCH_BREAK_RULE is None:
            return False
        if service_start_min - arrival_min < MIDDAY_LUNCH_BREAK_RULE.duration_min - 1e-9:
            return False
        earliest_lunch_start = max(arrival_min, MIDDAY_LUNCH_BREAK_RULE.window_start_min)
        latest_lunch_start = min(
            service_start_min - MIDDAY_LUNCH_BREAK_RULE.duration_min,
            MIDDAY_LUNCH_BREAK_RULE.latest_start_min,
        )
        return earliest_lunch_start <= latest_lunch_start + 1e-9

    def label_state_key(label: PricingLabel) -> tuple[int, bool, int, int]:
        pending_partner = pending_required_adjacent_partner(label)
        return (
            label.node,
            label.lunch_break_taken,
            -1 if pending_partner is None else pending_partner,
            partial_group_requirement_mask(label.visited_mask),
        )

    def mandatory_remainder_info(label: PricingLabel) -> tuple[int, float, float] | None:
        cache_key = (label.node, label.visited_mask)
        if cache_key in mandatory_remainder_cache:
            return mandatory_remainder_cache[cache_key]

        chain_requirement = forced_chain_requirement(label.node)
        if chain_requirement is None:
            mandatory_remainder_cache[cache_key] = None
            return None
        chain_mask, chain_time_lb = chain_requirement
        if chain_mask & label.visited_mask:
            mandatory_remainder_cache[cache_key] = None
            return None

        group_mask = partial_group_requirement_mask(label.visited_mask)
        if group_mask & ~chain_mask:
            chain_terminal_node = label.node
            while chain_terminal_node in branch_ctx.forced_successor:
                next_node = branch_ctx.forced_successor[chain_terminal_node]
                if next_node == DEPOT_NODE:
                    mandatory_remainder_cache[cache_key] = None
                    return None
                chain_terminal_node = next_node

        requirement_mask = chain_mask | group_mask
        if requirement_mask & label.visited_mask:
            mandatory_remainder_cache[cache_key] = None
            return None

        for stop_id, stop_bit in stop_bit_pairs:
            if not (requirement_mask & stop_bit):
                continue
            if separated_mask_by_stop[stop_id] & (label.visited_mask | (requirement_mask ^ stop_bit)):
                mandatory_remainder_cache[cache_key] = None
                return None

        requirement_demand, requirement_service = mask_resources(requirement_mask)
        chain_demand, chain_service = mask_resources(chain_mask)
        mandatory_remainder_cache[cache_key] = (
            requirement_mask,
            requirement_demand,
            (requirement_service - chain_service) + chain_time_lb,
        )
        return mandatory_remainder_cache[cache_key]

    def optimistic_remaining_reward(label: PricingLabel) -> float:
        cache_key = (label.node, label.visited_mask)
        if cache_key in optimistic_reward_cache:
            return optimistic_reward_cache[cache_key]

        if label.node in last_forced_stops:
            optimistic_reward_cache[cache_key] = 0.0
            return 0.0

        mandatory_info = mandatory_remainder_info(label)
        if mandatory_info is None:
            optimistic_reward_cache[cache_key] = 0.0
            return 0.0
        requirement_mask, requirement_demand, _ = mandatory_info

        mandatory_reward = 0.0
        residual_capacity = vehicle_type.capacity_kg - label.load_kg
        if residual_capacity <= 1e-9:
            optimistic_reward_cache[cache_key] = 0.0
            return 0.0

        optional_capacity = max(residual_capacity - requirement_demand, 0.0)
        mandatory_requirement_time = mandatory_info[2]
        if (
            MIDDAY_LUNCH_BREAK_RULE is not None
            and not label.lunch_break_taken
            and label.time_after_service + mandatory_requirement_time
            > MIDDAY_LUNCH_BREAK_RULE.latest_start_min + 1e-9
        ):
            mandatory_requirement_time += MIDDAY_LUNCH_BREAK_RULE.duration_min
        optional_service_time = max(
            end_limit - label.time_after_service - mandatory_requirement_time,
            0.0,
        )

        if native_reward_enabled:
            reward = native_reward_helper(
                positive_dual_stop_bits_ptr,
                positive_dual_rewards_ptr,
                positive_dual_demands_ptr,
                positive_dual_services_ptr,
                positive_dual_separated_masks_ptr,
                positive_dual_forced_pred_nodes_ptr,
                positive_dual_forced_pred_bits_ptr,
                len(positive_dual_stop_items),
                density_order_ptr,
                len(positive_dual_density_order),
                service_order_ptr,
                len(positive_dual_service_order),
                ctypes.c_uint64(label.visited_mask),
                ctypes.c_uint64(requirement_mask),
                ctypes.c_int32(label.node),
                optional_capacity,
                optional_service_time,
            )
            optimistic_reward_cache[cache_key] = reward
            return reward

        optional_mask = 0
        for stop_id, stop_bit, dual_reward, _, _ in positive_dual_stop_items:
            if label.visited_mask & stop_bit:
                continue
            if separated_mask_by_stop[stop_id] & label.visited_mask:
                continue
            forced_pred = branch_ctx.forced_predecessor.get(stop_id)
            if (
                forced_pred is not None
                and forced_pred != label.node
                and forced_pred in stop_mask
                and label.visited_mask & stop_mask[forced_pred]
            ):
                continue

            if requirement_mask & stop_bit:
                mandatory_reward += dual_reward
                continue

            optional_mask |= stop_bit

        optional_capacity_reward = 0.0
        if optional_capacity > 1e-9 and optional_mask:
            for stop_id, stop_bit, dual_reward, demand, _ in positive_dual_density_order:
                if not optional_mask & stop_bit:
                    continue
                if demand <= optional_capacity + 1e-9:
                    optional_capacity_reward += dual_reward
                    optional_capacity -= demand
                    continue
                optional_capacity_reward += dual_reward * (optional_capacity / max(demand, 1e-9))
                optional_capacity = 0.0
                break

        optional_service_reward = 0.0
        if optional_service_time > 1e-9 and optional_mask:
            for stop_id, stop_bit, dual_reward, _, service_min in positive_dual_service_order:
                if not optional_mask & stop_bit:
                    continue
                if service_min <= optional_service_time + 1e-9:
                    optional_service_reward += dual_reward
                    optional_service_time -= service_min
                    continue
                optional_service_reward += dual_reward * (
                    optional_service_time / max(service_min, 1e-9)
                )
                optional_service_time = 0.0
                break

        reward = mandatory_reward + min(optional_capacity_reward, optional_service_reward)

        optimistic_reward_cache[cache_key] = reward
        return reward

    def optimistic_completion_lower_bound(label: PricingLabel) -> float:
        mandatory_info = mandatory_remainder_info(label)
        if mandatory_info is None:
            return float("inf")
        _, requirement_demand, requirement_time_lb = mandatory_info
        if MIDDAY_LUNCH_BREAK_RULE is not None and not label.lunch_break_taken:
            earliest_completion_without_break = label.time_after_service + requirement_time_lb
            if earliest_completion_without_break > MIDDAY_LUNCH_BREAK_RULE.window_start_min + 1e-9:
                requirement_time_lb += MIDDAY_LUNCH_BREAK_RULE.duration_min
        if label.load_kg + requirement_demand > vehicle_type.capacity_kg + 1e-9:
            return float("inf")
        if label.time_after_service + requirement_time_lb > end_limit + 1e-9:
            return float("inf")
        if not can_return_from(label) and not candidate_successors(label):
            return float("inf")
        return label.reduced_cost - optimistic_remaining_reward(label)

    def can_return_from(label: PricingLabel) -> bool:
        if pending_required_adjacent_partner(label) is not None:
            return False
        if label.node in branch_ctx.forced_successor:
            return False
        if (label.node, DEPOT_NODE) in branch_ctx.forbidden_arcs:
            return False
        for group_mask in together_group_masks:
            visited_group_mask = label.visited_mask & group_mask
            if visited_group_mask not in {0, group_mask}:
                return False
        completion_delta = completion_delta_to_depot(label)
        if completion_delta is None:
            return False
        return label.time_after_service + completion_delta <= end_limit + 1e-9

    def complete_reduced_cost(label: PricingLabel) -> float:
        elapsed_before = label.time_after_service - start_limit
        return_delta = completion_delta_to_depot(label)
        if return_delta is None:
            return float("inf")
        return (
            label.reduced_cost
            + instance.distance_miles[label.node, DEPOT_NODE] * distance_unit_cost
            + incremental_labor_cost(instance, elapsed_before, return_delta)
        )

    def reconstruct_route(label_idx: int) -> RouteColumn:
        sequence: list[int] = []
        cursor = label_idx
        while cursor is not None:
            current_label = label_store[cursor]
            sequence.append(current_label.node)
            cursor = current_label.predecessor_idx
        sequence.reverse()
        route = evaluate_route_sequence(instance, vehicle_type, tuple(sequence))
        if route is None:
            raise RuntimeError(
                f"Label-setting pricing reconstructed an invalid route for "
                f"{vehicle_type.type_id}: {sequence}"
            )
        if not route_is_compatible(route, branch_ctx):
            raise RuntimeError(
                f"Label-setting pricing reconstructed a branch-incompatible route for "
                f"{vehicle_type.type_id}: {sequence}"
            )
        return route

    def exact_route_reduced_cost(route: RouteColumn) -> float:
        return route_reduced_cost(route, cover_duals=cover_duals, type_dual=type_dual)

    def add_label(label: PricingLabel) -> int | None:
        if (
            MIDDAY_LUNCH_BREAK_RULE is not None
            and not label.lunch_break_taken
            and label.time_after_service > MIDDAY_LUNCH_BREAK_RULE.latest_start_min + 1e-9
        ):
            return None
        optimistic_lb = optimistic_completion_lower_bound(label)
        if optimistic_lb >= -REDUCED_COST_TOL:
            return None

        label_time = label.time_after_service
        label_load = label.load_kg
        label_nonlabor = label.nonlabor_reduced_cost
        label_mask = label.visited_mask
        label_alive_local = label_alive
        label_store_local = label_store
        if numpy_acceleration_enabled:
            label_mask_u64 = np.uint64(label_mask)
            not_label_mask_u64 = np.bitwise_not(label_mask_u64)
        state_key = label_state_key(label)
        visited_count = label.visited_mask.bit_count()
        state_buckets = labels_by_state.get(state_key)
        if state_buckets is None:
            state_buckets = [None] * (len(stop_ids) + 1)
            labels_by_state[state_key] = state_buckets
        for incumbent_count, incumbent_bucket in enumerate(state_buckets):
            if incumbent_bucket is None:
                continue
            incumbent_can_dominate = incumbent_count <= visited_count
            label_can_dominate = incumbent_count >= visited_count
            if incumbent_can_dominate and (
                incumbent_bucket.min_time_after_service > label_time + 1e-9
                or incumbent_bucket.min_load_kg > label_load + 1e-9
                or incumbent_bucket.min_nonlabor_reduced_cost
                > label_nonlabor + 1e-9
                or (incumbent_bucket.intersection_visited_mask & ~label_mask) != 0
            ):
                incumbent_can_dominate = False
            if label_can_dominate and (
                incumbent_bucket.max_time_after_service < label_time - 1e-9
                or incumbent_bucket.max_load_kg < label_load - 1e-9
                or incumbent_bucket.max_nonlabor_reduced_cost
                < label_nonlabor - 1e-9
                or (label_mask & ~incumbent_bucket.union_visited_mask) != 0
            ):
                label_can_dominate = False
            if not incumbent_can_dominate and not label_can_dominate:
                continue
            use_numpy_bucket = (
                numpy_acceleration_enabled
                and len(incumbent_bucket.label_indices) >= EXACT_PRICING_NUMPY_MIN_BUCKET_SIZE
            )
            use_native_bucket = (
                native_acceleration_enabled
                and len(incumbent_bucket.label_indices) >= EXACT_PRICING_NATIVE_MIN_BUCKET_SIZE
            )
            if use_native_bucket:
                idx_array = bucket_index_array(incumbent_bucket)
                if idx_array.size == 0:
                    state_buckets[incumbent_count] = None
                    continue

                survivor_count = ctypes.c_size_t(0)
                native_survivor_count.value = 0
                native_label_mask.value = label_mask
                incumbent_dominates = native_dominance_helper(
                    idx_array.ctypes.data_as(ctypes.POINTER(ctypes.c_int32)),
                    idx_array.size,
                    label_time_ptr,
                    label_load_ptr,
                    label_nonlabor_ptr,
                    label_mask_ptr,
                    label_alive_ptr,
                    label_time,
                    label_load,
                    label_nonlabor,
                    native_label_mask,
                    int(incumbent_can_dominate),
                    int(label_can_dominate),
                    survivor_index_buffer_ptr,
                    ctypes.byref(native_survivor_count),
                )
                if incumbent_dominates:
                    return None

                survivor_total = native_survivor_count.value
                if survivor_total != idx_array.size:
                    removed_indices = idx_array[label_alive_array[idx_array] == 0]
                    for removed_idx in removed_indices.tolist():
                        label_alive_local[removed_idx] = False
                    survivor_indices = survivor_index_buffer[:survivor_total].copy()
                    if not refresh_bucket_from_indices(incumbent_bucket, survivor_indices):
                        state_buckets[incumbent_count] = None
                continue
            if use_numpy_bucket:
                idx_array = bucket_index_array(incumbent_bucket)
                if idx_array.size == 0:
                    state_buckets[incumbent_count] = None
                    continue

                alive_mask = label_alive_array[idx_array] != 0
                if not bool(np.all(alive_mask)):
                    survivor_indices = idx_array[alive_mask]
                    if not refresh_bucket_from_indices(incumbent_bucket, survivor_indices):
                        state_buckets[incumbent_count] = None
                        continue
                    idx_array = survivor_indices

                incumbent_times = label_time_array[idx_array]
                incumbent_loads = label_load_array[idx_array]
                incumbent_nonlabor = label_nonlabor_array[idx_array]
                incumbent_masks = label_visited_mask_array[idx_array]

                if incumbent_can_dominate:
                    incumbent_dominates_mask = (
                        (incumbent_times <= label_time + 1e-9)
                        & (incumbent_loads <= label_load + 1e-9)
                        & (incumbent_nonlabor <= label_nonlabor + 1e-9)
                        & (np.bitwise_and(incumbent_masks, not_label_mask_u64) == 0)
                    )
                    if bool(np.any(incumbent_dominates_mask)):
                        return None

                if label_can_dominate:
                    label_dominates_mask = (
                        (label_time <= incumbent_times + 1e-9)
                        & (label_load <= incumbent_loads + 1e-9)
                        & (label_nonlabor <= incumbent_nonlabor + 1e-9)
                        & (
                            np.bitwise_and(
                                label_mask_u64,
                                np.bitwise_not(incumbent_masks),
                            )
                            == 0
                        )
                    )
                    if bool(np.any(label_dominates_mask)):
                        dominated_indices = idx_array[label_dominates_mask]
                        label_alive_array[dominated_indices] = 0
                        for dominated_idx in dominated_indices.tolist():
                            label_alive_local[dominated_idx] = False
                        survivor_indices = idx_array[~label_dominates_mask]
                        if not refresh_bucket_from_indices(incumbent_bucket, survivor_indices):
                            state_buckets[incumbent_count] = None
                continue
            if incumbent_can_dominate and not label_can_dominate:
                bucket_needs_rebuild = False
                for idx in incumbent_bucket.label_indices:
                    if not label_alive_local[idx]:
                        bucket_needs_rebuild = True
                        break
                    incumbent = label_store_local[idx]
                    if (
                        incumbent.time_after_service <= label_time + 1e-9
                        and incumbent.load_kg <= label_load + 1e-9
                        and incumbent.nonlabor_reduced_cost <= label_nonlabor + 1e-9
                        and (incumbent.visited_mask & ~label_mask) == 0
                    ):
                        return None
                if not bucket_needs_rebuild:
                    continue
            elif label_can_dominate and not incumbent_can_dominate:
                bucket_needs_rebuild = False
                for idx in incumbent_bucket.label_indices:
                    if not label_alive_local[idx]:
                        bucket_needs_rebuild = True
                        break
                    incumbent = label_store_local[idx]
                    if (
                        label_time <= incumbent.time_after_service + 1e-9
                        and label_load <= incumbent.load_kg + 1e-9
                        and label_nonlabor <= incumbent.nonlabor_reduced_cost + 1e-9
                        and (label_mask & ~incumbent.visited_mask) == 0
                    ):
                        label_alive_local[idx] = False
                        bucket_needs_rebuild = True
                if not bucket_needs_rebuild:
                    continue

            surviving_bucket: list[int] = []
            bucket_min_time = float("inf")
            bucket_min_load = float("inf")
            bucket_min_nonlabor = float("inf")
            bucket_max_time = -float("inf")
            bucket_max_load = -float("inf")
            bucket_max_nonlabor = -float("inf")
            bucket_union_mask = 0
            bucket_intersection_mask = 0
            first_survivor = True
            bucket_changed = False
            for idx in incumbent_bucket.label_indices:
                if not label_alive_local[idx]:
                    bucket_changed = True
                    continue
                incumbent = label_store_local[idx]
                if (
                    incumbent_can_dominate
                    and incumbent.time_after_service <= label_time + 1e-9
                    and incumbent.load_kg <= label_load + 1e-9
                    and incumbent.nonlabor_reduced_cost <= label_nonlabor + 1e-9
                    and (incumbent.visited_mask & ~label_mask) == 0
                ):
                    return None
                if (
                    label_can_dominate
                    and label_time <= incumbent.time_after_service + 1e-9
                    and label_load <= incumbent.load_kg + 1e-9
                    and label_nonlabor <= incumbent.nonlabor_reduced_cost + 1e-9
                    and (label_mask & ~incumbent.visited_mask) == 0
                ):
                    label_alive_local[idx] = False
                    bucket_changed = True
                    continue
                surviving_bucket.append(idx)
                if incumbent.time_after_service < bucket_min_time:
                    bucket_min_time = incumbent.time_after_service
                if incumbent.load_kg < bucket_min_load:
                    bucket_min_load = incumbent.load_kg
                if incumbent.nonlabor_reduced_cost < bucket_min_nonlabor:
                    bucket_min_nonlabor = incumbent.nonlabor_reduced_cost
                if incumbent.time_after_service > bucket_max_time:
                    bucket_max_time = incumbent.time_after_service
                if incumbent.load_kg > bucket_max_load:
                    bucket_max_load = incumbent.load_kg
                if incumbent.nonlabor_reduced_cost > bucket_max_nonlabor:
                    bucket_max_nonlabor = incumbent.nonlabor_reduced_cost
                if first_survivor:
                    bucket_union_mask = incumbent.visited_mask
                    bucket_intersection_mask = incumbent.visited_mask
                    first_survivor = False
                else:
                    bucket_union_mask |= incumbent.visited_mask
                    bucket_intersection_mask &= incumbent.visited_mask
            if surviving_bucket:
                if bucket_changed:
                    state_buckets[incumbent_count] = PricingLabelBucket(
                        label_indices=surviving_bucket,
                        min_time_after_service=bucket_min_time,
                        min_load_kg=bucket_min_load,
                        min_nonlabor_reduced_cost=bucket_min_nonlabor,
                        max_time_after_service=bucket_max_time,
                        max_load_kg=bucket_max_load,
                        max_nonlabor_reduced_cost=bucket_max_nonlabor,
                        union_visited_mask=bucket_union_mask,
                        intersection_visited_mask=bucket_intersection_mask,
                        cached_index_array=(
                            np.asarray(surviving_bucket, dtype=np.int32)
                            if numpy_acceleration_enabled
                            else None
                        ),
                        cache_dirty=False,
                    )
            else:
                state_buckets[incumbent_count] = None

        label_idx = len(label_store)
        if numpy_acceleration_enabled:
            ensure_label_capacity(label_idx + 1)
        label_store.append(label)
        label_alive.append(True)
        if numpy_acceleration_enabled:
            label_time_array[label_idx] = label_time
            label_load_array[label_idx] = label_load
            label_nonlabor_array[label_idx] = label_nonlabor
            label_visited_mask_array[label_idx] = label_mask_u64
            label_alive_array[label_idx] = 1
        current_bucket = state_buckets[visited_count]
        if current_bucket is None:
            state_buckets[visited_count] = PricingLabelBucket(
                label_indices=[label_idx],
                min_time_after_service=label.time_after_service,
                min_load_kg=label.load_kg,
                min_nonlabor_reduced_cost=label.nonlabor_reduced_cost,
                max_time_after_service=label.time_after_service,
                max_load_kg=label.load_kg,
                max_nonlabor_reduced_cost=label.nonlabor_reduced_cost,
                union_visited_mask=label_mask,
                intersection_visited_mask=label_mask,
                cached_index_array=(
                    np.asarray([label_idx], dtype=np.int32) if numpy_acceleration_enabled else None
                ),
                cache_dirty=False,
            )
        else:
            current_bucket.label_indices.append(label_idx)
            if label.time_after_service < current_bucket.min_time_after_service:
                current_bucket.min_time_after_service = label.time_after_service
            if label.load_kg < current_bucket.min_load_kg:
                current_bucket.min_load_kg = label.load_kg
            if label.nonlabor_reduced_cost < current_bucket.min_nonlabor_reduced_cost:
                current_bucket.min_nonlabor_reduced_cost = label.nonlabor_reduced_cost
            if label.time_after_service > current_bucket.max_time_after_service:
                current_bucket.max_time_after_service = label.time_after_service
            if label.load_kg > current_bucket.max_load_kg:
                current_bucket.max_load_kg = label.load_kg
            if label.nonlabor_reduced_cost > current_bucket.max_nonlabor_reduced_cost:
                current_bucket.max_nonlabor_reduced_cost = label.nonlabor_reduced_cost
            current_bucket.union_visited_mask |= label_mask
            current_bucket.intersection_visited_mask &= label_mask
            if numpy_acceleration_enabled:
                if current_bucket.cached_index_array is not None and not current_bucket.cache_dirty:
                    current_bucket.cached_index_array = np.append(
                        current_bucket.cached_index_array,
                        np.int32(label_idx),
                    )
                else:
                    current_bucket.cache_dirty = True
        heapq.heappush(frontier, (optimistic_lb, label.reduced_cost, label_idx))
        return label_idx

    def candidate_successors(label: PricingLabel) -> list[int]:
        forced_next = branch_ctx.forced_successor.get(label.node)
        required_partner = pending_required_adjacent_partner(label)
        if forced_next is not None:
            if (
                (required_partner is None or forced_next == required_partner)
                and forced_next in stop_mask
                and label.visited_mask & stop_mask[forced_next] == 0
                and forced_next in potential_successors[label.node]
            ):
                return [forced_next]
            return []

        if required_partner is not None:
            forced_pred = branch_ctx.forced_predecessor.get(required_partner)
            if (
                required_partner in stop_mask
                and label.visited_mask & stop_mask[required_partner] == 0
                and required_partner in potential_successors[label.node]
                and not (separated_mask_by_stop[required_partner] & label.visited_mask)
                and (forced_pred is None or forced_pred == label.node)
            ):
                return [required_partner]
            return []

        if label.node in last_forced_stops:
            return []

        candidates: list[int] = []
        for next_stop in potential_successors[label.node]:
            if label.visited_mask & stop_mask[next_stop]:
                continue
            if separated_mask_by_stop[next_stop] & label.visited_mask:
                continue
            forced_pred = branch_ctx.forced_predecessor.get(next_stop)
            if forced_pred is not None and forced_pred != label.node:
                continue
            candidates.append(next_stop)
        return candidates

    improving_routes: list[tuple[float, RouteColumn]] = []
    improving_stop_sequences: set[tuple[int, ...]] = set()

    def maybe_add_route(label_idx: int) -> bool:
        label = label_store[label_idx]
        if not can_return_from(label):
            return False
        if complete_reduced_cost(label) >= -REDUCED_COST_TOL:
            return False
        route = reconstruct_route(label_idx)
        if (
            route.stop_sequence in excluded_stop_sequences
            or route.stop_sequence in improving_stop_sequences
        ):
            return False
        exact_rc = exact_route_reduced_cost(route)
        if exact_rc >= -REDUCED_COST_TOL:
            return False
        improving_stop_sequences.add(route.stop_sequence)
        improving_routes.append((exact_rc, route))
        return len(improving_routes) >= max_routes

    if progress is not None:
        progress.emit(
            f"{progress_label or f'pricing {vehicle_type.type_id}'}: starting label-setting "
            f"pricing over {sum(feasible_stop.values())} feasible stops"
        )

    for start_stop in potential_successors[DEPOT_NODE]:
        forced_pred = branch_ctx.forced_predecessor.get(start_stop)
        if forced_pred is not None and forced_pred != DEPOT_NODE:
            continue

        stop = instance.stops[start_stop]
        service_min = adjusted_stop_service_min_for_vehicle_type(
            instance, vehicle_type, start_stop
        )
        arrival = start_limit + instance.travel_minutes[DEPOT_NODE, start_stop]
        service_start = hard_service_start_after_arrival(start_stop, arrival)
        time_after_service = service_start + service_min
        start_transition_options: list[tuple[bool, float, float]] = []
        if not hard_service_start_is_feasible(start_stop, service_start):
            pass
        elif MIDDAY_LUNCH_BREAK_RULE is None:
            start_transition_options.append((False, service_start, time_after_service))
        elif lunch_can_fit_during_wait(arrival, service_start):
            start_transition_options.append((True, service_start, time_after_service))
        elif time_after_service <= MIDDAY_LUNCH_BREAK_RULE.latest_start_min + 1e-9:
            start_transition_options.append((False, service_start, time_after_service))

        lunch_window = lunch_break_interval_after(start_limit)
        if (
            MIDDAY_LUNCH_BREAK_RULE is not None
            and lunch_window is not None
            and not lunch_can_fit_during_wait(arrival, service_start)
        ):
            _, break_end = lunch_window
            lunch_arrival = break_end + instance.travel_minutes[DEPOT_NODE, start_stop]
            lunch_service_start = hard_service_start_after_arrival(start_stop, lunch_arrival)
            if hard_service_start_is_feasible(start_stop, lunch_service_start):
                lunch_time_after_service = lunch_service_start + service_min
                start_transition_options.append((True, lunch_service_start, lunch_time_after_service))

        for lunch_break_taken, service_start_min, service_finish_min in start_transition_options:
            if service_finish_min + instance.travel_minutes[start_stop, DEPOT_NODE] > end_limit + 1e-9:
                continue
            elapsed = service_finish_min - start_limit
            reduced_cost = (
                vehicle_type.fixed_daily_cost
                - type_dual
                + instance.distance_miles[DEPOT_NODE, start_stop] * distance_unit_cost
                + incremental_labor_cost(instance, 0.0, elapsed)
                + max(service_start_min - stop.latest_min, 0.0)
                / 60.0
                * instance.cost_params["Late_Delivery_Penalty_per_Hour"]
                - cover_duals[start_stop]
            )

            label_idx = add_label(
                PricingLabel(
                    node=start_stop,
                    visited_mask=stop_mask[start_stop],
                    load_kg=stop.demand_kg,
                    time_after_service=service_finish_min,
                    lunch_break_taken=lunch_break_taken,
                    reduced_cost=reduced_cost,
                    nonlabor_reduced_cost=reduced_cost - operating_labor_cost(instance, elapsed),
                    predecessor_idx=None,
                )
            )
            if label_idx is not None and maybe_add_route(label_idx):
                return [route for _, route in sorted(improving_routes, key=lambda item: item[0])]

    while frontier:
        _, _, label_idx = heapq.heappop(frontier)
        if not label_alive[label_idx]:
            continue

        explored_labels += 1
        if explored_labels % 250 == 0:
            check_time_limit()
            if progress is not None:
                progress.emit(
                    f"{progress_label or f'pricing {vehicle_type.type_id}'}: explored "
                    f"{explored_labels} labels, frontier {len(frontier)}"
                )

        label = label_store[label_idx]
        for next_stop in candidate_successors(label):
            next_stop_data = instance.stops[next_stop]
            next_stop_service_min = adjusted_stop_service_min_for_vehicle_type(
                instance, vehicle_type, next_stop
            )
            new_load = label.load_kg + next_stop_data.demand_kg
            if new_load > vehicle_type.capacity_kg + 1e-9:
                continue

            elapsed_before = label.time_after_service - start_limit
            transition_options: list[tuple[bool, float, float, float]] = []

            arrival = label.time_after_service + instance.travel_minutes[label.node, next_stop]
            service_start = hard_service_start_after_arrival(next_stop, arrival)
            time_after_service = service_start + next_stop_service_min
            if not hard_service_start_is_feasible(next_stop, service_start):
                pass
            elif MIDDAY_LUNCH_BREAK_RULE is None or label.lunch_break_taken:
                transition_options.append(
                    (
                        label.lunch_break_taken,
                        service_start,
                        time_after_service,
                        time_after_service - label.time_after_service,
                    )
                )
            elif lunch_can_fit_during_wait(arrival, service_start):
                transition_options.append(
                    (
                        True,
                        service_start,
                        time_after_service,
                        time_after_service - label.time_after_service,
                    )
                )
            elif time_after_service <= MIDDAY_LUNCH_BREAK_RULE.latest_start_min + 1e-9:
                transition_options.append(
                    (
                        False,
                        service_start,
                        time_after_service,
                        time_after_service - label.time_after_service,
                    )
                )

            if (
                MIDDAY_LUNCH_BREAK_RULE is not None
                and not label.lunch_break_taken
                and not lunch_can_fit_during_wait(arrival, service_start)
            ):
                lunch_window = lunch_break_interval_after(label.time_after_service)
                if lunch_window is not None:
                    _, break_end = lunch_window
                    lunch_arrival = break_end + instance.travel_minutes[label.node, next_stop]
                    lunch_service_start = hard_service_start_after_arrival(next_stop, lunch_arrival)
                    if hard_service_start_is_feasible(next_stop, lunch_service_start):
                        lunch_time_after_service = lunch_service_start + next_stop_service_min
                        transition_options.append(
                            (
                                True,
                                lunch_service_start,
                                lunch_time_after_service,
                                lunch_time_after_service - label.time_after_service,
                            )
                        )

            for lunch_break_taken, service_start_min, service_finish_min, delta_minutes in transition_options:
                if service_finish_min + instance.travel_minutes[next_stop, DEPOT_NODE] > end_limit + 1e-9:
                    continue

                new_reduced_cost = (
                    label.reduced_cost
                    + instance.distance_miles[label.node, next_stop] * distance_unit_cost
                    + incremental_labor_cost(instance, elapsed_before, delta_minutes)
                    + max(service_start_min - next_stop_data.latest_min, 0.0)
                    / 60.0
                    * instance.cost_params["Late_Delivery_Penalty_per_Hour"]
                    - cover_duals[next_stop]
                )

                new_label_idx = add_label(
                    PricingLabel(
                        node=next_stop,
                        visited_mask=label.visited_mask | stop_mask[next_stop],
                        load_kg=new_load,
                        time_after_service=service_finish_min,
                        lunch_break_taken=lunch_break_taken,
                        reduced_cost=new_reduced_cost,
                        nonlabor_reduced_cost=(
                            label.nonlabor_reduced_cost
                            + instance.distance_miles[label.node, next_stop] * distance_unit_cost
                            + max(service_start_min - next_stop_data.latest_min, 0.0)
                            / 60.0
                            * instance.cost_params["Late_Delivery_Penalty_per_Hour"]
                            - cover_duals[next_stop]
                        ),
                        predecessor_idx=label_idx,
                    )
                )
                if new_label_idx is None:
                    continue

                if maybe_add_route(new_label_idx):
                    return [route for _, route in sorted(improving_routes, key=lambda item: item[0])]

    check_time_limit()
    return [route for _, route in sorted(improving_routes, key=lambda item: item[0])]


def build_master_model(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    routes: list[RouteColumn],
    relax: bool,
    allow_artificial: bool = False,
) -> tuple[gp.Model, dict[str, Any]]:
    model = create_gurobi_model("route_based_master")
    model.Params.OutputFlag = 0

    route_indices = list(range(len(routes)))
    var_type = GRB.CONTINUOUS if relax else GRB.BINARY
    route_vars = model.addVars(route_indices, lb=0.0, ub=1.0, vtype=var_type, name="route")
    artificial_vars = (
        model.addVars(
            sorted(instance.stops),
            lb=0.0,
            ub=1.0,
            vtype=GRB.CONTINUOUS if relax else GRB.BINARY,
            name="artificial_cover",
        )
        if allow_artificial
        else {}
    )

    cover_constraints: dict[int, gp.Constr] = {}
    for stop_id in sorted(instance.stops):
        candidate_routes = [idx for idx, route in enumerate(routes) if stop_id in route.stop_set]
        if not candidate_routes and not allow_artificial:
            raise ValueError(f"No route column covers stop {stop_id}.")
        cover_expr = gp.quicksum(route_vars[idx] for idx in candidate_routes)
        if allow_artificial:
            cover_expr += artificial_vars[stop_id]
        cover_constraints[stop_id] = model.addConstr(
            cover_expr == 1,
            name=f"cover[{stop_id}]",
        )

    type_constraints: dict[str, gp.Constr] = {}
    for type_id, vehicle_type in vehicle_types.items():
        candidate_routes = [idx for idx, route in enumerate(routes) if route.type_id == type_id]
        type_constraints[type_id] = model.addConstr(
            gp.quicksum(route_vars[idx] for idx in candidate_routes) <= vehicle_type.count,
            name=f"type_limit[{type_id}]",
        )

    model.setObjective(
        gp.quicksum(routes[idx].cost * route_vars[idx] for idx in route_indices)
        + (
            gp.quicksum(
                ARTIFICIAL_COVER_PENALTY * artificial_vars[stop_id]
                for stop_id in sorted(instance.stops)
            )
            if allow_artificial
            else 0.0
        ),
        GRB.MINIMIZE,
    )

    model._data = {
        "route_vars": route_vars,
        "artificial_vars": artificial_vars,
        "cover_constraints": cover_constraints,
        "type_constraints": type_constraints,
        "routes": routes,
        "vehicle_types": vehicle_types,
    }
    return model, model._data


def build_incremental_node_master(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    routes: list[RouteColumn],
) -> NodeMasterLP:
    model = create_gurobi_model("route_based_node_master")
    model.Params.OutputFlag = 0
    model.ModelSense = GRB.MINIMIZE

    artificial_vars = model.addVars(
        sorted(instance.stops),
        lb=0.0,
        ub=1.0,
        obj=ARTIFICIAL_COVER_PENALTY,
        vtype=GRB.CONTINUOUS,
        name="artificial_cover",
    )

    cover_constraints: dict[int, gp.Constr] = {}
    for stop_id in sorted(instance.stops):
        cover_constraints[stop_id] = model.addConstr(
            artificial_vars[stop_id] == 1.0,
            name=f"cover[{stop_id}]",
        )

    type_constraints: dict[str, gp.Constr] = {}
    for type_id, vehicle_type in vehicle_types.items():
        type_constraints[type_id] = model.addConstr(
            gp.LinExpr() <= vehicle_type.count,
            name=f"type_limit[{type_id}]",
        )

    node_master = NodeMasterLP(
        model=model,
        routes=[],
        route_vars=[],
        artificial_vars={stop_id: artificial_vars[stop_id] for stop_id in sorted(instance.stops)},
        cover_constraints=cover_constraints,
        type_constraints=type_constraints,
    )
    for route in routes:
        add_route_column_to_master(node_master, route)
    model.update()
    return node_master


def add_route_column_to_master(node_master: NodeMasterLP, route: RouteColumn) -> gp.Var:
    column = gp.Column()
    for stop_id in route.stop_set:
        column.addTerms(1.0, node_master.cover_constraints[stop_id])
    column.addTerms(1.0, node_master.type_constraints[route.type_id])

    route_index = len(node_master.routes)
    route_var = node_master.model.addVar(
        lb=0.0,
        ub=1.0,
        obj=route.cost,
        vtype=GRB.CONTINUOUS,
        column=column,
        name=f"route[{route_index}]",
    )
    node_master.routes.append(route)
    node_master.route_vars.append(route_var)
    return route_var


def node_pricing_batch_size(node_label: str, base_count: int) -> int:
    if node_label in {"root", "node depth 0"}:
        return max(base_count * ROOT_PRICING_COLUMNS_MULTIPLIER, ROOT_PRICING_COLUMNS_MIN)
    return base_count


def solve_node_relaxation(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    route_pool: list[RouteColumn],
    route_key_set: set[tuple[str, tuple[int, ...]]],
    branch_state: BranchState,
    max_cg_iterations: int | None,
    pricing_time_limit: float | None,
    pricing_columns_per_type: int,
    pricing_workers: int | None,
    dual_stabilization_alpha: float,
    dual_stabilization_mode: str,
    deadline: float | None,
    progress: ProgressTracker | None = None,
    node_label: str = "root",
    resume_state: NodeRelaxationCheckpointState | None = None,
    checkpoint_callback: Callable[[NodeRelaxationCheckpointState], None] | None = None,
) -> NodeRelaxationResult | None:
    branch_ctx = build_branch_context(branch_state, instance, vehicle_types)
    if branch_ctx is None:
        return None

    compatible_routes = [route for route in route_pool if route_is_compatible(route, branch_ctx)]
    compatible_route_lookup = {
        route_column_key(route): route for route in compatible_routes
    }
    if resume_state is None:
        if node_label in {"root", "node depth 0"}:
            active_routes = compatible_routes[:]
        else:
            active_routes = select_initial_active_routes(
                instance=instance,
                vehicle_types=vehicle_types,
                compatible_routes=compatible_routes,
            )
        if not active_routes:
            active_routes = compatible_routes[:]
        last_lp_objective = float("nan")
        iteration = 0
        stabilized_cover_duals: dict[int, float] | None = None
        stabilized_type_duals: dict[str, float] | None = None
        previous_raw_cover_duals: dict[int, float] | None = None
        previous_raw_type_duals: dict[str, float] | None = None
    else:
        active_routes = []
        for route in resume_state.active_routes:
            key = route_column_key(route)
            resumed_route = compatible_route_lookup.get(key)
            if resumed_route is None:
                raise ValueError(
                    f"Checkpoint route {key} is no longer compatible with branch state {node_label}."
                )
            active_routes.append(resumed_route)
        if not active_routes:
            active_routes = compatible_routes[:]
        last_lp_objective = (
            resume_state.last_lp_objective
            if resume_state.last_lp_objective is not None
            else float("nan")
        )
        iteration = resume_state.iteration
        stabilized_cover_duals = (
            dict(resume_state.stabilized_cover_duals)
            if resume_state.stabilized_cover_duals is not None
            else None
        )
        stabilized_type_duals = (
            dict(resume_state.stabilized_type_duals)
            if resume_state.stabilized_type_duals is not None
            else None
        )
        previous_raw_cover_duals = (
            dict(resume_state.previous_raw_cover_duals)
            if resume_state.previous_raw_cover_duals is not None
            else None
        )
        previous_raw_type_duals = (
            dict(resume_state.previous_raw_type_duals)
            if resume_state.previous_raw_type_duals is not None
            else None
        )
    node_master = build_incremental_node_master(
        instance=instance,
        vehicle_types=vehicle_types,
        routes=active_routes,
    )
    if progress is not None:
        if resume_state is None:
            progress.emit(
                f"{node_label}: starting node relaxation with {len(active_routes)} active "
                f"columns out of {len(compatible_routes)} compatible global columns",
                force=True,
            )
        else:
            progress.emit(
                f"{node_label}: resuming node relaxation from CG iteration {iteration} with "
                f"{len(active_routes)} active columns out of {len(compatible_routes)} "
                "compatible global columns",
                force=True,
            )

    route_sequences_by_type: dict[str, set[tuple[int, ...]]] = {
        type_id: set() for type_id in vehicle_types
    }
    for type_id, stop_sequence in route_key_set:
        if type_id in route_sequences_by_type:
            route_sequences_by_type[type_id].add(stop_sequence)

    inactive_routes_by_type: dict[str, dict[tuple[str, tuple[int, ...]], RouteColumn]] = {
        type_id: {} for type_id in vehicle_types
    }
    active_route_keys = {route_column_key(route) for route in active_routes}
    for route in compatible_routes:
        key = route_column_key(route)
        if key not in active_route_keys:
            inactive_routes_by_type[route.type_id][key] = route

    root_full_pricing = node_label in {"root", "node depth 0"}
    effective_pricing_workers = (
        min(pricing_workers, len(vehicle_types))
        if pricing_workers is not None
        else min(os.cpu_count() or 1, len(vehicle_types))
    )
    if ROUTE_BASED_PARTIAL_PRICING and not root_full_pricing:
        effective_pricing_workers = 1
    pricing_executor: ProcessPoolExecutor | None = None
    if effective_pricing_workers > 1:
        pricing_executor = ProcessPoolExecutor(
            max_workers=effective_pricing_workers,
            mp_context=mp.get_context("spawn"),
            initializer=init_pricing_worker,
            initargs=(instance, branch_ctx),
        )
        if progress is not None:
            progress.emit(
                f"{node_label}: pricing vehicle types in parallel with "
                f"{effective_pricing_workers} worker processes",
                force=True,
            )
    elif progress is not None and ROUTE_BASED_PARTIAL_PRICING:
        if root_full_pricing:
            progress.emit(
                f"{node_label}: using sequential full pricing sweeps across vehicle types",
                force=True,
            )
        else:
            progress.emit(
                f"{node_label}: using sequential partial pricing sweeps across vehicle types",
                force=True,
            )

    pricing_batch_size = node_pricing_batch_size(node_label, pricing_columns_per_type)

    vehicle_type_list = list(vehicle_types.values())
    pricing_start_index = 0

    def price_vehicle_type(
        vehicle_type: VehicleTypeData,
        *,
        cover_duals: dict[int, float],
        type_dual: float,
        raw_pricing: bool,
        iteration_number: int,
    ) -> list[RouteColumn]:
        if progress is not None:
            prefix = "raw pricing" if raw_pricing else "pricing"
            progress.emit(
                f"{node_label}: CG iteration {iteration_number}, {prefix} "
                f"{vehicle_type.type_id} at LP bound {last_lp_objective:.2f}"
            )
        return solve_pricing_subproblem(
            instance=instance,
            vehicle_type=vehicle_type,
            cover_duals=cover_duals,
            type_dual=type_dual,
            branch_ctx=branch_ctx,
            max_routes=pricing_batch_size,
            excluded_stop_sequences=route_sequences_by_type[vehicle_type.type_id],
            pricing_time_limit=pricing_time_limit,
            deadline=deadline,
            progress=progress,
            progress_label=(
                f"{node_label}: {'raw pricing' if raw_pricing else 'pricing'} "
                f"{vehicle_type.type_id} in CG iteration {iteration_number}"
            ),
            exact_fallback_max_routes=(
                ROOT_EXACT_PRICING_FALLBACK_MAX_ROUTES
                if node_label in {"root", "node depth 0"}
                else EXACT_PRICING_FALLBACK_MAX_ROUTES
            ),
        )

    def collect_candidate_routes(
        vehicle_types_to_price: list[VehicleTypeData],
        cover_duals: dict[int, float],
        type_duals: dict[str, float],
        *,
        raw_pricing: bool,
        iteration_number: int,
    ) -> dict[str, list[RouteColumn]]:
        if pricing_executor is None:
            routes_by_type: dict[str, list[RouteColumn]] = {}
            for vehicle_type in vehicle_types_to_price:
                routes_by_type[vehicle_type.type_id] = price_vehicle_type(
                    vehicle_type,
                    cover_duals=cover_duals,
                    type_dual=type_duals[vehicle_type.type_id],
                    raw_pricing=raw_pricing,
                    iteration_number=iteration_number,
                )
            return routes_by_type

        if progress is not None:
            progress.emit(
                f"{node_label}: CG iteration {iteration_number}, running "
                f"{'raw ' if raw_pricing else ''}pricing for {len(vehicle_types_to_price)} vehicle types "
                f"in parallel"
            )

        future_by_type_id = {
            vehicle_type.type_id: pricing_executor.submit(
                run_pricing_worker,
                vehicle_type,
                cover_duals,
                type_duals[vehicle_type.type_id],
                pricing_batch_size,
                set(route_sequences_by_type[vehicle_type.type_id]),
                pricing_time_limit,
                deadline,
            )
            for vehicle_type in vehicle_types_to_price
        }
        routes_by_type: dict[str, list[RouteColumn]] = {}
        for vehicle_type in vehicle_types_to_price:
            type_id, routes = future_by_type_id[vehicle_type.type_id].result()
            routes_by_type[type_id] = routes
        return routes_by_type

    def emit_checkpoint() -> None:
        if checkpoint_callback is None:
            return
        checkpoint_callback(
            NodeRelaxationCheckpointState(
                iteration=iteration,
                last_lp_objective=(
                    None if math.isnan(last_lp_objective) else last_lp_objective
                ),
                active_routes=list(node_master.routes),
                stabilized_cover_duals=(
                    None if stabilized_cover_duals is None else dict(stabilized_cover_duals)
                ),
                stabilized_type_duals=(
                    None if stabilized_type_duals is None else dict(stabilized_type_duals)
                ),
                previous_raw_cover_duals=(
                    None
                    if previous_raw_cover_duals is None
                    else dict(previous_raw_cover_duals)
                ),
                previous_raw_type_duals=(
                    None
                    if previous_raw_type_duals is None
                    else dict(previous_raw_type_duals)
                ),
            )
        )

    try:
        while max_cg_iterations is None or iteration < max_cg_iterations:
            iteration += 1
            if progress is not None:
                progress.emit(
                    f"{node_label}: CG iteration {iteration}, solving master LP over "
                    f"{len(node_master.routes)} columns"
                )
            master_lp = node_master.model
            master_time_limit = remaining_time(deadline)
            if master_time_limit is not None:
                if master_time_limit <= 0:
                    raise TimeoutError("Overall time limit reached during node LP.")
                master_lp.Params.TimeLimit = master_time_limit
            optimize_with_heartbeat(
                master_lp,
                progress=progress,
                label=f"{node_label}: master LP iteration {iteration}",
            )
            if master_lp.Status == GRB.TIME_LIMIT:
                raise TimeoutError("Overall time limit reached during node LP.")
            if master_lp.Status != GRB.OPTIMAL:
                raise RuntimeError(
                    f"Column-generation node LP did not solve to optimality (status {master_lp.Status})."
                )

            last_lp_objective = master_lp.ObjVal
            raw_cover_duals = {
                stop_id: node_master.cover_constraints[stop_id].Pi
                for stop_id in sorted(instance.stops)
            }
            raw_type_duals = {
                type_id: node_master.type_constraints[type_id].Pi for type_id in vehicle_types
            }
            if root_full_pricing:
                current_alpha = 1.0
                stabilized_cover_duals = dict(raw_cover_duals)
                stabilized_type_duals = dict(raw_type_duals)
            else:
                current_alpha = adaptive_stabilization_alpha(
                    base_alpha=dual_stabilization_alpha,
                    raw_cover_duals=raw_cover_duals,
                    previous_raw_cover_duals=previous_raw_cover_duals,
                    raw_type_duals=raw_type_duals,
                    previous_raw_type_duals=previous_raw_type_duals,
                    mode=dual_stabilization_mode,
                )
                stabilized_cover_duals = stabilize_dual_vector(
                    raw_duals=raw_cover_duals,
                    previous_duals=stabilized_cover_duals,
                    alpha=current_alpha,
                )
                stabilized_type_duals = stabilize_dual_vector(
                    raw_duals=raw_type_duals,
                    previous_duals=stabilized_type_duals,
                    alpha=current_alpha,
                )
            previous_raw_cover_duals = raw_cover_duals
            previous_raw_type_duals = raw_type_duals

            new_routes: list[RouteColumn] = []
            reactivated_routes = select_inactive_reactivation_routes(
                inactive_routes_by_type=inactive_routes_by_type,
                cover_duals=raw_cover_duals,
                type_duals=raw_type_duals,
                max_per_type=pricing_batch_size,
            )
            for route in reactivated_routes:
                key = route_column_key(route)
                if key in active_route_keys:
                    continue
                inactive_routes_by_type[route.type_id].pop(key, None)
                active_route_keys.add(key)
                new_routes.append(route)
                add_route_column_to_master(node_master, route)
            if reactivated_routes and progress is not None:
                progress.emit(
                    f"{node_label}: CG iteration {iteration} reactivated {len(reactivated_routes)} "
                    "stored columns from global memory"
                )
            if reactivated_routes:
                if progress is not None:
                    progress.emit(
                        f"{node_label}: CG iteration {iteration} added {len(new_routes)} routes; "
                        f"global column pool is now {len(route_pool)}"
                    )
                node_master.model.update()
                emit_checkpoint()
                continue

            type_scan_order = [
                vehicle_type_list[(pricing_start_index + offset) % len(vehicle_type_list)]
                for offset in range(len(vehicle_type_list))
            ]
            found_new_route = False
            improving_type_index: int | None = None
            pricing_round_limit = ROOT_PRICING_EXTRA_SWEEPS if root_full_pricing else 1
            for pricing_round in range(pricing_round_limit):
                routes_by_type: dict[str, list[RouteColumn]] | None = None
                if root_full_pricing:
                    routes_by_type = collect_candidate_routes(
                        vehicle_types_to_price=type_scan_order,
                        cover_duals=stabilized_cover_duals,
                        type_duals=stabilized_type_duals,
                        raw_pricing=False,
                        iteration_number=iteration,
                    )
                routes_before_round = len(new_routes)
                for ordered_index, vehicle_type in enumerate(type_scan_order):
                    routes_before_vehicle = len(new_routes)
                    if root_full_pricing:
                        candidate_routes = routes_by_type[vehicle_type.type_id]
                    else:
                        candidate_routes = collect_candidate_routes(
                            vehicle_types_to_price=[vehicle_type],
                            cover_duals=stabilized_cover_duals,
                            type_duals=stabilized_type_duals,
                            raw_pricing=False,
                            iteration_number=iteration,
                        )[vehicle_type.type_id]
                    found_raw_improving_route = False
                    for route in candidate_routes:
                        if (
                            route_reduced_cost(
                                route,
                                cover_duals=raw_cover_duals,
                                type_dual=raw_type_duals[vehicle_type.type_id],
                            )
                            >= -REDUCED_COST_TOL
                        ):
                            continue
                        key = (route.type_id, route.stop_sequence)
                        if key in route_key_set:
                            continue
                        route_key_set.add(key)
                        route_sequences_by_type[route.type_id].add(route.stop_sequence)
                        active_route_keys.add(key)
                        route_pool.append(route)
                        new_routes.append(route)
                        add_route_column_to_master(node_master, route)
                        found_raw_improving_route = True

                    if not found_raw_improving_route and current_alpha < 1.0 - 1e-12:
                        if progress is not None:
                            progress.emit(
                                f"{node_label}: CG iteration {iteration}, no raw-improving routes "
                                f"found for {vehicle_type.type_id} under stabilized pricing; "
                                "certifying with raw dual pricing"
                            )
                        raw_pricing_results = collect_candidate_routes(
                            vehicle_types_to_price=[vehicle_type],
                            cover_duals=raw_cover_duals,
                            type_duals=raw_type_duals,
                            raw_pricing=True,
                            iteration_number=iteration,
                        )[vehicle_type.type_id]
                        for route in raw_pricing_results:
                            key = (route.type_id, route.stop_sequence)
                            if key in route_key_set:
                                continue
                            route_key_set.add(key)
                            route_sequences_by_type[route.type_id].add(route.stop_sequence)
                            active_route_keys.add(key)
                            route_pool.append(route)
                            new_routes.append(route)
                            add_route_column_to_master(node_master, route)

                    if len(new_routes) > routes_before_vehicle:
                        found_new_route = True
                        improving_type_index = (
                            pricing_start_index + ordered_index
                        ) % len(vehicle_type_list)
                        if not root_full_pricing:
                            break
                if not root_full_pricing:
                    break
                if len(new_routes) == routes_before_round:
                    break
                if len(new_routes) >= ROOT_PRICING_TARGET_NEW_ROUTES_PER_LP:
                    break

            if not new_routes:
                artificial_usage = sum(
                    node_master.artificial_vars[stop_id].X
                    for stop_id in sorted(instance.stops)
                )
                if artificial_usage > INTEGRALITY_TOL:
                    return None
                route_values = [route_var.X for route_var in node_master.route_vars]
                return NodeRelaxationResult(
                    lower_bound=last_lp_objective,
                    routes=list(node_master.routes),
                    route_values=route_values,
                    cg_iterations=iteration,
                )

            if found_new_route and improving_type_index is not None:
                pricing_start_index = (improving_type_index + 1) % len(vehicle_type_list)
            if progress is not None:
                progress.emit(
                    f"{node_label}: CG iteration {iteration} added {len(new_routes)} routes; "
                    f"global column pool is now {len(route_pool)}"
                )
            node_master.model.update()
            emit_checkpoint()

        raise RuntimeError(
            "Maximum column-generation iterations reached before node convergence. "
            "Increase --max-cg-iterations for an exact solve."
        )
    finally:
        if pricing_executor is not None:
            pricing_executor.shutdown()


def solve_restricted_integer_master(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    routes: list[RouteColumn],
    deadline: float | None,
    heuristic_time_limit: float | None,
    progress: ProgressTracker | None = None,
    progress_label: str = "restricted master heuristic",
) -> tuple[float, list[RouteColumn]] | None:
    if not routes:
        return None

    model, data = build_master_model(
        instance=instance,
        vehicle_types=vehicle_types,
        routes=routes,
        relax=False,
        allow_artificial=False,
    )

    effective_time_limit = heuristic_time_limit
    remaining = remaining_time(deadline)
    if remaining is not None:
        if remaining <= 0:
            raise TimeoutError("Overall time limit reached during restricted master heuristic.")
        effective_time_limit = (
            min(heuristic_time_limit, remaining)
            if heuristic_time_limit is not None
            else remaining
        )
    if effective_time_limit is not None:
        model.Params.TimeLimit = effective_time_limit

    optimize_with_heartbeat(model, progress=progress, label=progress_label)
    if model.Status == GRB.TIME_LIMIT and model.SolCount == 0:
        return None
    if model.SolCount == 0:
        return None

    route_vars = data["route_vars"]
    selected_routes = [routes[idx] for idx in range(len(routes)) if route_vars[idx].X > 0.5]
    return model.ObjVal, selected_routes


def is_integral_solution(route_values: list[float]) -> bool:
    return all(abs(value - round(value)) <= INTEGRALITY_TOL for value in route_values)


def compute_arc_values(
    routes: list[RouteColumn],
    route_values: list[float],
) -> dict[tuple[int, int], float]:
    arc_values: dict[tuple[int, int], float] = defaultdict(float)
    for route, value in zip(routes, route_values):
        if value <= INTEGRALITY_TOL:
            continue
        for arc in route.arc_set:
            arc_values[arc] += value
    return arc_values


def compute_assignment_values(
    routes: list[RouteColumn],
    route_values: list[float],
) -> dict[tuple[int, str], float]:
    assignment_values: dict[tuple[int, str], float] = defaultdict(float)
    for route, value in zip(routes, route_values):
        if value <= INTEGRALITY_TOL:
            continue
        for stop_id in route.stop_sequence:
            assignment_values[stop_id, route.type_id] += value
    return assignment_values


def compute_pair_together_values(
    routes: list[RouteColumn],
    route_values: list[float],
) -> dict[tuple[int, int], float]:
    pair_values: dict[tuple[int, int], float] = defaultdict(float)
    for route, value in zip(routes, route_values):
        if value <= INTEGRALITY_TOL or len(route.stop_sequence) < 2:
            continue
        ordered_stops = sorted(route.stop_set)
        for idx, stop_a in enumerate(ordered_stops):
            for stop_b in ordered_stops[idx + 1 :]:
                pair_values[stop_a, stop_b] += value
    return pair_values


def route_conflicts_with_forced_arc(route: RouteColumn, arc: tuple[int, int]) -> bool:
    i, j = arc
    if arc in route.arc_set:
        return False
    if i == DEPOT_NODE:
        return j in route.stop_set
    if j == DEPOT_NODE:
        return i in route.stop_set
    return i in route.stop_set or j in route.stop_set


def route_conflicts_with_forced_assignment(
    route: RouteColumn,
    stop_id: int,
    type_id: str,
) -> bool:
    return stop_id in route.stop_set and route.type_id != type_id


def route_conflicts_with_excluded_assignment(
    route: RouteColumn,
    stop_id: int,
    type_id: str,
) -> bool:
    return stop_id in route.stop_set and route.type_id == type_id


def select_ryan_foster_pair(
    instance: Instance,
    routes: list[RouteColumn],
    route_values: list[float],
    branch_ctx: BranchContext,
) -> tuple[int, int] | None:
    pair_values = compute_pair_together_values(routes, route_values)
    fixed_pairs = set(branch_ctx.together_stop_pairs) | set(branch_ctx.separated_stop_pairs)
    best_pair: tuple[int, int] | None = None
    best_score: tuple[float, float, int, int, float] | None = None

    stop_ids = sorted(instance.stops)
    for idx, stop_a in enumerate(stop_ids):
        for stop_b in stop_ids[idx + 1 :]:
            pair = (stop_a, stop_b)
            if pair in fixed_pairs:
                continue
            together_value = pair_values.get(pair, 0.0)
            if not (INTEGRALITY_TOL < together_value < 1.0 - INTEGRALITY_TOL):
                continue

            together_mass = 0.0
            separate_mass = 0.0
            together_columns = 0
            separate_columns = 0
            for route, route_value in zip(routes, route_values):
                if route_value <= INTEGRALITY_TOL:
                    continue
                has_a = stop_a in route.stop_set
                has_b = stop_b in route.stop_set
                if has_a and has_b:
                    together_mass += route_value
                    together_columns += 1
                elif has_a or has_b:
                    separate_mass += route_value
                    separate_columns += 1

            score = (
                min(together_mass, separate_mass),
                together_mass + separate_mass,
                min(together_columns, separate_columns),
                together_columns + separate_columns,
                -abs(together_value - 0.5),
            )
            if best_score is None or score > best_score:
                best_score = score
                best_pair = pair

    return best_pair


def select_fractional_arc(
    routes: list[RouteColumn],
    route_values: list[float],
    branch_ctx: BranchContext,
) -> tuple[int, int] | None:
    fixed_arcs = set(branch_ctx.forced_arcs) | set(branch_ctx.forbidden_arcs)
    arc_values = compute_arc_values(routes, route_values)
    fractional_arcs = {
        arc: value
        for arc, value in arc_values.items()
        if arc not in fixed_arcs and INTEGRALITY_TOL < value < 1.0 - INTEGRALITY_TOL
    }
    fractional_outgoing_count: dict[int, int] = defaultdict(int)
    fractional_incoming_count: dict[int, int] = defaultdict(int)
    for i, j in fractional_arcs:
        fractional_outgoing_count[i] += 1
        fractional_incoming_count[j] += 1

    best_arc: tuple[int, int] | None = None
    best_score: tuple[float, float, int, int, int, float] | None = None

    for arc, value in fractional_arcs.items():
        forbid_mass = 0.0
        force_mass = 0.0
        forbid_columns = 0
        force_columns = 0
        for route, route_value in zip(routes, route_values):
            if route_value <= INTEGRALITY_TOL:
                continue
            if arc in route.arc_set:
                forbid_mass += route_value
                forbid_columns += 1
            if route_conflicts_with_forced_arc(route, arc):
                force_mass += route_value
                force_columns += 1

        i, j = arc
        ambiguity = fractional_outgoing_count[i] + fractional_incoming_count[j]
        score = (
            min(forbid_mass, force_mass),
            forbid_mass + force_mass,
            min(forbid_columns, force_columns),
            ambiguity,
            0 if DEPOT_NODE in arc else 1,
            -abs(value - 0.5),
        )
        if best_score is None or score > best_score:
            best_score = score
            best_arc = arc

    return best_arc


def select_fractional_assignment(
    instance: Instance,
    vehicle_types: dict[str, VehicleTypeData],
    routes: list[RouteColumn],
    route_values: list[float],
    branch_ctx: BranchContext,
) -> tuple[int, str] | None:
    assignment_values = compute_assignment_values(routes, route_values)
    best_assignment: tuple[int, str] | None = None
    best_score: tuple[float, float, int, int, float] | None = None

    for stop_id in sorted(instance.stops):
        if stop_id in branch_ctx.forced_type_for_stop:
            continue
        for type_id in sorted(vehicle_types):
            if type_id in branch_ctx.excluded_types_by_stop.get(stop_id, set()):
                continue
            value = assignment_values.get((stop_id, type_id), 0.0)
            if not (INTEGRALITY_TOL < value < 1.0 - INTEGRALITY_TOL):
                continue
            exclude_mass = 0.0
            force_mass = 0.0
            exclude_columns = 0
            force_columns = 0
            for route, route_value in zip(routes, route_values):
                if route_value <= INTEGRALITY_TOL:
                    continue
                if route_conflicts_with_excluded_assignment(route, stop_id, type_id):
                    exclude_mass += route_value
                    exclude_columns += 1
                if route_conflicts_with_forced_assignment(route, stop_id, type_id):
                    force_mass += route_value
                    force_columns += 1
            score = (
                min(exclude_mass, force_mass),
                exclude_mass + force_mass,
                min(exclude_columns, force_columns),
                exclude_columns + force_columns,
                -abs(value - 0.5),
            )
            if best_score is None or score > best_score:
                best_score = score
                best_assignment = (stop_id, type_id)

    return best_assignment


def create_model(
    workbook_path: Path,
    max_cg_iterations: int | None = 5000,
    pricing_time_limit: float | None = None,
    pricing_columns_per_type: int = 3,
    pricing_workers: int | None = None,
    dual_stabilization_alpha: float = 0.5,
    dual_stabilization_mode: str = "adaptive",
) -> tuple[gp.Model, dict[str, Any]]:
    instance = load_instance(workbook_path)
    validate_shared_cost_assumptions(instance)
    vehicle_types = build_vehicle_types(instance)
    route_pool = initial_routes(instance, vehicle_types)
    route_key_set = {(route.type_id, route.stop_sequence) for route in route_pool}
    root_state = BranchState(
        forbidden_arcs=frozenset(),
        forced_arcs=frozenset(),
        excluded_assignments=frozenset(),
        forced_assignments=frozenset(),
        same_route_pairs=REQUIRED_SAME_ROUTE_PAIRS,
        different_route_pairs=frozenset(),
    )
    root_result = solve_node_relaxation(
        instance=instance,
        vehicle_types=vehicle_types,
        route_pool=route_pool,
        route_key_set=route_key_set,
        branch_state=root_state,
        max_cg_iterations=max_cg_iterations,
        pricing_time_limit=pricing_time_limit,
        pricing_columns_per_type=pricing_columns_per_type,
        pricing_workers=pricing_workers,
        dual_stabilization_alpha=dual_stabilization_alpha,
        dual_stabilization_mode=dual_stabilization_mode,
        deadline=None,
        progress=None,
        node_label="root",
    )
    if root_result is None:
        raise RuntimeError("The root branch-and-price node is infeasible.")

    model, data = build_master_model(
        instance=instance,
        vehicle_types=vehicle_types,
        routes=root_result.routes,
        relax=False,
        allow_artificial=False,
    )
    data.update(
        {
            "instance": instance,
            "cg_iterations": root_result.cg_iterations,
            "lp_objective": root_result.lower_bound,
            "column_count": len(root_result.routes),
            "pricing_columns_per_type": pricing_columns_per_type,
            "pricing_workers": pricing_workers,
            "dual_stabilization_alpha": dual_stabilization_alpha,
            "dual_stabilization_mode": dual_stabilization_mode,
            "note": (
                "This is the root restricted master over generated columns. "
                "Use --route-based to run the newer exact branch-and-price solver."
            ),
        }
    )
    return model, data


def branch_and_price(
    instance: Instance,
    workbook_path: Path,
    time_limit: float | None,
    mip_gap: float | None,
    max_cg_iterations: int | None,
    pricing_time_limit: float | None,
    pricing_columns_per_type: int,
    pricing_workers: int | None,
    dual_stabilization_alpha: float,
    dual_stabilization_mode: str,
    max_bp_nodes: int | None,
    heuristic_time_limit: float | None,
    progress: ProgressTracker | None = None,
    trace_logger: ObjectiveTraceLogger | None = None,
    checkpoint_manager: CheckpointManager | None = None,
    incumbent_report_writer: IncumbentReportWriter | None = None,
    resume_from: Path | None = None,
) -> BranchAndPriceResult:
    validate_shared_cost_assumptions(instance)
    vehicle_types = build_vehicle_types(instance)
    deadline = perf_counter() + time_limit if time_limit is not None else None
    known_seed_routes = construct_known_feasible_seed_routes(instance, vehicle_types)
    checkpoint_state = (
        load_branch_and_price_checkpoint(
            resume_from,
            workbook_path=workbook_path,
            instance=instance,
            vehicle_types=vehicle_types,
        )
        if resume_from is not None
        else None
    )

    if checkpoint_state is None:
        route_pool = initial_routes(instance, vehicle_types)
        active_nodes: list[tuple[float, int, int, BranchState]] = []
        next_sequence_id = 0

        def allocate_sequence_id() -> int:
            nonlocal next_sequence_id
            sequence_id = next_sequence_id
            next_sequence_id += 1
            return sequence_id

        root_state = BranchState(
            forbidden_arcs=frozenset(),
            forced_arcs=frozenset(),
            excluded_assignments=frozenset(),
            forced_assignments=frozenset(),
            same_route_pairs=REQUIRED_SAME_ROUTE_PAIRS,
            different_route_pairs=frozenset(),
        )
        heapq.heappush(active_nodes, (0.0, 0, allocate_sequence_id(), root_state))
        incumbent_obj: float | None = None
        incumbent_routes: list[RouteColumn] = []
        root_lp_objective: float | None = None
        nodes_processed = 0
        nodes_pruned = 0
        current_node: CurrentNodeCheckpointState | None = None
    else:
        route_pool = list(checkpoint_state.route_pool)
        active_nodes = list(checkpoint_state.active_nodes)
        next_sequence_id = checkpoint_state.next_sequence_id

        def allocate_sequence_id() -> int:
            nonlocal next_sequence_id
            sequence_id = next_sequence_id
            next_sequence_id += 1
            return sequence_id

        incumbent_obj = checkpoint_state.incumbent_obj
        incumbent_routes = list(checkpoint_state.incumbent_routes)
        root_lp_objective = checkpoint_state.root_lp_objective
        nodes_processed = checkpoint_state.nodes_processed
        nodes_pruned = checkpoint_state.nodes_pruned
        current_node = checkpoint_state.current_node
    route_key_set = {(route.type_id, route.stop_sequence) for route in route_pool}
    status = "UNKNOWN"

    def announce(message: str) -> None:
        print(message, flush=True)
        if progress is not None:
            progress.emit(message, force=True)

    def pending_node_count() -> int:
        return len(active_nodes) + (1 if current_node is not None else 0)

    def save_checkpoint(note: str) -> None:
        if checkpoint_manager is None:
            return
        checkpoint_manager.save_branch_and_price_state(
            workbook_path=workbook_path,
            route_pool=route_pool,
            active_nodes=active_nodes,
            next_sequence_id=next_sequence_id,
            incumbent_obj=incumbent_obj,
            incumbent_routes=incumbent_routes,
            root_lp_objective=root_lp_objective,
            nodes_processed=nodes_processed,
            nodes_pruned=nodes_pruned,
            current_node=current_node,
            status=status,
            note=note,
        )

    def maybe_save_incumbent_report(*, force: bool = False) -> None:
        if incumbent_report_writer is None or incumbent_obj is None:
            return
        saved_path = incumbent_report_writer.maybe_save(
            incumbent_obj,
            render_branch_and_price_route_report(
                instance,
                status="RUNNING" if status == "UNKNOWN" else status,
                objective=incumbent_obj,
                best_bound=current_global_bound(),
                gap=relative_gap(incumbent_obj, current_global_bound()),
                selected_routes=incumbent_routes,
                vehicle_types=vehicle_types,
                nodes_processed=nodes_processed,
                nodes_pruned=nodes_pruned,
                root_lp_objective=root_lp_objective,
                global_columns=len(route_pool),
            ),
            force=force,
        )
        if saved_path is not None and progress is not None:
            progress.emit(
                f"branch-and-price: saved incumbent route report to {saved_path}",
                force=True,
            )

    def current_global_bound(active_node_hint: float | None = None) -> float | None:
        if root_lp_objective is None and nodes_processed == 0 and current_node is None:
            return None
        candidate_bounds: list[float] = []
        if active_node_hint is not None:
            candidate_bounds.append(active_node_hint)
        if current_node is not None:
            candidate_bounds.append(current_node.lower_bound_hint)
        if active_nodes:
            candidate_bounds.append(min(bound for bound, *_ in active_nodes))
        if candidate_bounds:
            return min(candidate_bounds)
        if incumbent_obj is not None:
            return incumbent_obj
        return root_lp_objective

    last_logged_incumbent: float | None = None
    last_logged_bound: float | None = None

    def maybe_record_trace(
        event: str,
        *,
        active_node_hint: float | None = None,
        force: bool = False,
        note: str = "",
        node_depth: int | None = None,
    ) -> None:
        nonlocal last_logged_bound, last_logged_incumbent
        if trace_logger is None:
            return
        best_bound = current_global_bound(active_node_hint)
        incumbent_changed = optional_float_changed(last_logged_incumbent, incumbent_obj)
        bound_changed = optional_float_changed(last_logged_bound, best_bound)
        if not force and not incumbent_changed and not bound_changed:
            return
        gap = relative_gap(incumbent_obj, best_bound)
        trace_logger.record(
            event=event,
            status="RUNNING" if status == "UNKNOWN" else status,
            incumbent_objective=incumbent_obj,
            best_bound=best_bound,
            relative_gap=gap,
            nodes_processed=nodes_processed,
            nodes_pruned=nodes_pruned,
            global_columns=len(route_pool),
            queue_size=pending_node_count(),
            node_depth=node_depth,
            note=note,
        )
        last_logged_incumbent = incumbent_obj
        last_logged_bound = best_bound

    if checkpoint_state is None:
        announce(
            "Route-based startup: "
            f"{len(route_pool)} initial columns "
            f"({len(known_seed_routes)} known constructive seed routes)"
        )
        maybe_record_trace(
            "run_start",
            force=True,
            note=f"Initial route pool has {len(route_pool)} columns",
        )
        save_checkpoint(f"Initialized branch-and-price with {len(route_pool)} initial columns")

        bootstrap_added_routes = bootstrap_root_route_pool(
            instance=instance,
            vehicle_types=vehicle_types,
            route_pool=route_pool,
            route_key_set=route_key_set,
            branch_state=root_state,
            deadline=deadline,
            progress=progress,
        )
        if bootstrap_added_routes > 0:
            announce(
                "Route-based startup root bootstrap added "
                f"{bootstrap_added_routes} heuristic columns; "
                f"route pool now has {len(route_pool)} columns"
            )
            maybe_record_trace(
                "bound_update",
                force=True,
                note=(
                    "Root bootstrap added "
                    f"{bootstrap_added_routes} heuristic columns"
                ),
            )
            save_checkpoint(
                "Completed root bootstrap with "
                f"{bootstrap_added_routes} added heuristic columns"
            )

        try:
            initial_heuristic = solve_restricted_integer_master(
                instance=instance,
                vehicle_types=vehicle_types,
                routes=route_pool,
                deadline=deadline,
                heuristic_time_limit=heuristic_time_limit,
                progress=progress,
                progress_label="initial restricted-master heuristic",
            )
        except TimeoutError:
            initial_heuristic = None
        if initial_heuristic is not None:
            incumbent_obj, incumbent_routes = initial_heuristic
            announce(
                "Route-based startup incumbent from seeded restricted master: "
                f"{incumbent_obj:.2f}"
            )
            maybe_record_trace(
                "incumbent_update",
                force=True,
                note="Initial restricted-master incumbent",
            )
            maybe_save_incumbent_report(force=True)
            save_checkpoint("Captured initial restricted-master incumbent")
        else:
            announce("Route-based startup did not find an initial restricted-master incumbent.")
    else:
        announce(
            f"Resuming branch-and-price from {resume_from} with {len(route_pool)} "
            f"global columns and {pending_node_count()} pending nodes"
        )
        maybe_record_trace(
            "run_resume",
            force=True,
            note=f"Resumed from checkpoint {resume_from}",
        )
        if incumbent_routes:
            maybe_save_incumbent_report(force=True)

    while current_node is not None or active_nodes:
        if max_bp_nodes is not None and nodes_processed >= max_bp_nodes:
            status = "NODE_LIMIT"
            break
        if deadline is not None and remaining_time(deadline) <= 0:
            status = "TIME_LIMIT"
            break

        if current_node is None:
            lower_bound_hint, depth, sequence_id, branch_state = heapq.heappop(active_nodes)
            current_node = CurrentNodeCheckpointState(
                lower_bound_hint=lower_bound_hint,
                depth=depth,
                sequence_id=sequence_id,
                branch_state=branch_state,
                relaxation_state=None,
            )
            save_checkpoint(f"Started node at depth {depth}")

        lower_bound_hint = current_node.lower_bound_hint
        depth = current_node.depth
        branch_state = current_node.branch_state
        if incumbent_obj is not None and lower_bound_hint >= incumbent_obj - INTEGRALITY_TOL:
            nodes_pruned += 1
            current_node = None
            maybe_record_trace(
                "bound_update",
                note="Pruned queued node by incumbent cutoff",
                node_depth=depth,
            )
            save_checkpoint("Pruned pending node by incumbent cutoff")
            continue

        if progress is not None:
            incumbent_text = "none" if incumbent_obj is None else f"{incumbent_obj:.2f}"
            progress.emit(
                f"Exploring node depth {depth}, queue {pending_node_count() - 1}, "
                f"best hint {lower_bound_hint:.2f}, incumbent {incumbent_text}"
            )

        def save_node_relaxation_checkpoint(
            node_checkpoint: NodeRelaxationCheckpointState,
        ) -> None:
            if current_node is None:
                return
            current_node.relaxation_state = node_checkpoint
            save_checkpoint(
                f"Saved node depth {current_node.depth} after CG iteration "
                f"{node_checkpoint.iteration}"
            )

        try:
            node_result = solve_node_relaxation(
                instance=instance,
                vehicle_types=vehicle_types,
                route_pool=route_pool,
                route_key_set=route_key_set,
                branch_state=branch_state,
                max_cg_iterations=max_cg_iterations,
                pricing_time_limit=pricing_time_limit,
                pricing_columns_per_type=pricing_columns_per_type,
                pricing_workers=pricing_workers,
                dual_stabilization_alpha=dual_stabilization_alpha,
                dual_stabilization_mode=dual_stabilization_mode,
                deadline=deadline,
                progress=progress,
                node_label=f"node depth {depth}",
                resume_state=current_node.relaxation_state,
                checkpoint_callback=save_node_relaxation_checkpoint,
            )
        except TimeoutError:
            if progress is not None:
                progress.emit("Time limit reached while processing the current node", force=True)
            status = "TIME_LIMIT"
            save_checkpoint(f"Time limit reached while processing node depth {depth}")
            break

        nodes_processed += 1
        if node_result is None:
            current_node = None
            maybe_record_trace(
                "bound_update",
                note="Pruned infeasible node",
                node_depth=depth,
            )
            save_checkpoint(f"Pruned infeasible node at depth {depth}")
            continue
        if root_lp_objective is None:
            root_lp_objective = node_result.lower_bound
            if progress is not None:
                progress.emit(
                    f"Root relaxation closed at {root_lp_objective:.2f}",
                    force=True,
                )
            maybe_record_trace(
                "bound_update",
                active_node_hint=node_result.lower_bound,
                force=True,
                note="Root relaxation closed",
                node_depth=depth,
            )
            save_checkpoint("Closed root relaxation")

        node_lb = node_result.lower_bound
        if incumbent_obj is not None and node_lb >= incumbent_obj - INTEGRALITY_TOL:
            nodes_pruned += 1
            current_node = None
            maybe_record_trace(
                "bound_update",
                note="Pruned processed node by incumbent cutoff",
                node_depth=depth,
            )
            save_checkpoint(f"Pruned processed node at depth {depth} by incumbent cutoff")
            continue

        heuristic_obj: float | None = None
        if nodes_processed == 1 or incumbent_obj is None:
            try:
                heuristic_solution = solve_restricted_integer_master(
                    instance=instance,
                    vehicle_types=vehicle_types,
                    routes=node_result.routes,
                    deadline=deadline,
                    heuristic_time_limit=heuristic_time_limit,
                    progress=progress,
                    progress_label=f"restricted-master heuristic at node depth {depth}",
                )
            except TimeoutError:
                if progress is not None:
                    progress.emit(
                        "Time limit reached while solving the restricted-master heuristic",
                        force=True,
                    )
                status = "TIME_LIMIT"
                save_checkpoint(
                    f"Time limit reached during restricted-master heuristic at depth {depth}"
                )
                break
            if heuristic_solution is not None:
                heuristic_obj, heuristic_routes = heuristic_solution
                if incumbent_obj is None or heuristic_obj < incumbent_obj - INTEGRALITY_TOL:
                    incumbent_obj = heuristic_obj
                    incumbent_routes = heuristic_routes
                    if progress is not None:
                        progress.emit(
                            f"New incumbent from restricted master: {incumbent_obj:.2f}",
                            force=True,
                        )
                    maybe_record_trace(
                        "incumbent_update",
                        active_node_hint=node_lb,
                        force=True,
                        note="Improved incumbent from restricted master",
                        node_depth=depth,
                    )
                    maybe_save_incumbent_report()
                    save_checkpoint(f"Updated incumbent from restricted master at depth {depth}")

        if is_integral_solution(node_result.route_values):
            selected_routes = [
                route
                for route, value in zip(node_result.routes, node_result.route_values)
                if value > 0.5
            ]
            if incumbent_obj is None or node_lb < incumbent_obj - INTEGRALITY_TOL:
                incumbent_obj = node_lb
                incumbent_routes = selected_routes
                if progress is not None:
                    progress.emit(
                        f"Found integral node solution: {incumbent_obj:.2f}",
                        force=True,
                    )
                maybe_record_trace(
                    "incumbent_update",
                    force=True,
                    note="Found integral node solution",
                    node_depth=depth,
                )
                maybe_save_incumbent_report()
                save_checkpoint(f"Updated incumbent from integral node at depth {depth}")
            current_node = None
            save_checkpoint(f"Fathomed integral node at depth {depth}")
            continue

        if heuristic_obj is not None and abs(heuristic_obj - node_lb) <= 1e-5:
            current_node = None
            save_checkpoint(
                f"Fathomed node depth {depth} because heuristic matched the LP bound"
            )
            continue

        branch_ctx = build_branch_context(branch_state, instance, vehicle_types)
        if branch_ctx is None:
            current_node = None
            save_checkpoint(f"Skipped inconsistent branch context at depth {depth}")
            continue

        child_states: list[BranchState] = []
        branch_pair = select_ryan_foster_pair(
            instance=instance,
            routes=node_result.routes,
            route_values=node_result.route_values,
            branch_ctx=branch_ctx,
        )
        if branch_pair is not None:
            if progress is not None:
                progress.emit(
                    f"Branching on Ryan-Foster pair ({branch_pair[0]}, {branch_pair[1]}) "
                    f"at node depth {depth}"
                )
            child_states = [
                BranchState(
                    forbidden_arcs=branch_state.forbidden_arcs,
                    forced_arcs=branch_state.forced_arcs,
                    excluded_assignments=branch_state.excluded_assignments,
                    forced_assignments=branch_state.forced_assignments,
                    same_route_pairs=branch_state.same_route_pairs,
                    different_route_pairs=frozenset(
                        set(branch_state.different_route_pairs) | {branch_pair}
                    ),
                ),
                BranchState(
                    forbidden_arcs=branch_state.forbidden_arcs,
                    forced_arcs=branch_state.forced_arcs,
                    excluded_assignments=branch_state.excluded_assignments,
                    forced_assignments=branch_state.forced_assignments,
                    same_route_pairs=frozenset(set(branch_state.same_route_pairs) | {branch_pair}),
                    different_route_pairs=branch_state.different_route_pairs,
                ),
            ]
        else:
            branch_arc = select_fractional_arc(
            routes=node_result.routes,
            route_values=node_result.route_values,
            branch_ctx=branch_ctx,
            )
            if branch_arc is not None:
                if progress is not None:
                    progress.emit(
                        f"Branching on arc {branch_arc[0]} -> {branch_arc[1]} at node depth {depth}"
                    )
                child_states = [
                    BranchState(
                        forbidden_arcs=frozenset(set(branch_state.forbidden_arcs) | {branch_arc}),
                        forced_arcs=branch_state.forced_arcs,
                        excluded_assignments=branch_state.excluded_assignments,
                        forced_assignments=branch_state.forced_assignments,
                        same_route_pairs=branch_state.same_route_pairs,
                        different_route_pairs=branch_state.different_route_pairs,
                    ),
                    BranchState(
                        forbidden_arcs=branch_state.forbidden_arcs,
                        forced_arcs=frozenset(set(branch_state.forced_arcs) | {branch_arc}),
                        excluded_assignments=branch_state.excluded_assignments,
                        forced_assignments=branch_state.forced_assignments,
                        same_route_pairs=branch_state.same_route_pairs,
                        different_route_pairs=branch_state.different_route_pairs,
                    ),
                ]
            else:
                branch_assignment = select_fractional_assignment(
                    instance=instance,
                    vehicle_types=vehicle_types,
                    routes=node_result.routes,
                    route_values=node_result.route_values,
                    branch_ctx=branch_ctx,
                )
                if branch_assignment is not None:
                    stop_id, type_id = branch_assignment
                    if progress is not None:
                        progress.emit(
                            f"Branching on stop {stop_id} type assignment {type_id} at node depth {depth}"
                        )
                    child_states = [
                        BranchState(
                            forbidden_arcs=branch_state.forbidden_arcs,
                            forced_arcs=branch_state.forced_arcs,
                            excluded_assignments=frozenset(
                                set(branch_state.excluded_assignments) | {(stop_id, type_id)}
                            ),
                            forced_assignments=branch_state.forced_assignments,
                            same_route_pairs=branch_state.same_route_pairs,
                            different_route_pairs=branch_state.different_route_pairs,
                        ),
                        BranchState(
                            forbidden_arcs=branch_state.forbidden_arcs,
                            forced_arcs=branch_state.forced_arcs,
                            excluded_assignments=branch_state.excluded_assignments,
                            forced_assignments=frozenset(
                                set(branch_state.forced_assignments) | {(stop_id, type_id)}
                            ),
                            same_route_pairs=branch_state.same_route_pairs,
                            different_route_pairs=branch_state.different_route_pairs,
                        ),
                    ]

        if not child_states:
            raise RuntimeError(
                "Branch-and-price could not find a valid branching object on a fractional node."
            )

        for child_state in child_states:
            if build_branch_context(child_state, instance, vehicle_types) is None:
                continue
            heapq.heappush(
                active_nodes,
                (node_lb, depth + 1, allocate_sequence_id(), child_state),
            )

        current_node = None
        maybe_record_trace(
            "bound_update",
            note="Updated global bound after branching",
            node_depth=depth,
        )
        save_checkpoint(f"Branched node at depth {depth}")

        best_bound = min(bound for bound, *_ in active_nodes) if active_nodes else node_lb
        if incumbent_obj is not None and mip_gap is not None:
            gap = relative_gap(incumbent_obj, best_bound)
            if gap is not None and gap <= mip_gap + 1e-12:
                if progress is not None:
                    progress.emit(
                        f"Gap target reached with incumbent {incumbent_obj:.2f} and "
                        f"bound {best_bound:.2f}",
                        force=True,
                    )
                status = "GAP_LIMIT"
                maybe_record_trace(
                    "gap_limit_reached",
                    force=True,
                    note="Reached requested relative gap target",
                    node_depth=depth,
                )
                save_checkpoint("Reached requested relative gap target")
                break

    if status == "UNKNOWN":
        if current_node is not None or active_nodes:
            status = "TIME_LIMIT" if deadline is not None and remaining_time(deadline) <= 0 else "NODE_LIMIT"
        elif incumbent_obj is None:
            status = "INFEASIBLE"
        else:
            status = "OPTIMAL"

    pending_bounds: list[float] = []
    if current_node is not None:
        pending_bounds.append(current_node.lower_bound_hint)
    if active_nodes:
        pending_bounds.append(min(bound for bound, *_ in active_nodes))

    if root_lp_objective is None and nodes_processed == 0 and not pending_bounds:
        best_bound = None
    elif pending_bounds:
        best_bound = min(pending_bounds)
    elif incumbent_obj is not None:
        best_bound = incumbent_obj
    else:
        best_bound = root_lp_objective

    if trace_logger is not None:
        trace_logger.record(
            event="solve_end",
            status=status,
            incumbent_objective=incumbent_obj,
            best_bound=best_bound,
            relative_gap=relative_gap(incumbent_obj, best_bound),
            nodes_processed=nodes_processed,
            nodes_pruned=nodes_pruned,
            global_columns=len(route_pool),
            queue_size=pending_node_count(),
            note="Branch-and-price finished",
        )

    if incumbent_routes:
        maybe_save_incumbent_report(force=True)

    save_checkpoint(f"Branch-and-price exited with status {status}")

    return BranchAndPriceResult(
        status=status,
        objective=incumbent_obj,
        best_bound=best_bound,
        gap=relative_gap(incumbent_obj, best_bound),
        selected_routes=incumbent_routes,
        vehicle_types=vehicle_types,
        nodes_processed=nodes_processed,
        nodes_pruned=nodes_pruned,
        global_columns=len(route_pool),
        root_lp_objective=root_lp_objective,
    )


def solve_model(
    workbook_path: Path,
    time_limit: float | None,
    mip_gap: float | None,
    max_cg_iterations: int | None,
    pricing_time_limit: float | None,
    pricing_columns_per_type: int,
    pricing_workers: int | None,
    dual_stabilization_alpha: float,
    dual_stabilization_mode: str,
    max_bp_nodes: int | None,
    heuristic_time_limit: float | None = HEURISTIC_MASTER_TIME_LIMIT,
    progress: ProgressTracker | None = None,
    trace_logger: ObjectiveTraceLogger | None = None,
    checkpoint_manager: CheckpointManager | None = None,
    incumbent_report_writer: IncumbentReportWriter | None = None,
    resume_from: Path | None = None,
) -> BranchAndPriceResult:
    instance = load_instance(workbook_path)
    print_stop_reference(instance)
    result = branch_and_price(
        instance=instance,
        workbook_path=workbook_path,
        time_limit=time_limit,
        mip_gap=mip_gap,
        max_cg_iterations=max_cg_iterations,
        pricing_time_limit=pricing_time_limit,
        pricing_columns_per_type=pricing_columns_per_type,
        pricing_workers=pricing_workers,
        dual_stabilization_alpha=dual_stabilization_alpha,
        dual_stabilization_mode=dual_stabilization_mode,
        max_bp_nodes=max_bp_nodes,
        heuristic_time_limit=heuristic_time_limit,
        progress=progress,
        trace_logger=trace_logger,
        checkpoint_manager=checkpoint_manager,
        incumbent_report_writer=incumbent_report_writer,
        resume_from=resume_from,
    )

    print(f"Status: {result.status}")
    print("Formulation: Branch-and-Price Set Partitioning")
    print(f"Columns generated: {result.global_columns}")
    print(f"Nodes processed: {result.nodes_processed}")
    print(f"Nodes pruned: {result.nodes_pruned}")
    if result.root_lp_objective is not None:
        print(f"Root LP objective: {result.root_lp_objective:.2f}")
    if result.best_bound is not None:
        print(f"Best bound: {result.best_bound:.2f}")
    if result.objective is not None:
        print(f"Objective value: {result.objective:.2f}")
    if result.gap is not None:
        print(f"Relative gap: {100.0 * result.gap:.2f}%")
    print()

    if not result.selected_routes:
        print("No feasible incumbent route set found.")
        return result

    usage_counter: dict[str, int] = defaultdict(int)
    for route in sorted(
        result.selected_routes,
        key=lambda candidate: (candidate.type_id, candidate.return_min, candidate.stop_sequence),
    ):
        vehicle_type = result.vehicle_types[route.type_id]
        usage_counter[route.type_id] += 1
        vehicle_position = usage_counter[route.type_id] - 1
        vehicle_id = vehicle_type.vehicle_ids[vehicle_position]

        print(f"Vehicle {vehicle_id} ({vehicle_type.vehicle_type})")
        print(
            "  "
            f"Depart {minutes_to_clock(vehicle_start_limit(instance, vehicle_type))}, "
            f"Return {minutes_to_clock(route.return_min)}, "
            f"Load {route.load_kg:.0f} / {vehicle_type.capacity_kg:.0f} kg"
        )
        print(
            "  "
            f"Active {route.active_min:.1f} min, Distance {route.distance_mi:.2f} mi, "
            f"Route cost {route.cost:.2f}"
        )
        if route.lunch_break_start_min is not None:
            lunch_end = route.lunch_break_start_min + MIDDAY_LUNCH_BREAK_RULE.duration_min
            print(
                "  "
                f"Lunch break {minutes_to_clock(route.lunch_break_start_min)}-"
                f"{minutes_to_clock(lunch_end)}"
            )
        route_with_depot = (DEPOT_NODE,) + route.stop_sequence + (DEPOT_NODE,)
        print(f"  Route IDs: {format_route_stop_ids(route_with_depot)}")
        print(
            "  Route detail: "
            f"{format_route_stop_details(instance, route_with_depot)}"
        )
        for stop_id in route.stop_sequence:
            start = route.service_start_min[stop_id]
            early = max(instance.stops[stop_id].earliest_min - start, 0.0)
            late = route.late_min[stop_id]
            print(
                "    "
                f"{format_stop_detail_label(instance.stops[stop_id])}: "
                f"start {minutes_to_clock(start)}, "
                f"early {early:.1f} min, late {late:.1f} min"
            )
        print()

    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Solve the workbook-driven VRP. By default this uses the older compact arc-flow "
            "MIP with vehicle symmetry breaking; pass --route-based to run the newer "
            "route-based branch-and-price solver."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="Path to the Excel workbook.",
    )
    parser.add_argument(
        "--time-limit",
        type=float,
        default=None,
        help="Optional overall solve time limit in seconds.",
    )
    parser.add_argument(
        "--mip-gap",
        type=float,
        default=None,
        help="Optional relative MIP/optimality gap target for the active solver.",
    )
    parser.add_argument(
        "--mip-profile",
        type=str,
        default=None,
        help=(
            "Optional named Gurobi tuning profile for the default arc-flow MIP. "
            "Built-in profiles are stored in gurobi_mip_profiles.json: "
            "hybrid, two-minute, max-accuracy. Only used without --route-based."
        ),
    )
    parser.add_argument(
        "--mip-persist-dir",
        type=Path,
        default=None,
        help=(
            "Autosave directory for the default arc-flow MIP. "
            "The solver writes intermediate incumbent .sol files into timestamped runs/ "
            "subdirectories there during the run, "
            "keeps a reusable resume.sol, and automatically reloads that start on later runs. "
            "Defaults to .mip_persist/<workbook-stem> beside the workbook. "
            "Only used without --route-based."
        ),
    )
    parser.add_argument(
        "--route-based",
        "--branch-and-price",
        dest="route_based",
        action="store_true",
        help="Use the newer route-based branch-and-price solver instead of the default arc-flow MIP.",
    )
    parser.add_argument(
        "--max-cg-iterations",
        type=parse_optional_int_limit,
        default=5000,
        help=(
            "Maximum column-generation iterations allowed at each tree node. "
            "Use 0, 'none', or 'unlimited' for no cap. Only used with --route-based."
        ),
    )
    parser.add_argument(
        "--pricing-time-limit",
        type=float,
        default=None,
        help=(
            "Optional per-pricing time limit in seconds. "
            "If a pricing solve hits this limit, the exact solve aborts. "
            "Only used with --route-based."
        ),
    )
    parser.add_argument(
        "--pricing-columns-per-type",
        type=parse_positive_int,
        default=3,
        help=(
            "Maximum number of distinct improving routes to add per vehicle type in each "
            "column-generation iteration. Only used with --route-based."
        ),
    )
    parser.add_argument(
        "--pricing-workers",
        type=parse_worker_count,
        default=None,
        help=(
            "Number of worker processes for pricing across vehicle types. "
            "Default: auto-select up to the number of vehicle types. Use 1 to disable "
            "parallel pricing. You can also pass 'auto'. Only used with --route-based."
        ),
    )
    parser.add_argument(
        "--dual-stabilization-alpha",
        type=parse_closed_unit_interval,
        default=0.5,
        help=(
            "Base exponential smoothing weight for pricing duals. "
            "In adaptive mode this is the center value; use 1.0 for raw duals. "
            "Only used with --route-based."
        ),
    )
    parser.add_argument(
        "--dual-stabilization-mode",
        choices=("adaptive", "fixed"),
        default="adaptive",
        help=(
            "Use a fixed stabilization alpha every iteration, or adapt it based on "
            "how much the raw duals move between CG iterations. Only used with --route-based."
        ),
    )
    parser.add_argument(
        "--max-bp-nodes",
        type=int,
        default=None,
        help="Optional cap on processed branch-and-price nodes. Only used with --route-based.",
    )
    parser.add_argument(
        "--unlimited",
        action="store_true",
        help=(
            "Run with no user-facing time, gap, node, pricing, or column-generation limits. "
            "In default arc-flow mode this removes time and gap limits; in route-based mode "
            "this also removes the restricted-master heuristic time cap."
        ),
    )
    parser.add_argument(
        "--progress",
        action="store_true",
        help="Show throttled progress updates during long solves.",
    )
    parser.add_argument(
        "--progress-interval",
        type=float,
        default=30.0,
        help="Minimum seconds between non-forced progress updates.",
    )
    parser.add_argument(
        "--objective-trace",
        type=Path,
        default=None,
        help=(
            "CSV path for incumbent and bound snapshots over time. "
            "Default: auto-generate a timestamped file beside the workbook."
        ),
    )
    parser.add_argument(
        "--no-objective-trace",
        action="store_true",
        help="Disable writing the objective trace CSV.",
    )
    parser.add_argument(
        "--incumbent-report-path",
        type=Path,
        default=None,
        help=(
            "Optional text file path for the latest human-readable incumbent route report. "
            "The solver refreshes this file whenever the incumbent objective improves by a "
            "meaningful amount. Default: auto-generate beside the workbook."
        ),
    )
    parser.add_argument(
        "--incumbent-report-min-improvement",
        type=float,
        default=DEFAULT_INCUMBENT_REPORT_MIN_IMPROVEMENT,
        help=(
            "Minimum objective decrease required before rewriting the incumbent route report. "
            "Default: 1.0."
        ),
    )
    parser.add_argument(
        "--checkpoint-path",
        type=Path,
        default=None,
        help=(
            "Optional JSON checkpoint path. When set, save resumable branch-and-price "
            "state after each completed CG iteration and major tree updates. "
            "Only used with --route-based."
        ),
    )
    parser.add_argument(
        "--resume-from",
        type=Path,
        default=None,
        help=(
            "Resume branch-and-price from a previously saved checkpoint JSON file. "
            "If --checkpoint-path is omitted, the resumed run keeps writing to this file. "
            "Only used with --route-based."
        ),
    )
    return parser.parse_args()


def main() -> None:
    global _INTERRUPT_SIGNAL_NAME
    _INTERRUPT_SIGNAL_NAME = None
    _LATEST_INTERRUPT_ARTIFACTS.clear()
    args = parse_args()
    route_based_enabled = args.route_based
    selected_mip_profile: ArcFlowMIPProfile | None = None
    route_based_only_overrides: list[str] = []
    if not route_based_enabled:
        if args.max_cg_iterations != 5000:
            route_based_only_overrides.append("--max-cg-iterations")
        if args.pricing_time_limit is not None:
            route_based_only_overrides.append("--pricing-time-limit")
        if args.pricing_columns_per_type != 3:
            route_based_only_overrides.append("--pricing-columns-per-type")
        if args.pricing_workers is not None:
            route_based_only_overrides.append("--pricing-workers")
        if abs(args.dual_stabilization_alpha - 0.5) > 1e-12:
            route_based_only_overrides.append("--dual-stabilization-alpha")
        if args.dual_stabilization_mode != "adaptive":
            route_based_only_overrides.append("--dual-stabilization-mode")
        if args.max_bp_nodes is not None:
            route_based_only_overrides.append("--max-bp-nodes")
        if args.checkpoint_path is not None:
            route_based_only_overrides.append("--checkpoint-path")
        if args.resume_from is not None:
            route_based_only_overrides.append("--resume-from")
    if route_based_only_overrides:
        override_list = ", ".join(route_based_only_overrides)
        raise SystemExit(f"{override_list} requires --route-based.")
    if route_based_enabled and args.mip_persist_dir is not None:
        raise SystemExit("--mip-persist-dir is only supported for the default arc-flow MIP path.")
    if route_based_enabled and args.mip_profile is not None:
        raise SystemExit("--mip-profile is only supported for the default arc-flow MIP path.")
    if args.mip_profile is not None:
        try:
            mip_profiles = load_arc_flow_mip_profiles()
        except ValueError as exc:
            raise SystemExit(str(exc)) from exc
        selected_mip_profile = mip_profiles.get(args.mip_profile)
        if selected_mip_profile is None:
            available_profiles = ", ".join(sorted(mip_profiles))
            raise SystemExit(
                f"Unknown arc-flow MIP profile '{args.mip_profile}'. "
                f"Available profiles: {available_profiles}."
            )
    heuristic_time_limit = HEURISTIC_MASTER_TIME_LIMIT
    if args.unlimited:
        args.time_limit = None
        args.mip_gap = None
        if route_based_enabled:
            args.max_cg_iterations = None
            args.pricing_time_limit = None
            args.max_bp_nodes = None
            heuristic_time_limit = None
    progress = ProgressTracker(
        enabled=args.progress,
        interval_seconds=args.progress_interval,
    )
    if args.incumbent_report_min_improvement < 0:
        raise SystemExit("--incumbent-report-min-improvement must be non-negative.")
    objective_trace_path = (
        None
        if args.no_objective_trace
        else (args.objective_trace or default_objective_trace_path(args.workbook))
    )
    incumbent_report_path = (
        args.incumbent_report_path or default_incumbent_report_path(args.workbook)
    )
    checkpoint_path = (args.checkpoint_path or args.resume_from) if route_based_enabled else None
    checkpoint_manager = CheckpointManager(checkpoint_path)
    mip_persistence_manager = (
        None
        if route_based_enabled
        else MIPPersistenceManager(args.mip_persist_dir or default_mip_persist_dir(args.workbook))
    )
    trace_logger = ObjectiveTraceLogger(objective_trace_path, args.workbook)
    incumbent_report_writer = IncumbentReportWriter(
        incumbent_report_path,
        min_improvement=args.incumbent_report_min_improvement,
    )
    try:
        configured_wls_license = resolve_gurobi_wls_license()
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc
    previous_handlers = install_interrupt_handlers()

    try:
        if trace_logger.path is not None:
            record_interrupt_artifact("objective_trace", trace_logger.path)
            print(f"Objective trace: {trace_logger.path}")
            trace_logger.record(
                event="run_start",
                status="RUNNING",
                note=f"Workbook: {args.workbook}",
            )
        if checkpoint_manager.path is not None:
            record_interrupt_artifact("checkpoint", checkpoint_manager.path)
            print(f"Checkpoint path: {checkpoint_manager.path}")
        if incumbent_report_writer.path is not None:
            print(f"Incumbent report: {incumbent_report_writer.path}")
        if configured_wls_license is not None:
            print("Using Gurobi WLS credentials from .env, environment variables, or fallback script settings.")
        print("Active business rules:")
        for description in active_business_rule_descriptions():
            print(f"  - {description}")
        if route_based_enabled:
            solve_model(
                workbook_path=args.workbook,
                time_limit=args.time_limit,
                mip_gap=args.mip_gap,
                max_cg_iterations=args.max_cg_iterations,
                pricing_time_limit=args.pricing_time_limit,
                pricing_columns_per_type=args.pricing_columns_per_type,
                pricing_workers=args.pricing_workers,
                dual_stabilization_alpha=args.dual_stabilization_alpha,
                dual_stabilization_mode=args.dual_stabilization_mode,
                max_bp_nodes=args.max_bp_nodes,
                heuristic_time_limit=heuristic_time_limit,
                progress=progress,
                trace_logger=trace_logger,
                checkpoint_manager=checkpoint_manager,
                incumbent_report_writer=incumbent_report_writer,
                resume_from=args.resume_from,
            )
        else:
            solve_arc_flow_model(
                workbook_path=args.workbook,
                time_limit=args.time_limit,
                mip_gap=args.mip_gap,
                mip_profile=selected_mip_profile,
                skip_mip_profile_params=(
                    frozenset({"TimeLimit", "MIPGap"}) if args.unlimited else frozenset()
                ),
                progress=progress,
                trace_logger=trace_logger,
                persistence_manager=mip_persistence_manager,
                incumbent_report_writer=incumbent_report_writer,
            )
    except KeyboardInterrupt:
        interrupt_note = "Interrupted by user"
        if _INTERRUPT_SIGNAL_NAME is not None:
            interrupt_note = f"Interrupted by {_INTERRUPT_SIGNAL_NAME}"
        if trace_logger.path is not None:
            trace_logger.record(
                event="interrupted",
                status="INTERRUPTED",
                note=interrupt_note,
            )
            print(f"\n{interrupt_note}. Partial objective trace saved to {trace_logger.path}")
        if checkpoint_manager.path is not None:
            print(f"{interrupt_note}. Latest checkpoint retained at {checkpoint_manager.path}")
        if incumbent_report_writer.path is not None and incumbent_report_writer.path.is_file():
            print(
                f"{interrupt_note}. Latest incumbent route report is at "
                f"{incumbent_report_writer.path}"
            )
        saved_interrupt_report_path = write_interrupt_artifact_report(
            default_interrupt_artifact_report_path(args.workbook),
            workbook_path=args.workbook,
            interrupt_note=interrupt_note,
        )
        print(
            f"{interrupt_note}. Interrupt artifact report saved to "
            f"{saved_interrupt_report_path}"
        )
        raise SystemExit(130) from None
    finally:
        dispose_gurobi_env()
        restore_interrupt_handlers(previous_handlers)
        trace_logger.close()


if __name__ == "__main__":
    main()
